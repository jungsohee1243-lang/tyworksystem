import io
import os
import re
import string
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import unicodedata
from typing import Optional, Tuple, List, Dict, Any

import pandas as pd
import requests
import pdfplumber
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


COLUMNS = [
    "비엘", "쉬퍼", "쉬퍼주소", "컨사이니", "컨사이니 사업자번호", "컨사이니 주소",
    "선명", "항차", "출발지", "도착지", "품명", "마크", "수량", "중량", "CBM",
    "원본파일명", "확인필요"
]


def group_words_to_lines(words, y_tol=3):
    words = sorted(words, key=lambda w: (w["top"], w["x0"]))
    lines = []
    for w in words:
        cy = (w["top"] + w["bottom"]) / 2
        for line in lines:
            if abs(line["cy"] - cy) <= y_tol:
                line["words"].append(w)
                line["cy"] = (line["cy"] * line["n"] + cy) / (line["n"] + 1)
                line["n"] += 1
                break
        else:
            lines.append({"cy": cy, "n": 1, "words": [w]})

    result = []
    for line in sorted(lines, key=lambda x: x["cy"]):
        ws = sorted(line["words"], key=lambda w: w["x0"])
        result.append({
            "top": min(w["top"] for w in ws),
            "bottom": max(w["bottom"] for w in ws),
            "x0": min(w["x0"] for w in ws),
            "x1": max(w["x1"] for w in ws),
            "text": " ".join(w["text"] for w in ws),
            "words": ws,
        })
    return result


def text_from_words(words, y_tol=3):
    return "\n".join(line["text"] for line in group_words_to_lines(words, y_tol=y_tol)).strip()


def words_in_region(page, x0, top, x1, bottom, *, mode="inside"):
    words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
    selected = []
    for w in words:
        if w["top"] < top or w["bottom"] > bottom:
            continue
        if mode == "start":
            ok = x0 <= w["x0"] < x1
        else:
            ok = w["x0"] >= x0 and w["x1"] <= x1
        if ok:
            selected.append(w)
    return selected


def text_in_region(page, x0, top, x1, bottom, *, mode="inside"):
    return text_from_words(words_in_region(page, x0, top, x1, bottom, mode=mode))


def extract_business_no(text):
    """컨사이니명 안의 000-00-00000 형식 사업자번호만 추출합니다."""
    m = re.search(r"(\d{3}-\d{2}-\d{5})", text or "")
    return m.group(1) if m else ""


def clean_company_with_paren(text):
    """컨사이니명 정리.
    - 000-00-00000 형식 사업자번호만 제거
    - 사업자번호 제거 후 남는 빈 괄호 (), （）만 제거
    - 일반 괄호 문구는 임의로 삭제하지 않음
    예: ZZZIP GUESTHOUSE（105-20-88541） -> ZZZIP GUESTHOUSE
        GOGOSS(452-64-00260) -> GOGOSS
        HOMLUX CO., LTD. 563-87-03514 -> HOMLUX CO., LTD.
    """
    text = text or ""
    text = re.sub(r"[\(（]?\s*\b\d{3}-\d{2}-\d{5}\b\s*[\)）]?", " ", text)
    text = re.sub(r"[\(（]\s*[\)）]", " ", text)
    text = re.sub(r"\s+", " ", text).strip(" .,-")
    return text


def clean_consignee_address_and_bizno(address_text, existing_bizno=""):
    """컨사이니 주소 안에 000-00-00000 형식 사업자번호가 있으면 사업자번호 칸으로 분리합니다.
    전화번호(예: 010-3970-7762)는 000-00-00000 형식이 아니므로 주소에 그대로 유지합니다.
    """
    address_text = address_text or ""
    biz_no = existing_bizno or ""

    if not biz_no:
        found = extract_business_no(address_text)
        if found:
            biz_no = found

    if biz_no:
        # 사업자번호만 주소에서 제거. 일반 전화번호는 제거하지 않음.
        address_text = re.sub(rf"[\(（]?\s*{re.escape(biz_no)}\s*[\)）]?", " ", address_text)

    # 사업자번호 제거 후 생긴 빈 괄호만 제거
    address_text = re.sub(r"[\(（]\s*[\)）]", " ", address_text)
    address_text = re.sub(r"[ \t]+", " ", address_text)
    address_text = "\n".join(line.strip(" .,-") for line in address_text.splitlines() if line.strip(" .,-"))
    return address_text.strip(), biz_no


def clean_mark(text):
    lines = []
    for ln in (text or "").splitlines():
        ln = ln.strip()
        if not ln:
            continue
        if re.search(r"marks\s*&?\s*nos|container\s*seal", ln, re.I):
            continue
        lines.append(ln)
    return "\n".join(lines).strip()


def clean_description(text):
    stop_patterns = [
        r"^FREIGHT\b",
        r"^SHIPPED\s+ON\s+BOARD\b",
        r"^Above\s+Particulars\b",
        r"^\d{4}[-./]\d{1,2}[-./]\d{1,2}$",
    ]
    lines = []
    seen_said_to_contain = False

    for ln in (text or "").splitlines():
        ln = re.sub(r"\b\d+(?:\.\d+)?\s*KGS\b", "", ln, flags=re.I)
        ln = re.sub(r"\b\d+(?:\.\d+)?\s*CBM\b", "", ln, flags=re.I)
        ln = ln.strip()
        if not ln:
            continue
        if any(re.search(p, ln, re.I) for p in stop_patterns):
            break

        # 품명은 SAID TO CONTAIN 아래부터 시작. 안내 문구 자체는 제외.
        if re.search(r"SHIPPER['’]?S\s+LOAD\s+COUNT", ln, re.I):
            continue
        if re.search(r"SAID\s+TO\s+CONTAIN", ln, re.I):
            seen_said_to_contain = True
            # 같은 줄 뒤에 품명이 붙는 예외가 있으면 뒤쪽만 살림
            tail = re.split(r"SAID\s+TO\s+CONTAIN", ln, flags=re.I)[-1].strip(" :-")
            if tail:
                lines.append(tail)
            continue

        lines.append(ln)

    return "\n".join(lines).strip()


def clean_description_words(words):
    clean = []
    for w in words:
        t = w["text"].strip()
        if re.fullmatch(r"\d+(?:\.\d+)?\s*KGS", t, re.I):
            continue
        if re.fullmatch(r"\d+(?:\.\d+)?\s*CBM", t, re.I):
            continue
        if re.fullmatch(r"\d+\s*(?:PKGS?|CTNS?|CARTONS?|PCS)", t, re.I):
            continue
        clean.append(w)
    return clean_description(text_from_words(clean))


def safe_search(pattern, text, group=1, flags=re.I):
    m = re.search(pattern, text or "", flags)
    return m.group(group).strip() if m else ""


def normalize_port_code(text):
    """Convert common port text to requested customs-style code."""
    raw = (text or "").strip()
    key = re.sub(r"\s+", " ", raw.upper().replace("，", ","))
    key = key.replace(" ", "")
    mapping = {
        "SHIDAO": "CNSHD",
        "SHIDAOCHINA": "CNSHD",
        "SHIDAO,CHINA": "CNSHD",
        "YANTAICHINA": "CNYNT",
        "YANTAI,CHINA": "CNYNT",
        "WEIHAICHINA": "CNWEI",
        "WEIHAI,CHINA": "CNWEI",
        "GUNSAN,KOREA": "KRKUV",
        "GUNSANKOREA": "KRKUV",
        "INCHEON,KOREA": "KRINC",
        "INCHEONKOREA": "KRINC",
        "INCHON": "KRINC",
        "INCHON,KOREA": "KRINC",
        "INCHONKOREA": "KRINC",
    }
    return mapping.get(key, raw)



def split_vessel_voyage_text(text):
    """선명/항차 보정.
    - HANSUNG으로 시작하는 경우 HANSUNG INCHEON까지 선명으로 고정
    - 뒤의 숫자 3~4자리(+E 선택)를 항차로 분리
    """
    raw = re.sub(r"\s+", " ", text or " ").strip()
    m = re.search(r"\bHANSUNG\s+INCHEON\s+(\d{3,4}E?)\b", raw, re.I)
    if m:
        voyage = m.group(1).upper()
        if re.fullmatch(r"\d{3,4}", voyage):
            voyage = f"{voyage}E"
        return "HANSUNG INCHEON", voyage

    m = re.search(r"\b([A-Z][A-Z0-9]*(?:\s+[A-Z0-9]+)*?)\s+(\d{3,4}E?)\b", raw, re.I)
    if m:
        vessel = re.sub(r"\s+", " ", m.group(1)).strip()
        voyage = m.group(2).upper()
        if re.fullmatch(r"\d{3,4}", voyage):
            voyage = f"{voyage}E"
        return vessel, voyage
    return "", ""

def extract_qty_from_mark(mark_text):
    """Fallback quantity recognition from mark ranges, e.g. C/T:1-71, FR-xxx-001 - FR-xxx-037, LJ1-63."""
    txt = (mark_text or "").replace("\n", " ")
    # C/T:1-71, CT: 1 ~ 71
    m = re.search(r"C\s*/?\s*T\s*[:：]?\s*(\d+)\s*[-~]\s*(\d+)", txt, re.I)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        return str(b - a + 1) if b >= a else str(b)
    # FR-1519594-001 - FR-1519594-037 / GR-...001 - ...005
    nums = re.findall(r"(?:^|[^A-Z0-9])(\d{1,4})(?=\s*(?:$|[^A-Z0-9]))", txt, re.I)
    if len(nums) >= 2:
        # use the last two small sequence numbers if it looks like a range
        a, b = int(nums[-2]), int(nums[-1])
        if 0 <= a <= b <= 9999:
            return str(b - a + 1)
    # LJ1-63, JZF9-1-9: use last number as count when no explicit range exists
    tail = re.search(r"[-\s](\d{1,4})\s*$", txt)
    if tail:
        return str(int(tail.group(1)))
    return ""






def extract_wooyoung_pdf(page, filename):
    """WOOYOUNG / INUF 텍스트형 BL 양식 전용 추출.
    해당 양식은 칸이 명확하므로 일반 BL 로직을 건드리지 않고 좌측/우측 구역 기준으로만 보정합니다.
    """
    warnings = []
    w, h = page.width, page.height
    text = page.extract_text() or ""
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    def region_lines(x0, top, x1, bottom):
        txt = text_in_region(page, x0, top, x1, bottom, mode="inside")
        return [re.sub(r"\s+", " ", ln).strip() for ln in txt.splitlines() if ln.strip()]

    def remove_labels(block, *labels):
        out = []
        for ln in block:
            if any(re.fullmatch(lab, ln, re.I) for lab in labels):
                continue
            out.append(ln)
        return out

    # B/L No.: 우측 상단 또는 좌측 하단의 Bill of Lading No. 아래 숫자
    bl = ""
    bl_blocks = [
        region_lines(w * 0.44, h * 0.125, w * 0.60, h * 0.165),
        region_lines(w * 0.07, h * 0.865, w * 0.26, h * 0.900),
    ]
    for block in bl_blocks:
        cand = " ".join(block)
        m = re.search(r"\b(\d{10,20})\b", cand)
        if m:
            bl = m.group(1)
            break
    if not bl:
        # fallback: 파일 전체에서 긴 숫자 후보 중 B/L에 가까운 값 사용
        nums = re.findall(r"\b\d{12,20}\b", text)
        if nums:
            bl = nums[-1]

    # SHIPPER: 좌측 상단 SHIPPER 칸. 첫 줄=쉬퍼, 나머지=쉬퍼주소
    shipper_block = remove_labels(
        region_lines(w * 0.07, h * 0.085, w * 0.47, h * 0.164),
        r"SHIPPER"
    )
    shipper = shipper_block[0] if shipper_block else ""
    shipper_addr = "\n".join(shipper_block[1:]).strip()

    # CONSIGNEE: 좌측 CONSIGNEE 칸. 첫 줄=컨사이니, 나머지=컨사이니주소
    consignee_block = remove_labels(
        region_lines(w * 0.07, h * 0.155, w * 0.47, h * 0.235),
        r"CONSIGNEE", r"NOTIFY PARTY", r"SAME AS CONSIGNEE.*"
    )
    consignee_raw = consignee_block[0] if consignee_block else ""
    consignee_no = extract_business_no(consignee_raw)
    consignee = clean_company_with_paren(consignee_raw)
    consignee_addr = "\n".join(consignee_block[1:]).strip()
    consignee_addr, consignee_no = clean_consignee_address_and_bizno(consignee_addr, consignee_no)

    # Vessel/Voyage No. 칸 아래 값: 마지막 3~4자리+E = 항차, 앞부분 = 선명
    vessel_region = " ".join(region_lines(w * 0.07, h * 0.335, w * 0.34, h * 0.365))
    vessel_region = re.sub(r"Vessel\s*/\s*Voyage\s+No\.?", " ", vessel_region, flags=re.I)
    vessel, voyage = split_vessel_voyage_text(vessel_region)

    # PORT OF LOADING / DISCHARGE 칸 아래 값
    pol_raw = " ".join(region_lines(w * 0.07, h * 0.374, w * 0.21, h * 0.402))
    pod_raw = " ".join(region_lines(w * 0.24, h * 0.374, w * 0.45, h * 0.402))
    pol_raw = re.sub(r"PORT\s+OF\s+LOADING", " ", pol_raw, flags=re.I).strip()
    pod_raw = re.sub(r"PORT\s+OF\s+DISCHARGE", " ", pod_raw, flags=re.I).strip()
    pol = normalize_port_code(pol_raw)
    pod = normalize_port_code(pod_raw)

    # 기존 표 영역 인식은 유지: N/M 줄부터 수량/품명/중량/CBM 추출
    start_idx = None
    for i, ln in enumerate(lines):
        if re.search(r"^N/M\b", ln, re.I):
            start_idx = i
            break

    mark = ""
    pkg = ""
    weight = ""
    cbm = ""
    description_lines = []

    if start_idx is not None:
        first = lines[start_idx]
        mark = "N/M"

        m_pkg = re.search(r"\b(\d+)\s*(?:CTNS?|PKGS?|PACKAGES?|CARTONS?|PCS)\b", first, re.I)
        if m_pkg:
            pkg = m_pkg.group(1)

        desc_first = re.sub(r"^N/M\s*", "", first, flags=re.I)
        desc_first = re.sub(r"\b\d+\s*(?:CTNS?|PKGS?|PACKAGES?|CARTONS?|PCS)\b", "", desc_first, flags=re.I).strip()
        desc_first = re.sub(r"\b\d+(?:\.\d+)?\s*KGS?\b", "", desc_first, flags=re.I).strip()
        desc_first = re.sub(r"\b\d+(?:\.\d+)?\s*CBM\b", "", desc_first, flags=re.I).strip()
        if desc_first:
            description_lines.append(desc_first)

        for ln in lines[start_idx + 1:]:
            if re.search(r"SHIPPER['’]?S\s+LOAD|S\.T\.C|SHIPPED\s+ON\s+BOARD|FREIGHT\s+PREPAID|Total Number|SAY TOTAL", ln, re.I):
                break
            ln_clean = re.sub(r"\b\d+(?:\.\d+)?\s*KGS?\b", "", ln, flags=re.I)
            ln_clean = re.sub(r"\b\d+(?:\.\d+)?\s*CBM\b", "", ln_clean, flags=re.I).strip()
            if ln_clean:
                description_lines.append(ln_clean)

    m_w = re.search(r"\b(\d+(?:\.\d+)?)\s*KGS?\b", text, re.I)
    if m_w:
        weight = m_w.group(1)
    m_c = re.search(r"\b(\d+(?:\.\d+)?)\s*CBM\b", text, re.I)
    if m_c:
        cbm = m_c.group(1)

    description = "\n".join(description_lines).strip()

    if not bl:
        warnings.append("비엘 미인식")
    if not shipper:
        warnings.append("쉬퍼 미인식")
    if not consignee:
        warnings.append("컨사이니 미인식")
    if not vessel:
        warnings.append("선명 미인식")
    if not voyage:
        warnings.append("항차 미인식")
    if not pol:
        warnings.append("출발지 미인식")
    if not pod:
        warnings.append("도착지 미인식")
    if not mark:
        warnings.append("마크 미인식")
    if not description:
        warnings.append("품명 미인식")
    if not weight:
        warnings.append("중량 확인필요")
    if not cbm:
        warnings.append("CBM 확인필요")
    if not pkg:
        warnings.append("수량 확인필요")

    return {
        "비엘": bl,
        "쉬퍼": shipper,
        "쉬퍼주소": shipper_addr,
        "컨사이니": consignee,
        "컨사이니 사업자번호": consignee_no,
        "컨사이니 주소": consignee_addr,
        "선명": vessel,
        "항차": voyage,
        "출발지": pol,
        "도착지": pod,
        "품명": description,
        "마크": mark,
        "수량": pkg,
        "중량": weight,
        "CBM": cbm,
        "원본파일명": filename,
        "확인필요": ", ".join(warnings),
    }


def _strip_label_text(text):
    """라벨 문구가 같이 잡힌 경우 값만 남기기."""
    text = text or ""
    text = re.sub(r"^.*?\b(?:Consignee|Shipper|Notify\s+party|Ocean\s+vessel|Port\s+of\s+loading|Port\s+of\s+discharge|Place\s+of\s+receipt|Place\s+of\s+delivery)\b.*?:?", "", text, flags=re.I).strip()
    return text


def _normalize_space_keep_lines(lines):
    cleaned = []
    for ln in lines:
        ln = re.sub(r"\s+", " ", (ln or "")).strip()
        if ln:
            cleaned.append(ln)
    return cleaned


def _block_by_left_labels(page, start_pattern, end_pattern, *, x_limit_ratio=0.55):
    """SILKROAD 특수 BL용: 왼쪽 상단 영역의 제목 기준으로 블록 추출."""
    w = page.width
    words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
    lines = group_words_to_lines(words, y_tol=4)
    left_lines = [ln for ln in lines if ln["x0"] < w * x_limit_ratio]

    start_y = None
    end_y = None
    for ln in left_lines:
        if start_y is None and re.search(start_pattern, ln["text"], re.I):
            start_y = ln["bottom"]
            continue
        if start_y is not None and re.search(end_pattern, ln["text"], re.I):
            end_y = ln["top"]
            break
    if start_y is None:
        return []
    if end_y is None:
        end_y = page.height

    block = []
    for ln in left_lines:
        if ln["top"] >= start_y - 1 and ln["bottom"] <= end_y + 1:
            # 라벨 줄이 같이 걸리는 경우 제외
            if re.search(start_pattern, ln["text"], re.I) or re.search(end_pattern, ln["text"], re.I):
                continue
            block.append(ln["text"])
    return _normalize_space_keep_lines(block)


def _extract_after_label_near(page, label_pattern, *, x0_ratio=0, x1_ratio=0.55, y_after=None, y_before=None):
    """라벨과 같은 줄 또는 바로 아래 줄에 있는 값을 추출."""
    w, h = page.width, page.height
    words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
    lines = group_words_to_lines(words, y_tol=4)
    candidates = []
    for i, ln in enumerate(lines):
        if not (w * x0_ratio <= ln["x0"] <= w * x1_ratio):
            continue
        if y_after is not None and ln["top"] < y_after:
            continue
        if y_before is not None and ln["top"] > y_before:
            continue
        if re.search(label_pattern, ln["text"], re.I):
            # 같은 줄 뒤쪽에 값이 같이 있으면 우선 사용
            same = re.sub(label_pattern + r".*?", "", ln["text"], flags=re.I).strip(" :")
            if same:
                candidates.append(same)
            # 다음 1~2줄 값도 확인
            for nxt in lines[i+1:i+3]:
                if nxt["top"] - ln["bottom"] < h * 0.045:
                    candidates.append(nxt["text"])
            break
    return " ".join(_normalize_space_keep_lines(candidates)).strip()


def extract_silkroad_pdf(page, filename):
    """SREJY / SILKROAD LOGISTICS 특수 BL 전용 추출.
    일반 BL 로직은 건드리지 않고, 해당 양식일 때만 보정합니다.
    """
    warnings = []
    w, h = page.width, page.height
    text = page.extract_text() or ""
    flat = re.sub(r"\s+", " ", text)

    words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
    lines_obj = group_words_to_lines(words, y_tol=4)
    line_texts = [ln["text"] for ln in lines_obj]

    # B/L No.
    bl = safe_search(r"\b(SREJY[A-Z0-9]{6,}|[A-Z]{3,}[A-Z0-9]{8,})\b", text)

    # 상단 좌측 실제 값 기준 추출: pdfplumber에서 라벨은 누락될 수 있어 값의 위치/패턴을 기준으로 보정
    left_top = [ln for ln in lines_obj if ln["x0"] < w * 0.45 and ln["top"] < h * 0.31]

    # Shipper: 상단 첫 회사명부터 사업자번호 포함 컨사이니 시작 전까지
    shipper_lines = []
    for ln in left_top:
        t = ln["text"].strip()
        if not t or t == bl:
            continue
        if extract_business_no(t):
            break
        # B/L No. 우측 텍스트 방지
        if ln["x0"] < w * 0.10:
            shipper_lines.append(t)
    shipper = shipper_lines[0] if shipper_lines else ""
    shipper_addr = "\n".join(shipper_lines[1:]).strip()

    # Consignee: 사업자번호 있는 줄부터 Notify(SILKROAD LOGISTICS CO., LTD) 전까지
    consignee_lines = []
    collecting = False
    for ln in left_top:
        t = ln["text"].strip()
        if not collecting and extract_business_no(t):
            collecting = True
        if collecting:
            # Notify Party 값 시작 전까지만
            if t.upper().startswith("SILKROAD LOGISTICS CO") and not extract_business_no(t):
                break
            consignee_lines.append(t)
    consignee_raw = consignee_lines[0] if consignee_lines else ""
    consignee_no = extract_business_no(consignee_raw)
    consignee = clean_company_with_paren(consignee_raw)
    consignee_addr = "\n".join(consignee_lines[1:]).strip()
    consignee_addr, consignee_no = clean_consignee_address_and_bizno(consignee_addr, consignee_no)

    # 선명/항차: HUADONG PEARL8 7264E 처럼 한 줄에 붙거나 뒤에 항구가 붙어도 분리
    vessel = ""
    voyage = ""
    for src in line_texts + [flat]:
        m = re.search(r"\b([A-Z][A-Z0-9]*(?:\s+[A-Z0-9]+)*?)\s+([0-9]{3,5}E)\b", src, re.I)
        if m:
            cand_v = re.sub(r"\s+", " ", m.group(1)).strip()
            cand_y = m.group(2).strip()
            if re.search(r"HUADONG|PEARL|HANSUNG|GRAND|PEACE|INCHEON", cand_v, re.I):
                vessel, voyage = cand_v, cand_y
                break
    if re.fullmatch(r"\d+", voyage or ""):
        voyage = f"{voyage}E"

    # 출발지/도착지: 해당 특수 양식은 SHIDAO / INCHEON 값을 코드로 변환
    pol_raw = "SHIDAO,CHINA" if re.search(r"SHIDAO\s*,?\s*CHINA", text, re.I) else ""
    pod_raw = "INCHEON,KOREA" if re.search(r"INCHEON\s*,?\s*KOREA", text, re.I) else ""
    pol = normalize_port_code(pol_raw)
    pod = normalize_port_code(pod_raw)

    # 마크: 표 왼쪽 마크 값만. 수량/품명 안내문구가 같은 줄에 붙으면 제거
    mark_candidates = []
    for ln in lines_obj:
        t = ln["text"].strip()
        if ln["top"] > h * 0.43 and ln["top"] < h * 0.58 and ln["x0"] < w * 0.20:
            if re.match(r"^(SILK|SMARTGO|[A-Z]{2,}[-A-Z0-9]*)", t, re.I):
                t = re.split(r"\b\d+\s*(?:PACKAGES?|PKGS?|CTNS?|CARTONS?|CARTON|PCS|BOXES|BOX)\b", t, flags=re.I)[0].strip()
                t = re.split(r"\bSAID\s+TO\s+CONTAIN\b|SHIPPER", t, flags=re.I)[0].strip()
                if t:
                    mark_candidates.append(t)
    mark = clean_mark("\n".join(mark_candidates))

    # 수량/중량/CBM: 단위 앞 숫자만
    m_pkg = re.search(r"\b(\d+(?:\.\d+)?)\s*(?:PACKAGES?|PKGS?|CTNS?|CARTONS?|CARTON|PCS|BOXES|BOX)\b", text, re.I)
    pkg = m_pkg.group(1) if m_pkg else ""
    m_w = re.search(r"\b(\d+(?:\.\d+)?)\s*KGS?\b", text, re.I)
    weight = m_w.group(1) if m_w else ""
    m_c = re.search(r"\b(\d+(?:\.\d+)?)\s*CBM\b", text, re.I)
    cbm = m_c.group(1) if m_c else ""

    # 품명: SAID TO CONTAIN 아래 실제 품명줄만. 좌표/문구 두 방식 병행
    desc_lines = []
    # 1순위: 표 품명칸 좌표 기준. 안내문구, 중량/CBM, SURRENDERED 제외
    for ln in lines_obj:
        t = ln["text"].strip()
        if ln["top"] > h * 0.48 and ln["top"] < h * 0.57 and w * 0.25 < ln["x0"] < w * 0.60:
            if re.search(r"SHIPPER|SAID\s+TO\s+CONTAIN|SURRENDERED|LOADED|FREIGHT|KGS?|CBM|PACKAGES?", t, re.I):
                continue
            if t:
                desc_lines.append(t)
    description = "\n".join(desc_lines).strip()
    # 2순위: 텍스트 기준
    if not description:
        m = re.search(r"SAID\s+TO\s+CONTAIN\s*:?\s*(.*?)(?:\n\s*(?:SURRENDERED|LOADED\s+ON\s+BOARD|FREIGHT\s+COLLECT|CFS/CFS|DESTINATION|AS\s+ARRANGED)\b)", text, flags=re.I | re.S)
        if m:
            description = "\n".join(
                ln.strip()
                for ln in m.group(1).splitlines()
                if ln.strip() and not re.search(r"SHIPPER|SAID\s+TO\s+CONTAIN", ln, re.I)
            ).strip()

    if not bl:
        warnings.append("비엘 미인식")
    if not consignee_addr:
        warnings.append("컨사이니 주소 확인필요")
    if not vessel:
        warnings.append("선명 확인필요")
    if not voyage:
        warnings.append("항차 확인필요")
    if not pod:
        warnings.append("도착지 확인필요")
    if not mark:
        warnings.append("마크 미인식")
    if not description:
        warnings.append("품명 미인식")
    if not weight:
        warnings.append("중량 확인필요")
    if not cbm:
        warnings.append("CBM 확인필요")
    if not pkg:
        warnings.append("수량 확인필요")

    return {
        "비엘": bl,
        "쉬퍼": shipper,
        "쉬퍼주소": shipper_addr,
        "컨사이니": consignee,
        "컨사이니 사업자번호": consignee_no,
        "컨사이니 주소": consignee_addr,
        "선명": vessel,
        "항차": voyage,
        "출발지": pol,
        "도착지": pod,
        "품명": description,
        "마크": mark,
        "수량": pkg,
        "중량": weight,
        "CBM": cbm,
        "원본파일명": filename,
        "확인필요": ", ".join(warnings),
    }



def fix_split_business_no_text(text):
    """452-\n81-03361처럼 줄바꿈으로 끊어진 사업자번호를 452-81-03361로 보정합니다."""
    text = text or ""
    text = re.sub(r"(\d{3})-\s*\n\s*(\d{2}-\d{5})", r"\1-\2", text)
    text = re.sub(r"(\d{3})-\s+(\d{2}-\d{5})", r"\1-\2", text)
    return text


def extract_first_bl_number(text):
    text = text or ""
    # HKD 계열은 영문+숫자 혼합 BL 우선
    m = re.search(r"\b(HKD[A-Z0-9]{8,25})\b", text, re.I)
    if m:
        return m.group(1).upper()
    m = re.search(r"Bill\s*of\s*Lading\s*No\.?\s*\n\s*([A-Z0-9]{8,30})", text, re.I)
    if m:
        return m.group(1).strip()
    m = re.search(r"\b([A-Z]{2,}[A-Z0-9]{6,})\b", text, re.I)
    return m.group(1).strip() if m else ""


def normalize_broken_unit_text(text):
    text = text or ""
    # PDF 텍스트 겹침으로 2C TN S / 2 BO XE S처럼 끊긴 단위를 복구
    text = re.sub(r"(\d+)\s*C\s*T\s*N\s*S", r"\1CTNS", text, flags=re.I)
    text = re.sub(r"(\d+)\s*P\s*K\s*G\s*S", r"\1PKGS", text, flags=re.I)
    text = re.sub(r"(\d+)\s*P\s*L\s*T\s*S", r"\1PLTS", text, flags=re.I)
    text = re.sub(r"(\d+)\s+BO\s+\d+\s*C\s*XE\s*TN\s*S\s*S", r"\1 BOXES", text, flags=re.I)
    text = re.sub(r"(\d+)\s*BO\s*XE\s*S", r"\1BOXES", text, flags=re.I)
    return text


def clean_line_remove_bl(line, bl):
    line = line or ""
    if bl:
        line = line.replace(bl, " ")
    return re.sub(r"\s+", " ", line).strip(" ,")


def extract_xingwen_hkd_pdf(page, filename):
    """HKD / XINGWEN / HAOKUNDA 계열 특수 BL 전용 추출.
    일반 BL은 건드리지 않고, 이 계열만 좌상단 Shipper/Consignee 칸을 줄 흐름 기준으로 읽습니다.
    """
    warnings = []
    raw_text = page.extract_text() or ""
    text = normalize_broken_unit_text(fix_split_business_no_text(raw_text))
    w, h = page.width, page.height
    words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
    lines_obj = group_words_to_lines(words, y_tol=3)

    bl = extract_first_bl_number(text)

    # 좌측 상단 라인들: 오른쪽으로 조금 넘어간 텍스트도 같은 줄이면 포함하되, BL 번호는 제거
    geo_lines = []
    for ln in lines_obj:
        t = normalize_broken_unit_text(clean_line_remove_bl(ln["text"], bl))
        if not t:
            continue
        geo_lines.append({**ln, "text": t})

    # Shipper: 상단 첫 구역. x 좌측 시작 라인만 사용, BL 번호가 같은 줄에 붙으면 제거됨.
    shipper_lines = []
    for ln in geo_lines:
        if ln["top"] < h * 0.105 and ln["x0"] < w * 0.18:
            t = ln["text"].strip()
            if re.fullmatch(r"SHIPPER", t, re.I):
                continue
            shipper_lines.append(t)
    shipper = shipper_lines[0] if shipper_lines else ""
    shipper_addr = "\n".join(shipper_lines[1:]).strip()

    # Consignee: 선으로 나뉜 컨사이니 구역. 첫 줄=컨사이니, 나머지=주소.
    consignee_lines = []
    for ln in geo_lines:
        if h * 0.105 <= ln["top"] < h * 0.215 and ln["x0"] < w * 0.18:
            t = ln["text"].strip()
            if re.fullmatch(r"CONSIGNEE", t, re.I):
                continue
            # Notify 라벨/값은 컨사이니 주소에서 제외
            if re.fullmatch(r"NOTIFY\s+PARTY", t, re.I) or t.upper().startswith("SAME AS ABOVE"):
                continue
            consignee_lines.append(t)

    consignee_raw = ""
    addr_start = 1
    if consignee_lines:
        if len(consignee_lines) >= 2 and re.search(r"\d{3}-$", consignee_lines[0]) and re.fullmatch(r"\d{2}-\d{5}", consignee_lines[1]):
            consignee_raw = consignee_lines[0] + consignee_lines[1]
            addr_start = 2
        else:
            consignee_raw = consignee_lines[0]
            addr_start = 1
    consignee_zone_text = "\n".join(consignee_lines)
    # 컨사이니 첫 줄에는 사업자번호가 있으면 그대로 유지
    consignee = re.sub(r"\s+", " ", consignee_raw).strip()
    consignee_no = extract_business_no(consignee_zone_text) or extract_business_no(consignee)
    consignee_addr = "\n".join(consignee_lines[addr_start:]).strip()
    consignee_addr, consignee_no = clean_consignee_address_and_bizno(consignee_addr, consignee_no)

    # 선명/항차
    vessel, voyage = split_vessel_voyage_text(text)

    # 출발지/도착지: Ocean Vessel 라인 기준. 같은 줄에 WEIHAI CHINA이 붙은 경우도 처리.
    pol_raw = ""
    pod_raw = ""
    for i, ln in enumerate(geo_lines):
        t = ln["text"]
        if re.search(r"HANSUNG\s+INCHEON\s+\d{3,4}E?", t, re.I):
            after = re.split(r"HANSUNG\s+INCHEON\s+\d{3,4}E?", t, flags=re.I)[-1].strip()
            if after:
                pol_raw = after
            # 다음 1~2개 라인에서 POL/POD 후보 찾기
            for nxt in geo_lines[i+1:i+4]:
                nt = nxt["text"].strip()
                if not pol_raw and re.search(r"WEIHAI|SHIDAO|YANTAI", nt, re.I):
                    pol_raw = nt
                if not pod_raw and re.search(r"INCHEON|INCHON", nt, re.I):
                    pod_raw = nt
            break
    if not pol_raw:
        m = re.search(r"\b(WEIHAI\s+CHINA|WEIHAI,\s*CHINA|SHIDAO(?:,?\s*CHINA)?)\b", text, re.I)
        pol_raw = m.group(1) if m else ""
    if not pod_raw:
        m = re.search(r"\b(INCHEON\s*,?\s*KOREA|INCHON\s*,?\s*KOREA)\b", text, re.I)
        pod_raw = m.group(1) if m else ""
    pol = normalize_port_code(pol_raw)
    pod = normalize_port_code(pod_raw)

    # 수량/중량/CBM: 단위 앞 숫자만. 겹쳐 깨진 2C TN S도 보정된 text에서 추출.
    m_pkg = re.search(r"\b(\d+(?:\.\d+)?)\s*(?:CTNS?|CARTONS?|PKGS?|PACKAGES?|PLTS?|BOXES?|PCS)\b", text, re.I)
    pkg = m_pkg.group(1) if m_pkg else ""
    m_w = re.search(r"\b(\d+(?:\.\d+)?)\s*KGS?\b", text, re.I)
    weight = m_w.group(1) if m_w else ""
    m_c = re.search(r"\b(\d+(?:\.\d+)?)\s*CBM\b", text, re.I)
    cbm = m_c.group(1) if m_c else ""

    # 품명: SAID TO CONTAIN 아래 실제 품명만. SURRENDERED 제외.
    description_lines = []
    in_desc = False
    for ln in geo_lines:
        if ln["top"] < h * 0.38:
            continue
        t = normalize_broken_unit_text(ln["text"].strip())
        up = t.upper()
        if re.search(r"SAID\s+TO\s+CONTAIN", up):
            in_desc = True
            tail = re.split(r"SAID\s+TO\s+CONTAIN", t, flags=re.I)[-1].strip(" :-")
            if tail:
                description_lines.append(tail)
            continue
        if not in_desc:
            continue
        if re.search(r"^(SURRENDERED|COPY|FREIGHT\b|SHIPPED\b|LOADED\b|ONE\b|SAY\b|TOTAL\b|YST\b|E-MAIL|TEL\b|FAX\b|\d{4}[-./]\d{1,2}[-./]\d{1,2})", up):
            break
        if re.search(r"\b(?:KGS?|CBM|CTNS?|PKGS?|PLTS?|BOXES?)\b", up):
            continue
        # E04617 TELESCOPE처럼 마크/실번호와 품명이 같은 줄에 겹친 경우 앞 코드는 제외하고 뒤 품명만 살림
        m_code_desc = re.match(r"^(?:[A-Z]{4}\d{7}|[A-Z]{1,4}\d{1,5}[-~]\d{1,5}|E\d{3,}|\d{4,6})\s+(.+)$", t, re.I)
        if m_code_desc:
            tail_desc = m_code_desc.group(1).strip()
            if tail_desc and not re.search(r"SAID\s+TO\s+CONTAIN", tail_desc, re.I):
                description_lines.append(tail_desc)
            continue
        if re.match(r"^(OL-|[A-Z]{4}\d{7}\b|[A-Z]{1,4}\d{1,5}[-~]\d{1,5}|E\d{3,}|\d{4,6}$)", t, re.I):
            continue
        description_lines.append(t)
    description = "\n".join(description_lines).strip()

    # 마크: 표 왼쪽/연속 텍스트. OL- 한 줄은 구역을 넘어가도 한 마크로 유지.
    mark_lines = []
    for ln in geo_lines:
        if ln["top"] < h * 0.39:
            continue
        t = normalize_broken_unit_text(ln["text"].strip())
        up = t.upper()
        if re.search(r"^(FREIGHT\b|SURRENDERED|COPY|SHIPPED\b|ONE\b|SAY\b|YST\b|E-MAIL|TEL\b|FAX\b)", up):
            break
        if up.startswith("OL-"):
            t = re.split(r"SHIPPER", t, flags=re.I)[0].strip()
            t = re.sub(r"(\d+)BOXES", r"\1 BOXES", t, flags=re.I)
            mark_lines.append(t)
            continue
        # HHXU3325532 SAID TO CONTAIN처럼 마크와 안내문구가 같은 줄이면 앞 마크는 살림
        if re.search(r"SAID\s+TO\s+CONTAIN", up):
            prefix_said = re.split(r"SAID\s+TO\s+CONTAIN", t, flags=re.I)[0].strip()
            if prefix_said and re.match(r"^([A-Z]{4}\d{7}|[A-Z]{1,4}\d{1,5}[-~]\d{1,5}|E\d{3,}|\d{4,6})$", prefix_said, re.I):
                mark_lines.append(prefix_said)
            continue
        if ln["x0"] < w * 0.18:
            prefix = re.split(r"\b\d+(?:\.\d+)?\s*(?:CTNS?|CARTONS?|PKGS?|PACKAGES?|PLTS?|BOXES?|PCS|KGS?|CBM)\b|SHIPPER", t, flags=re.I)[0].strip()
            if prefix and not re.search(r"HANSUNG|WEIHAI|INCHEON|KOREA", prefix, re.I):
                mark_lines.append(prefix)
        # 품명 뒤쪽의 컨테이너/실번호 계열. E04617 TELESCOPE면 E04617만 마크로 사용.
        else:
            m_mark = re.match(r"^([A-Z]{4}\d{7}|[A-Z]{1,4}\d{1,5}[-~]\d{1,5}|E\d{3,}|\d{4,6})(?:\s+.+)?$", t, re.I)
            if m_mark:
                mark_lines.append(m_mark.group(1))
    mark = clean_mark("\n".join(dict.fromkeys([m for m in mark_lines if m]).keys()))

    if not bl:
        warnings.append("비엘 미인식")
    if not shipper:
        warnings.append("쉬퍼 확인필요")
    if not consignee:
        warnings.append("컨사이니 확인필요")
    if not consignee_addr:
        warnings.append("컨사이니 주소 확인필요")
    if not vessel:
        warnings.append("선명 확인필요")
    if not voyage:
        warnings.append("항차 확인필요")
    if not description:
        warnings.append("품명 미인식")
    if not pkg:
        warnings.append("수량 확인필요")
    if not weight:
        warnings.append("중량 확인필요")
    if not cbm:
        warnings.append("CBM 확인필요")

    return {
        "비엘": bl,
        "쉬퍼": shipper,
        "쉬퍼주소": shipper_addr,
        "컨사이니": consignee,
        "컨사이니 사업자번호": consignee_no,
        "컨사이니 주소": consignee_addr,
        "선명": vessel,
        "항차": voyage,
        "출발지": pol,
        "도착지": pod,
        "품명": description,
        "마크": mark,
        "수량": pkg,
        "중량": weight,
        "CBM": cbm,
        "원본파일명": filename,
        "확인필요": ", ".join(warnings),
    }


def extract_one_pdf(file_bytes, filename, desc_right_ratio=0.89, table_bottom_ratio=0.69):
    warnings = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        if not pdf.pages:
            return {**{c: "" for c in COLUMNS}, "원본파일명": filename, "확인필요": "PDF 페이지 없음"}

        page = pdf.pages[0]
        all_text = page.extract_text() or ""


        # HKD / XINGWEN / HAOKUNDA 계열 특수 양식
        if re.search(r"\bHKD[A-Z0-9]{8,25}\b", all_text.upper()) and ("HANSUNG INCHEON" in all_text.upper()):
            return extract_xingwen_hkd_pdf(page, filename)

        # SILKROAD / SREJY 특수 양식은 WOOYOUNG 조건보다 먼저 분기해야 함
        # (일부 PDF 텍스트에 MUTIMODAL 문구가 같이 잡혀도 SILKROAD로 처리)
        if ("SILKROAD LOGISTICS" in all_text.upper()) or re.search(r"\bSREJY[A-Z0-9]{6,}\b", all_text.upper()):
            return extract_silkroad_pdf(page, filename)

        # WOOYOUNG / INUF 특수 양식
        if ("WOOYOUNG (CHINA) LOGISTICS" in all_text.upper()) or ("MUTIMODAL TRANSPORT BILL OF LADING" in all_text.upper()):
            return extract_wooyoung_pdf(page, filename)

        w, h = page.width, page.height

        bl_region = text_in_region(page, w * 0.58, 0, w, h * 0.18)
        bl = safe_search(r"\b([A-Z0-9]{8,})\b", bl_region)

        shipper_block = text_in_region(page, 0, h * 0.02, w * 0.57, h * 0.10)
        shipper_lines = [x.strip() for x in shipper_block.splitlines() if x.strip()]
        shipper = shipper_lines[0] if shipper_lines else ""
        shipper_addr = "\n".join(shipper_lines[1:]).strip()
        if shipper and not shipper_addr:
            shipper_addr = "CHINA"

        consignee_block = text_in_region(page, 0, h * 0.09, w * 0.57, h * 0.16)
        consignee_lines = [x.strip() for x in consignee_block.splitlines() if x.strip()]
        consignee_raw = consignee_lines[0] if consignee_lines else ""
        consignee_no = extract_business_no(consignee_raw)
        consignee = clean_company_with_paren(consignee_raw)
        consignee_addr = "\n".join(consignee_lines[1:])
        consignee_addr, consignee_no = clean_consignee_address_and_bizno(consignee_addr, consignee_no)

        vessel_region = text_in_region(page, 0, h * 0.31, w * 0.25, h * 0.37).replace("\n", " ")
        vessel = ""
        voyage = ""
        m = re.search(r"([A-Z]+(?:\s+[A-Z0-9]+)*)\s+([0-9A-Z]+)$", vessel_region)
        if m:
            vessel, voyage = m.group(1).strip(), m.group(2).strip()
        else:
            vessel = vessel_region.strip()
        if re.fullmatch(r"\d+", voyage or ""):
            voyage = f"{voyage}E"

        # HANSUNG 양식 보정: HANSUNG INCHEON까지 선명, 오른쪽 3~4자리(+E)는 항차
        hv, hy = split_vessel_voyage_text(all_text)
        if hv and hv.upper().startswith("HANSUNG INCHEON"):
            vessel, voyage = hv, hy

        pol_raw = text_in_region(page, w * 0.25, h * 0.31, w * 0.48, h * 0.37).replace("\n", " ").strip()
        pod_raw = text_in_region(page, 0, h * 0.35, w * 0.28, h * 0.40).replace("\n", " ").strip()
        pol = normalize_port_code(pol_raw)
        pod = normalize_port_code(pod_raw)

        table_top = h * 0.39
        table_bottom = h * table_bottom_ratio

        mark_words = words_in_region(page, 0, table_top, w * 0.28, table_bottom, mode="start")
        mark = clean_mark(text_from_words(mark_words))

        pkg_region = text_in_region(page, w * 0.27, table_top, w * 0.39, table_top + h * 0.08, mode="start")
        pkg = safe_search(r"(\d+)\s*(?:PKGS?|PACKAGES?|CTNS?|CARTONS?|PCS)", pkg_region)
        if not pkg:
            pkg = extract_qty_from_mark(mark)

        desc_left = w * 0.37
        desc_right = w * desc_right_ratio
        desc_words = words_in_region(page, desc_left, table_top, desc_right, table_bottom, mode="start")
        description = clean_description_words(desc_words)

        weight_region = text_in_region(page, w * 0.76, table_top, w * 0.90, table_top + h * 0.08, mode="start").replace("\n", " ")
        weight = safe_search(r"(\d+(?:\.\d+)?)\s*KGS", weight_region)

        cbm_region = text_in_region(page, w * 0.88, table_top, w, table_top + h * 0.08, mode="start").replace("\n", " ")
        cbm = safe_search(r"(\d+(?:\.\d+)?)\s*CBM", cbm_region)

        if consignee_raw and "(" in consignee_raw and consignee.endswith("()"):
            consignee = consignee.replace("()", "").strip()
        if not mark:
            warnings.append("마크 미인식")
        if not description:
            warnings.append("품명 미인식")
        if not weight:
            warnings.append("중량 확인필요")
        if not cbm:
            warnings.append("CBM 확인필요")
        if not pkg:
            warnings.append("수량 확인필요")

        return {
            "비엘": bl,
            "쉬퍼": shipper,
            "쉬퍼주소": shipper_addr,
            "컨사이니": consignee,
            "컨사이니 사업자번호": consignee_no,
            "컨사이니 주소": consignee_addr,
            "선명": vessel,
            "항차": voyage,
            "출발지": pol,
            "도착지": pod,
            "품명": description,
            "마크": mark,
            "수량": pkg,
            "중량": weight,
            "CBM": cbm,
            "원본파일명": filename,
            "확인필요": ", ".join(warnings),
        }

def make_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="변환결과")
        ws = writer.book["변환결과"]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
            cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
        widths = {
            "A": 22, "B": 30, "C": 42, "D": 24, "E": 20, "F": 45,
            "G": 20, "H": 12, "I": 18, "J": 18, "K": 55, "L": 35,
            "M": 10, "N": 12, "O": 10, "P": 34, "Q": 28,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 95
    output.seek(0)
    return output.getvalue()

KAKAO_API_URL = "https://dapi.kakao.com/v2/local/search/address.json"

HOUSE_NUM = r"\d{1,4}(?:-\d{1,4})?"
ADMIN_JIBUN = r"(?:동|리|면|읍|가)"
ROAD_SUFFIX = r"(?:로|길|대로|거리)"

GENERAL_ROAD_PATTERN = rf"[0-9가-힣A-Za-z\.]+{ROAD_SUFFIX}\s*{HOUSE_NUM}"
COMPLEX_ROAD_PATTERNS = [
    rf"[0-9가-힣A-Za-z\.]+로\s*\d+[가-힣]?(?:번길|번로|길)\s*{HOUSE_NUM}",
    rf"[0-9가-힣A-Za-z\.]+로\d+[가-힣]?(?:번길|번로|길)\s*{HOUSE_NUM}",
    rf"[0-9가-힣A-Za-z\.]+길\s*\d+(?:번길|번로)\s*{HOUSE_NUM}",
    rf"[0-9가-힣A-Za-z\.]+길\d+(?:번길|번로)\s*{HOUSE_NUM}",
    rf"[0-9가-힣A-Za-z\.]+대로\s*\d+(?:번길|번로)\s*{HOUSE_NUM}",
    rf"[0-9가-힣A-Za-z\.]+대로\d+(?:번길|번로)\s*{HOUSE_NUM}",
    rf"[0-9가-힣A-Za-z\.]+거리\s*\d+(?:번길|번로)\s*{HOUSE_NUM}",
    rf"[0-9가-힣A-Za-z\.]+거리\d+(?:번길|번로)\s*{HOUSE_NUM}",
]
JIBUN_PATTERN = rf"[0-9가-힣A-Za-z\.]+{ADMIN_JIBUN}\s*{HOUSE_NUM}(?:번지)?"

APT_KEYWORDS = [
    "아파트", "APT", "Apartment", "APARTMENT", "타워", "빌라", "맨션", "하우스",
    "래미안", "푸르지오", "자이", "힐스테이트", "아이파크", "e편한세상", "롯데캐슬", "더샵",
    "센트럴파크", "트리플시티", "리버파크", "스카이", "시티", "주공", "LH", "현대", "삼성",
    "대림", "대우", "경남", "쌍용", "SK", "SK뷰", "서희", "휴먼시아", "임광", "롯데캐슬"
]
APT_NAME_RE = re.compile("|".join(re.escape(x) for x in APT_KEYWORDS), re.IGNORECASE)
DONG_HO_RE = re.compile(r"\b\d+\s*동\b.*\b\d+\s*호\b|\b\d+\s*호\b|\b\d+\s*동\b")
NUM_UNIT_RE = re.compile(r"\b\d{1,4}\s*-\s*\d{1,4}\b")

REGION_HINTS_ORDERED = [
    ("chungcheongbuk-do", "충청북도"),
    ("chungcheongnam-do", "충청남도"),
    ("gyeongsangbuk-do", "경상북도"),
    ("gyeongsangnam-do", "경상남도"),
    ("changnyeong-gun", "창녕군"),
    ("jeungpyeong-gun", "증평군"),
    ("yeongdeungpo-gu", "영등포구"),
    ("cheongyang-gun", "청양군"),
    ("dongducheon-si", "동두천시"),
    ("masanhoewon-gu", "마산회원구"),
    ("yangpyeong-gun", "양평군"),
    ("dongdaemun-gu", "동대문구"),
    ("hampyeong-gun", "함평군"),
    ("hongcheon-gun", "홍천군"),
    ("hongseong-gun", "홍성군"),
    ("jangheung-gun", "장흥군"),
    ("pyeongtaek-si", "평택시"),
    ("sancheong-gun", "산청군"),
    ("yeongcheon-si", "영천시"),
    ("yeongdeok-gun", "영덕군"),
    ("cheongwon-gu", "청원구"),
    ("cheorwon-gun", "철원군"),
    ("dalseong-gun", "달성군"),
    ("eumseong-gun", "음성군"),
    ("eunpyeong-gu", "은평구"),
    ("gangneung-si", "강릉시"),
    ("gapyeong-gun", "가평군"),
    ("geochang-gun", "거창군"),
    ("geumcheon-gu", "금천구"),
    ("geumjeong-gu", "금정구"),
    ("goryeong-gun", "고령군"),
    ("gyeongsan-si", "경산시"),
    ("hapcheon-gun", "합천군"),
    ("heungdeok-gu", "흥덕구"),
    ("ilsandong-gu", "일산동구"),
    ("jeollabuk-do", "전라북도"),
    ("namyangju-si", "남양주시"),
    ("seodaemun-gu", "서대문구"),
    ("seongdong-gu", "성동구"),
    ("sunchang-gun", "순창군"),
    ("uijeongbu-si", "의정부시"),
    ("yangcheon-gu", "양천구"),
    ("yeongtong-gu", "영통구"),
    ("bupyeong-gu", "부평구"),
    ("changwon-si", "창원시"),
    ("cheongju-si", "청주시"),
    ("chilgok-gun", "칠곡군"),
    ("deogyang-gu", "덕양구"),
    ("gangdong-gu", "강동구"),
    ("ganghwa-gun", "강화군"),
    ("goheung-gun", "고흥군"),
    ("gwacheon-si", "과천시"),
    ("gwangsan-gu", "광산구"),
    ("gwonseon-gu", "권선구"),
    ("gyeonggi-do", "경기도"),
    ("gyeongju-si", "경주시"),
    ("haeundae-gu", "해운대구"),
    ("hwaseong-si", "화성시"),
    ("ilsanseo-gu", "일산서구"),
    ("jungnang-gu", "중랑구"),
    ("michuhol-gu", "미추홀구"),
    ("sangdang-gu", "상당구"),
    ("seongbuk-gu", "성북구"),
    ("seongju-gun", "성주군"),
    ("seongnam-si", "성남시"),
    ("seongsan-gu", "성산구"),
    ("suncheon-si", "순천시"),
    ("yecheon-gun", "예천군"),
    ("yeongam-gun", "영암군"),
    ("anseong-si", "안성시"),
    ("bucheon-si", "부천시"),
    ("bundang-gu", "분당구"),
    ("cheonan-si", "천안시"),
    ("chungju-si", "충주시"),
    ("daedeok-gu", "대덕구"),
    ("dangjin-si", "당진시"),
    ("deokjin-gu", "덕진구"),
    ("donghae-si", "동해시"),
    ("dongjak-gu", "동작구"),
    ("dongnae-gu", "동래구"),
    ("dongnam-gu", "동남구"),
    ("dongtan-gu", "동탄구"),
    ("eojin-dong", "어진동"),
    ("gangnam-gu", "강남구"),
    ("gangseo-gu", "강서구"),
    ("gangwon-do", "강원도"),
    ("giheung-gu", "기흥구"),
    ("gijang-gun", "기장군"),
    ("gyeyang-gu", "계양구"),
    ("jeonbuk-do", "전북도"),
    ("jungwon-gu", "중원구"),
    ("miryang-si", "밀양시"),
    ("namdong-gu", "남동구"),
    ("namhae-gun", "남해군"),
    ("pocheon-si", "포천시"),
    ("sangnok-gu", "상록구"),
    ("siheung-si", "시흥시"),
    ("suseong-gu", "수성구"),
    ("yanggu-gun", "양구군"),
    ("yangsan-si", "양산시"),
    ("yeongdo-gu", "영도구"),
    ("yeongju-si", "영주시"),
    ("yongsan-gu", "용산구"),
    ("yuseong-gu", "유성구"),
    ("andong-si", "안동시"),
    ("anyang-si", "안양시"),
    ("cheoin-gu", "처인구"),
    ("dalseo-gu", "달서구"),
    ("danwon-gu", "단원구"),
    ("dobong-gu", "도봉구"),
    ("dongan-gu", "동안구"),
    ("gimhae-si", "김해시"),
    ("gongju-si", "공주시"),
    ("goyang-si", "고양시"),
    ("gwanak-gu", "관악구"),
    ("haman-gun", "함안군"),
    ("icheon-si", "이천시"),
    ("jangan-gu", "장안구"),
    ("jeonju-si", "전주시"),
    ("jinhae-gu", "진해구"),
    ("jongno-gu", "종로구"),
    ("nonsan-si", "논산시"),
    ("paldal-gu", "팔달구"),
    ("pohang-si", "포항시"),
    ("sasang-gu", "사상구"),
    ("sejong-si", "세종시"),
    ("seobuk-gu", "서북구"),
    ("seocho-gu", "서초구"),
    ("seowon-gu", "서원구"),
    ("sokcho-si", "속초시"),
    ("songpa-gu", "송파구"),
    ("uljin-gun", "울진군"),
    ("wanju-gun", "완주군"),
    ("wansan-gu", "완산구"),
    ("yangju-si", "양주시"),
    ("yeonsu-gu", "연수구"),
    ("yongin-si", "용인시"),
    ("ansan-si", "안산시"),
    ("geoje-si", "거제시"),
    ("gimje-si", "김제시"),
    ("gimpo-si", "김포시"),
    ("gunpo-si", "군포시"),
    ("hanam-si", "하남시"),
    ("iksan-si", "익산시"),
    ("jinju-si", "진주시"),
    ("manan-gu", "만안구"),
    ("mokpo-si", "목포시"),
    ("suwon-si", "수원시"),
    ("ulju-gun", "울주군"),
    ("wonju-si", "원주시"),
    ("wonmi-gu", "원미구"),
    ("yeoju-si", "여주시"),
    ("yeosu-si", "여수시"),
    ("asan-si", "아산시"),
    ("daejeon", "대전"),
    ("dong-gu", "동구"),
    ("dongtan", "동탄"),
    ("gumi-si", "구미시"),
    ("guro-gu", "구로"),
    ("gwangju", "광주"),
    ("incheon", "인천"),
    ("jung-gu", "중구"),
    ("mapo-gu", "마포구"),
    ("naju-si", "나주시"),
    ("osan-si", "오산시"),
    ("paju-si", "파주시"),
    ("saha-gu", "사하구"),
    ("sosa-gu", "소사구"),
    ("suji-gu", "수지구"),
    ("buk-gu", "북구"),
    ("nam-gu", "남구"),
    ("pohang", "포항"),
    ("seo-gu", "서구"),
    ("yongin", "용인"),
    ("ansan", "안산"),
    ("busan", "부산"),
    ("daegu", "대구"),
    ("ulsan", "울산"),
]

ROAD_PARTS = ["로", "길", "대로", "거리"]
ADMIN_PARTS = ["동", "리", "면", "읍", "가", "시", "군", "구", "도"]

def detect_columns(df: pd.DataFrame) -> Tuple[str, str]:
    cols = list(df.columns)
    if len(cols) < 2:
        raise ValueError("엑셀에 최소 2개 컬럼이 필요합니다. (송장, 주소)")
    invoice_candidates = ["송장", "송장번호", "HAWB", "HAWB NO", "운송장", "invoice", "Invoice"]
    address_candidates = ["주소", "Address", "address", "收件地址", "地址"]
    invoice_col = next((c for c in invoice_candidates if c in df.columns), cols[0])
    address_col = next((c for c in address_candidates if c in df.columns), cols[1])
    return invoice_col, address_col

def compact_korean_spacing(text: str) -> str:
    s = str(text)

    s = re.sub(r"(\d+\.\d+)\s*([가-힣]+로)", r"\1\2", s)
    s = re.sub(r"(\d+)\s*([가-힣]+로)", r"\1\2", s)

    for suffix in ROAD_PARTS + ADMIN_PARTS:
        pattern = rf"((?:[가-힣]\s+){1,10}){suffix}\b"
        while True:
            m = re.search(pattern, s)
            if not m:
                break
            prefix = re.sub(r"\s+", "", m.group(1))
            s = s[:m.start()] + prefix + suffix + s[m.end():]

    s = re.sub(r"([가-힣]+)\s+([가-힣]+로)\b", r"\1\2", s)
    s = re.sub(r"([가-힣]+)\s+([가-힣]+길)\b", r"\1\2", s)
    s = re.sub(r"([가-힣]+)\s+([가-힣]+대로)\b", r"\1\2", s)

    s = re.sub(r"([가-힣]+)\s*로\s*(\d+)\s*([가-힣]?)\s*길", r"\1로\2\3길", s)
    s = re.sub(r"([가-힣]+)\s*로\s*(\d+)\s*번\s*길", r"\1로 \2번길", s)
    s = re.sub(r"([가-힣]+)\s*로\s*(\d+)\s*번\s*로", r"\1로 \2번로", s)
    s = re.sub(r"([가-힣]+)\s*길\s*(\d+)\s*번\s*길", r"\1길 \2번길", s)
    s = re.sub(r"([가-힣]+)\s*대로\s*(\d+)\s*번\s*길", r"\1대로 \2번길", s)

    s = re.sub(r"(\d+[가-힣]?(?:번길|번로|길))\s*(" + HOUSE_NUM + r")", r"\1 \2", s)
    s = re.sub(r"(" + ROAD_SUFFIX + r")\s*(" + HOUSE_NUM + r")", r"\1 \2", s)
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s

def is_apt_only(addr: str) -> bool:
    a = compact_korean_spacing(str(addr))
    has_road = re.search(GENERAL_ROAD_PATTERN, a) is not None
    has_jibun = re.search(JIBUN_PATTERN, a) is not None
    has_complex = any(re.search(p, a) is not None for p in COMPLEX_ROAD_PATTERNS)
    if has_road or has_jibun or has_complex:
        return False
    has_apt_name = APT_NAME_RE.search(a) is not None
    has_dongho = DONG_HO_RE.search(a) is not None or NUM_UNIT_RE.search(a) is not None
    return has_apt_name or has_dongho

def strip_room_only_tail(text: str) -> str:
    s = str(text).strip()
    s = re.sub(r"(\b" + HOUSE_NUM + r")[- ]\d{1,4}호\b", r"\1", s)
    s = re.sub(r"\b\d+\s*동\s*\d+\s*호\b", "", s)
    s = re.sub(r"\b\d+\s*호\b", "", s)
    s = re.sub(r"\s{2,}", " ", s).strip(" ,")
    return s

def extract_core_korean_address(addr: str) -> Optional[str]:
    a = compact_korean_spacing(str(addr))
    a = strip_room_only_tail(a)

    for pattern in COMPLEX_ROAD_PATTERNS:
        m = re.search(pattern, a)
        if m:
            return m.group(0).strip()

    road_match = re.search(GENERAL_ROAD_PATTERN, a)
    if road_match:
        return road_match.group(0).strip()

    jibun_match = re.search(JIBUN_PATTERN, a)
    if jibun_match:
        return jibun_match.group(0).strip()

    return None

def extract_region_hints(addr: str) -> List[str]:
    a = str(addr).lower()
    found = []
    matched_spans = []

    for eng, kor in REGION_HINTS_ORDERED:
        start = a.find(eng)
        if start != -1:
            end = start + len(eng)
            overlap = any(not (end <= s or start >= e) for s, e in matched_spans)
            if not overlap and kor not in found:
                found.append(kor)
                matched_spans.append((start, end))
    return found

def order_region_hints(hints: List[str]) -> List[str]:
    provinces, cities, districts, towns, dongs, others = [], [], [], [], [], []
    for h in hints:
        if h.endswith("도"):
            provinces.append(h)
        elif h in ["서울", "부산", "대구", "인천", "광주", "대전", "울산", "세종"] or h.endswith("시"):
            cities.append(h)
        elif h.endswith("구") or h.endswith("군"):
            districts.append(h)
        elif h.endswith("읍") or h.endswith("면"):
            towns.append(h)
        elif h.endswith("동"):
            dongs.append(h)
        else:
            others.append(h)

    ordered = provinces + cities + districts + towns + dongs + others
    out = []
    for x in ordered:
        if x not in out:
            out.append(x)
    return out

def extract_trailing_locality(addr: str) -> Optional[str]:
    s = compact_korean_spacing(str(addr))
    matches = re.findall(r"([가-힣A-Za-z0-9\.]+(?:동|리|읍|면))\b", s)
    for m in matches:
        if re.fullmatch(r"\d+동", m):
            continue
        if re.fullmatch(r"\d+리", m):
            continue
        return m
    return None

def build_region_enriched_query(raw_addr: str, core_addr: str) -> str:
    hints = order_region_hints(extract_region_hints(raw_addr))
    trailing = extract_trailing_locality(raw_addr)
    parts = []
    for h in hints:
        if h not in parts:
            parts.append(h)
    if trailing and trailing not in parts:
        parts.append(trailing)
    parts.append(core_addr)
    return " ".join(parts).strip()

def search_kakao_address(address: str, api_key: str, session: requests.Session) -> Dict[str, Any]:
    headers = {"Authorization": f"KakaoAK {api_key}"}
    params = {"query": str(address).strip()}
    resp = session.get(KAKAO_API_URL, headers=headers, params=params, timeout=10)
    resp.raise_for_status()
    return resp.json()

def parse_kakao_docs(docs: list) -> Tuple[str, str, str, str]:
    doc = docs[0]
    road_addr = doc.get("road_address") or {}
    jibun_addr = doc.get("address") or {}
    return road_addr.get("address_name", ""), jibun_addr.get("address_name", ""), doc.get("y", ""), doc.get("x", "")

def try_query(query: str, api_key: str, session: requests.Session) -> Dict[str, Any]:
    data = search_kakao_address(query, api_key, session)
    docs = data.get("documents", [])
    out = {"query": query, "count": len(docs), "road": "", "jibun": "", "lat": "", "lon": ""}
    if len(docs) == 1:
        road_name, jibun_name, lat, lon = parse_kakao_docs(docs)
        out.update({"road": road_name, "jibun": jibun_name, "lat": lat, "lon": lon})
    return out

def classify_kakao_result(raw_address: str, api_key: str, session: requests.Session) -> Dict[str, Any]:
    addr = str(raw_address).strip()
    result = {
        "원본주소": raw_address, "조회주소": "", "판정": "",
        "도로명주소": "", "지번주소": "", "위도": "", "경도": "", "오류사유": ""
    }

    if not addr or addr.lower() == "nan":
        result["판정"] = "오류"; result["오류사유"] = "빈 주소"; return result
    if is_apt_only(addr):
        result["판정"] = "오류"; result["오류사유"] = "아파트명/동호만 존재"; return result

    try:
        r1 = try_query(addr, api_key, session)
        if r1["count"] == 1:
            result.update({"조회주소": r1["query"], "판정": "정상", "도로명주소": r1["road"], "지번주소": r1["jibun"], "위도": r1["lat"], "경도": r1["lon"]})
            return result

        core_addr = extract_core_korean_address(addr)
        if not core_addr:
            result["판정"] = "오류"; result["오류사유"] = "핵심 한국 주소 추출 실패"; return result

        region_query = build_region_enriched_query(addr, core_addr)
        r2 = try_query(region_query, api_key, session) if region_query != core_addr else {"count": 0}
        if r2["count"] == 1:
            result.update({"조회주소": r2["query"], "판정": "정상", "도로명주소": r2["road"], "지번주소": r2["jibun"], "위도": r2["lat"], "경도": r2["lon"]})
            return result

        r3 = try_query(core_addr, api_key, session)
        if r3["count"] == 1:
            result.update({"조회주소": r3["query"], "판정": "정상", "도로명주소": r3["road"], "지번주소": r3["jibun"], "위도": r3["lat"], "경도": r3["lon"]})
            return result

        compact_core = compact_korean_spacing(core_addr)
        region_query2 = build_region_enriched_query(addr, compact_core)
        r4 = try_query(region_query2, api_key, session) if region_query2 != compact_core else {"count": 0}
        if r4["count"] == 1:
            result.update({"조회주소": r4["query"], "판정": "정상", "도로명주소": r4["road"], "지번주소": r4["jibun"], "위도": r4["lat"], "경도": r4["lon"]})
            return result

        r5 = try_query(compact_core, api_key, session) if compact_core != core_addr else {"count": 0}
        if r5["count"] == 1:
            result.update({"조회주소": r5["query"], "판정": "정상", "도로명주소": r5["road"], "지번주소": r5["jibun"], "위도": r5["lat"], "경도": r5["lon"]})
            return result

        result["판정"] = "오류"
        result["조회주소"] = region_query2 if region_query2 else compact_core
        max_count = max(r1["count"], r2["count"], r3["count"], r4["count"], r5["count"])
        result["오류사유"] = "조회 실패" if max_count == 0 else f"후보 다수(최대 {max_count}건)"
        return result

    except requests.HTTPError as e:
        status_code = e.response.status_code if e.response is not None else ""
        result["판정"] = "오류"
        if status_code == 401:
            result["오류사유"] = "API 키 인증 오류"
        elif status_code == 429:
            result["오류사유"] = "호출 한도 초과"
        else:
            result["오류사유"] = f"HTTP 오류 {status_code}"
        return result
    except requests.RequestException:
        result["판정"] = "오류"; result["오류사유"] = "네트워크 오류"; return result
    except Exception as e:
        result["판정"] = "오류"; result["오류사유"] = f"처리 오류: {e}"; return result



def normalize_addr_for_detail_match(text: str) -> str:
    """상세주소 분리를 위해 주소 문자열의 공백/쉼표/괄호/시도 표기를 느슨하게 정리합니다."""
    s = compact_korean_spacing(str(text or ""))
    s = re.sub(r"[\[\]\(\)（）,，]", " ", s)
    replacements = {
        "서울특별시": "서울", "부산광역시": "부산", "대구광역시": "대구", "인천광역시": "인천",
        "광주광역시": "광주", "대전광역시": "대전", "울산광역시": "울산", "세종특별자치시": "세종",
        "경기도": "경기", "강원특별자치도": "강원", "강원도": "강원",
        "충청북도": "충북", "충청남도": "충남", "전라북도": "전북", "전북특별자치도": "전북",
        "전라남도": "전남", "경상북도": "경북", "경상남도": "경남", "제주특별자치도": "제주",
    }
    for old, new in replacements.items():
        s = s.replace(old, new)
    s = re.sub(r"\b(서울|부산|대구|인천|광주|대전|울산|세종)시\b", r"\1", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _addr_compact(text: str) -> str:
    """공백/구두점 차이를 제거한 비교용 문자열."""
    s = normalize_addr_for_detail_match(text)
    return re.sub(r"[\s,，./·\-~～]", "", s)


def _normalize_raw_preserve_detail(text: str) -> str:
    """원본 상세주소를 최대한 보존한 상태로 비교용 기본 정리만 합니다."""
    s = compact_korean_spacing(str(text or ""))
    replacements = {
        "서울특별시": "서울", "부산광역시": "부산", "대구광역시": "대구", "인천광역시": "인천",
        "광주광역시": "광주", "대전광역시": "대전", "울산광역시": "울산", "세종특별자치시": "세종",
        "경기도": "경기", "강원특별자치도": "강원", "강원도": "강원",
        "충청북도": "충북", "충청남도": "충남", "전라북도": "전북", "전북특별자치도": "전북",
        "전라남도": "전남", "경상북도": "경북", "경상남도": "경남", "제주특별자치도": "제주",
    }
    for old, new in replacements.items():
        s = s.replace(old, new)
    s = re.sub(r"\b(서울|부산|대구|인천|광주|대전|울산|세종)시\b", r"\1", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _build_preserve_compact_index_map(text: str):
    """원본 위치 보존용 compact 문자열과 원문 인덱스 매핑을 만듭니다."""
    norm = _normalize_raw_preserve_detail(text)
    compact_chars = []
    index_map = []
    for i, ch in enumerate(norm):
        if re.match(r"[\s,，./·\-~～\[\]\(\)（）]", ch):
            continue
        compact_chars.append(ch)
        index_map.append(i)
    return norm, "".join(compact_chars), index_map


def _sanitize_detail_text(text: str) -> str:
    """상세주소/최종주소에서 제거할 특수표시 정리."""
    s = str(text or "")
    # 요청 특수문자 제거: 알림벨, 별표류, 역슬래시
    s = re.sub(r"[🔔★☆\\]", "", s)
    # 빈 괄호/중복 공백 정리
    s = re.sub(r"[\(（]\s*[\)）]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _clean_detail_suffix(detail: str) -> str:
    detail = str(detail or "")
    # 상세주소 앞에 붙은 구분자 정리. 괄호는 상세주소일 수 있어 보존합니다.
    detail = re.sub(r"^[\s,，./\-~～]+", "", detail)
    # 지번주소 뒤에 바로 붙는 "번지"는 상세주소가 아니라 기본주소 보조표기라 제거
    detail = re.sub(r"^(?:번지|번지내|지번)\s*", "", detail)
    detail = _sanitize_detail_text(detail)
    detail = re.sub(r"\s+", " ", detail).strip(" ,，-/~～")
    # API 주소가 원본의 괄호 안 주소와 매칭된 경우 suffix가 ')' 하나만 남는 문제 방지
    if detail in {"(", ")", "（", "）", "[]", "{}"}:
        return ""
    return detail


def _match_suffix_after_candidate(raw: str, candidate: str):
    """candidate가 원본 안에서 끝나는 위치를 찾습니다.

    핵심 보정:
    - 공백/쉼표/하이픈 차이는 무시해서 `도림로12`와 `도림로 12`를 같은 주소로 봅니다.
    - 원본 안에 같은 후보가 여러 번 있으면, 뒤쪽 상세주소가 잘리지 않도록 "가장 앞쪽의 정상 매칭"보다
      실제 기본주소가 끝나는 위치가 자연스러운 매칭을 사용합니다.
    - 매칭 뒤에 `번지`가 있으면 기본주소 부가표기라 보고 상세주소 앞에서 제거합니다.
    """
    raw_norm, raw_compact, idx_map = _build_preserve_compact_index_map(raw)
    cand_compact = _addr_compact(candidate)
    if not raw_norm or not raw_compact or not cand_compact:
        return False, ""

    positions = [m.start() for m in re.finditer(re.escape(cand_compact), raw_compact)]
    if not positions:
        return False, ""

    # 같은 후보가 여러 번 잡히면 더 긴 suffix를 남기는 앞쪽 매칭을 우선합니다.
    # 예: 원본 끝 괄호의 행정동 `(명동2가)` 같은 짧은 매칭이 뒤에서 잡혀 상세주소가 사라지는 문제 방지.
    best_suffix = None
    for pos in positions:
        end_compact_idx = pos + len(cand_compact) - 1
        if end_compact_idx >= len(idx_map):
            continue
        end_norm_idx = idx_map[end_compact_idx] + 1
        suffix = _clean_detail_suffix(raw_norm[end_norm_idx:])
        if best_suffix is None or len(suffix) > len(best_suffix):
            best_suffix = suffix

    return True, (best_suffix or "")



def _road_or_jibun_tail_candidates(text: str) -> list:
    """API 주소에서 도로명+건물번호 또는 동/리/가+지번 꼬리 후보를 뽑습니다."""
    # compact_korean_spacing 과정에서 `동탄구 동탄오산로`가 붙어
    # `동탄구동탄오산로 86-8`처럼 잘못 긴 후보가 생길 수 있으므로,
    # 먼저 원문 공백 기준으로 마지막 도로명/지번 꼬리를 따로 뽑습니다.
    raw_s = str(text or "")
    raw_s = re.sub(r"[\[\]\(\)（）,，]", " ", raw_s)
    raw_s = re.sub(r"\s+", " ", raw_s).strip()
    candidates = []

    raw_road_patterns = [
        r"[가-힣A-Za-z0-9\.]+(?:대로|로|길|거리)\s*\d+[가-힣]?(?:번길|번로|길)\s*\d{1,5}(?:[-~～]\d{1,5})?",
        r"[가-힣A-Za-z0-9\.]+(?:대로|로|길|거리)\d+[가-힣]?(?:번길|번로|길)\s*\d{1,5}(?:[-~～]\d{1,5})?",
        r"[가-힣A-Za-z0-9\.]+(?:대로|로|길|거리)\s*\d{1,5}(?:[-~～]\d{1,5})?",
    ]
    for pat in raw_road_patterns:
        for m in re.finditer(pat, raw_s, flags=re.I):
            cand = re.sub(r"\s+", " ", m.group(0)).strip()
            if cand and cand not in candidates:
                candidates.append(cand)

    raw_jibun_patterns = [
        r"[가-힣A-Za-z0-9\.]+(?:동|리|가)\s*\d{1,5}(?:[-~～]\d{1,5})?",
    ]
    for pat in raw_jibun_patterns:
        for m in re.finditer(pat, raw_s, flags=re.I):
            cand = re.sub(r"\s+", " ", m.group(0)).strip()
            if cand and cand not in candidates:
                candidates.append(cand)

    s = normalize_addr_for_detail_match(text)

    road_patterns = [
        # 동탄오산로 86-8 / 명동10길 19-10 / 원적로488번길 5-1
        r"[가-힣A-Za-z][0-9가-힣A-Za-z\.]{1,}(?:대로|로|길|거리)\s*\d+[가-힣]?(?:번길|번로|길)\s*\d{1,5}(?:[-~～]\d{1,5})?",
        r"[가-힣A-Za-z][0-9가-힣A-Za-z\.]{1,}\d+(?:번길|번로|길)\s*\d{1,5}(?:[-~～]\d{1,5})?",
        r"[가-힣A-Za-z][0-9가-힣A-Za-z\.]{1,}(?:대로|로|길|거리)\s*\d{1,5}(?:[-~～]\d{1,5})?",
    ]
    for pat in road_patterns:
        for m in re.finditer(pat, s, flags=re.I):
            cand = re.sub(r"\s+", " ", m.group(0)).strip()
            if cand and cand not in candidates:
                candidates.append(cand)

    jibun_patterns = [
        r"[가-힣A-Za-z0-9\.]+(?:동|리|가)\s*\d{1,5}(?:[-~～]\d{1,5})?",
        r"[가-힣A-Za-z][가-힣A-Za-z0-9\.]+(?:읍|면)\s+[가-힣A-Za-z0-9\.]+(?:리|동)?\s*\d{1,5}(?:[-~～]\d{1,5})?",
    ]
    for pat in jibun_patterns:
        for m in re.finditer(pat, s, flags=re.I):
            cand = re.sub(r"\s+", " ", m.group(0)).strip()
            if cand and cand not in candidates:
                candidates.append(cand)

    # 긴 후보가 짧은 핵심 후보로 끝나는 경우(예: 동탄구동탄오산로 86-8 vs 동탄오산로 86-8)는
    # 행정구역명이 붙은 긴 후보를 버리고, 실제 도로명/지번 꼬리만 남깁니다.
    filtered = []
    compact_pairs = [(c, _addr_compact(c)) for c in candidates]
    for cand, cc in compact_pairs:
        # `명동10길 19-10` 안에서 `10길 19-10`처럼 숫자로 시작하는 짧은 후보가 따로 잡히면 제거
        if re.match(r"^\d", cc or "") and any(cand != other and oc and oc.endswith(cc) and len(oc) > len(cc) for other, oc in compact_pairs):
            continue
        # `명동10길 19-10` 안에서 `명동10`이 지번처럼 오인식되는 경우 제거
        if any(cand != other and oc and oc.startswith(cc) and len(oc) > len(cc) and re.search(r"(?:대로|로|길|거리)", other) for other, oc in compact_pairs):
            continue
        if any(cand != other and oc and cc.endswith(oc) and len(cc) > len(oc) for other, oc in compact_pairs):
            continue
        filtered.append(cand)
    return filtered

def _last_road_jibun_tail(text: str) -> str:
    """API 주소의 마지막 핵심 주소(도로명+번호 또는 동/리/가+지번)를 반환합니다."""
    candidates = _road_or_jibun_tail_candidates(text)
    if not candidates:
        return ""
    # 실제 주소 끝부분을 우선: API 주소에서 가장 뒤에 나오는 후보, 동률이면 긴 후보
    norm = normalize_addr_for_detail_match(text)
    candidates.sort(key=lambda c: (norm.rfind(c), len(_addr_compact(c))), reverse=True)
    return candidates[0]


def _candidate_list_for_base(base: str) -> list:
    base = str(base or "").strip()
    if not base:
        return []

    candidates = []

    # 1순위: 원본에 가장 잘 들어맞는 마지막 핵심 주소
    # 예: `경기 안산시 상록구 이동 612-22` -> `이동 612-22`
    #     `서울 영등포구 도림로 12` -> `도림로 12`
    #     `경기 화성시 동탄오산로 86-8` -> `동탄오산로 86-8`
    tail = _last_road_jibun_tail(base)
    if tail:
        candidates.append(tail)

    # 2순위: API 전체 기본주소. 원본에도 시/군/구가 같이 있을 때 정확도를 올림.
    candidates.append(base)

    # 3순위: 후보 전체. 행정구역명이 붙은 긴 후보보다 실제 핵심 꼬리 후보를 우선합니다.
    all_tails = _road_or_jibun_tail_candidates(base)
    compact_pairs = [(c, _addr_compact(c)) for c in all_tails]
    for cand, cc in compact_pairs:
        # `명동10길 19-10` 안에서 `10길 19-10`처럼 숫자로 시작하는 짧은 후보가 따로 잡히면 제거
        if re.match(r"^\d", cc or "") and any(cand != other and oc and oc.endswith(cc) and len(oc) > len(cc) for other, oc in compact_pairs):
            continue
        # `명동10길 19-10` 안에서 `명동10`이 지번처럼 오인식되는 경우 제거
        if any(cand != other and oc and oc.startswith(cc) and len(oc) > len(cc) and re.search(r"(?:대로|로|길|거리)", other) for other, oc in compact_pairs):
            continue
        if any(cand != other and oc and cc.endswith(oc) and len(cc) > len(oc) for other, oc in compact_pairs):
            continue
        candidates.append(cand)

    # 중복 제거 + 너무 짧은 후보 제거. `가`, `동` 같은 1글자 오인식 방지.
    deduped = []
    for cand in candidates:
        cc = _addr_compact(cand)
        if len(cc) < 4:
            continue
        if cand not in deduped:
            deduped.append(cand)

    # 기본적으로 마지막 핵심주소를 먼저 보되, 같은 후보군 안에서는 긴 후보 우선.
    return sorted(deduped, key=lambda x: (x != tail, len(_addr_compact(x))), reverse=False)


def _score_detail_match(kind: str, candidate: str, suffix: str, raw_addr: str) -> tuple:
    """도로명/지번 후보 중 상세주소포함 기준으로 쓸 주소를 고르기 위한 점수."""
    cand_len = len(_addr_compact(candidate))
    suffix_clean = _clean_detail_suffix(suffix)
    suffix_len = len(suffix_clean)

    # 상세주소가 너무 길게 남으면 도로명/지번 기준이 원본과 어긋났을 가능성이 큼.
    # 예: 원본은 지번인데 도로명 기준으로 붙여서 `장충동1가 35-13 306호`가 상세주소가 되는 경우.
    bad_leftover = bool(re.search(r"[가-힣0-9A-Za-z\.]+(?:동|리|가)\s*\d{1,5}(?:[-~～]\d{1,5})?", suffix_clean))

    # 원본 주소 형태 힌트: 원본에 지번 패턴이 강하면 지번 후보 우선, 도로명 패턴이 강하면 도로명 후보 우선.
    raw_norm = normalize_addr_for_detail_match(raw_addr)
    raw_looks_jibun = bool(re.search(r"[가-힣0-9A-Za-z\.]+(?:동|리|가)\s*\d{1,5}(?:[-~～]\d{1,5})?", raw_norm))
    raw_looks_road = bool(re.search(r"[가-힣0-9A-Za-z\.]+(?:대로|로|길|거리)\s*\d{1,5}(?:[-~～]\d{1,5})?", raw_norm))

    kind_bonus = 0
    if kind == "지번" and raw_looks_jibun:
        kind_bonus += 1000
    if kind == "도로명" and raw_looks_road:
        kind_bonus += 1000

    # ')' 하나처럼 의미 없는 잔여문자가 남은 후보보다, 실제 상세주소가 길게 남는 후보를 우선한다.
    detail_bonus = min(suffix_len, 60) * 10 if suffix_len > 0 else 0
    bad_penalty = -2000 if bad_leftover else 0

    return (kind_bonus + detail_bonus + bad_penalty, suffix_len, cand_len)


def extract_detail_with_base(raw_addr: str, road_addr: str = "", jibun_addr: str = ""):
    """원본주소와 더 잘 맞는 API 기본주소(도로명/지번)를 선택하고 상세주소를 추출합니다.

    - API 도로명주소/지번주소는 정상 시트처럼 둘 다 유지합니다.
    - `상세주소포함`은 원본주소에서 실제로 매칭되는 쪽(도로명 또는 지번)을 골라 하나만 만듭니다.
    - 기본주소의 마지막 핵심주소 뒤에 따라오는 내용은 그대로 상세주소로 붙입니다.
      예: `이동 612-22 번지 지하` -> `경기 안산시 상록구 이동 612-22 지하`
          `도림로12 2층` -> `서울 ... 도림로 12 2층`
          `동탄오산로 86-8 오산동 ...` -> `경기 ... 동탄오산로 86-8 오산동 ...`
    """
    raw = str(raw_addr or "").strip()
    road = str(road_addr or "").strip()
    jibun = str(jibun_addr or "").strip()

    matched_results = []
    for kind, base in [("도로명", road), ("지번", jibun)]:
        if not base:
            continue
        for cand in _candidate_list_for_base(base):
            matched, suffix = _match_suffix_after_candidate(raw, cand)
            if matched:
                matched_results.append({
                    "kind": kind,
                    "base": base,
                    "candidate": cand,
                    "suffix": suffix,
                    "score": _score_detail_match(kind, cand, suffix, raw),
                })

    if matched_results:
        matched_results.sort(key=lambda x: x["score"], reverse=True)
        best = matched_results[0]
        return best["base"], best["suffix"], best["kind"]

    # 매칭 실패 시 기존 정책: 도로명 우선, 없으면 지번
    return road or jibun, "", "도로명" if road else "지번"

def extract_detail_address(raw_addr: str, road_addr: str = "", jibun_addr: str = "", query_addr: str = "") -> str:
    """호환용: 상세주소만 반환합니다."""
    _, detail, _ = extract_detail_with_base(raw_addr, road_addr, jibun_addr)
    return _clean_detail_suffix(detail)


def append_detail_to_base(base_addr: str, detail_addr: str) -> str:
    base = str(base_addr or "").strip()
    detail = _clean_detail_suffix(detail_addr)
    if not base:
        return ""
    return _sanitize_detail_text(f"{base} {detail}".strip() if detail else base)


def build_detail_address_df(df_ok: pd.DataFrame) -> pd.DataFrame:
    """정상 결과 기준으로 상세주소가 포함된 주소 시트 1개를 생성합니다.

    정상 시트처럼 API 도로명주소/지번주소는 둘 다 보여주고,
    상세주소포함 컬럼만 원본주소와 더 잘 맞는 기준주소 1개를 골라 상세주소를 붙입니다.
    예: 원본이 지번주소이면 API 지번주소 기준으로 306호를 붙이고, 도로명주소 뒤에 지번 일부가 붙는 문제를 방지합니다.
    """
    if df_ok is None or df_ok.empty:
        return pd.DataFrame(columns=[
            "송장", "원본주소", "조회주소", "도로명주소", "지번주소", "API기본주소", "상세주소", "상세주소포함", "위도", "경도"
        ])

    out = df_ok.copy()

    base_detail = out.apply(
        lambda r: extract_detail_with_base(
            r.get("원본주소", ""),
            r.get("도로명주소", ""),
            r.get("지번주소", ""),
        ),
        axis=1,
    )
    out["API기본주소"] = base_detail.apply(lambda x: x[0])
    out["상세주소"] = base_detail.apply(lambda x: x[1])
    out["상세주소포함"] = out.apply(lambda r: append_detail_to_base(r.get("API기본주소", ""), r.get("상세주소", "")), axis=1)

    cols = ["송장", "원본주소", "조회주소", "도로명주소", "지번주소", "API기본주소", "상세주소", "상세주소포함", "위도", "경도"]
    return out[[c for c in cols if c in out.columns]]

def to_excel_bytes(df_all: pd.DataFrame, df_ok: pd.DataFrame, df_err: pd.DataFrame, df_detail: Optional[pd.DataFrame] = None) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df_all.to_excel(writer, index=False, sheet_name="전체결과")
        df_ok.to_excel(writer, index=False, sheet_name="정상")
        df_err.to_excel(writer, index=False, sheet_name="오류")
        if df_detail is None:
            df_detail = build_detail_address_df(df_ok)
        df_detail.to_excel(writer, index=False, sheet_name="상세주소포함")
        workbook = writer.book
        for sheet_name, df in {"전체결과": df_all, "정상": df_ok, "오류": df_err, "상세주소포함": df_detail}.items():
            ws = writer.sheets[sheet_name]
            ws.freeze_panes(1, 0)
            header_fmt = workbook.add_format({"bold": True, "bg_color": "#D9EAF7", "border": 1, "align": "center", "valign": "vcenter"})
            for i, col in enumerate(df.columns):
                width = max(12, min(52, max(len(str(col)), int(df[col].astype(str).str.len().quantile(0.9)) if len(df) else 12) + 2))
                ws.set_column(i, i, width)
                ws.write(0, i, col, header_fmt)
    return output.getvalue()

st.set_page_config(
    page_title="TY LOGIS 업무 자동화 시스템",
    layout="wide",
    initial_sidebar_state="expanded"
)

import base64

def img_to_base64(path):
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()

if "login" not in st.session_state:
    st.session_state.login = False
if "user" not in st.session_state:
    st.session_state.user = ""
if "page" not in st.session_state:
    st.session_state.page = "main"
if "dash_schedule" not in st.session_state:
    st.session_state.dash_schedule = None   # 작업일정 이미지 bytes
if "dash_ty" not in st.session_state:
    st.session_state.dash_ty = None         # TY 현황표 이미지 bytes
if "dash_ky" not in st.session_state:
    st.session_state.dash_ky = None         # KY 현황표 이미지 bytes
if "dash_schedule_name" not in st.session_state:
    st.session_state.dash_schedule_name = ""
if "dash_ty_name" not in st.session_state:
    st.session_state.dash_ty_name = ""
if "dash_ky_name" not in st.session_state:
    st.session_state.dash_ky_name = ""

# ── GitHub 이미지 저장/로드 함수 ──
import base64 as _b64

def _gh_headers():
    token = st.secrets.get("GITHUB_TOKEN", "")
    return {"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"}

def _gh_repo():
    return st.secrets.get("GITHUB_REPO", "")

def github_upload_image(image_bytes, filename):
    """이미지를 GitHub 레포 dashboard/ 폴더에 업로드."""
    import requests as _req
    repo = _gh_repo()
    if not repo:
        return False
    url = f"https://api.github.com/repos/{repo}/contents/dashboard/{filename}"
    # 기존 파일 SHA 조회 (업데이트 시 필요)
    r = _req.get(url, headers=_gh_headers())
    sha = r.json().get("sha") if r.status_code == 200 else None
    data = {"message": f"update {filename}", "content": _b64.b64encode(image_bytes).decode()}
    if sha:
        data["sha"] = sha
    r2 = _req.put(url, headers=_gh_headers(), json=data)
    return r2.status_code in (200, 201)

def github_load_image(filename):
    """GitHub에서 이미지 로드."""
    import requests as _req
    repo = _gh_repo()
    if not repo:
        return None
    url = f"https://api.github.com/repos/{repo}/contents/dashboard/{filename}"
    r = _req.get(url, headers=_gh_headers())
    if r.status_code == 200:
        return _b64.b64decode(r.json()["content"])
    return None

def github_delete_image(filename):
    """GitHub에서 이미지 삭제."""
    import requests as _req
    repo = _gh_repo()
    if not repo:
        return False
    url = f"https://api.github.com/repos/{repo}/contents/dashboard/{filename}"
    r = _req.get(url, headers=_gh_headers())
    if r.status_code != 200:
        return False
    sha = r.json().get("sha")
    data = {"message": f"delete {filename}", "sha": sha}
    r2 = _req.delete(url, headers=_gh_headers(), json=data)
    return r2.status_code == 200

def load_dashboard_from_github():
    """앱 시작 시 GitHub에서 대시보드 이미지 불러오기."""
    if st.session_state.get("_gh_loaded"):
        return
    for key, fname in [("dash_schedule", "schedule.png"), ("dash_ty", "ty.png"), ("dash_ky", "ky.png")]:
        if not st.session_state.get(key):
            img = github_load_image(fname)
            if img:
                st.session_state[key] = img
    st.session_state["_gh_loaded"] = True

# ── 사용자 계정 관리 (st.secrets 우선, 없으면 session_state 내장) ──
def _default_users():
    return {
        "admin": {"password": "admin2024!", "role": "admin"},
        "ty":    {"password": "ty1234",     "role": "user"},
        "yst":   {"password": "yst1234",    "role": "user"},
    }

def load_users():
    """secrets에 [users] 섹션이 있으면 그걸 쓰고, 없으면 session_state 내장 계정 사용."""
    try:
        raw = dict(st.secrets.get("users", {}))
        if raw:
            # secrets 형식: admin = "admin2024!|admin"  또는  admin = "admin2024!"
            users = {}
            for k, v in raw.items():
                parts = str(v).split("|")
                users[k] = {"password": parts[0], "role": parts[1] if len(parts) > 1 else "user"}
            return users
    except Exception:
        pass
    if "users_db" not in st.session_state:
        st.session_state.users_db = _default_users()
    return st.session_state.users_db

def save_users(users_dict):
    """session_state에 저장 (secrets 미사용 환경용)."""
    st.session_state.users_db = users_dict

def check_login(user, pw):
    users = load_users()
    u = users.get(user)
    if u and u["password"] == pw:
        return True, u.get("role", "user")
    return False, None

def is_admin():
    return st.session_state.get("role") == "admin"

logo_html = '<img src="data:image/png;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAFWAnMDASIAAhEBAxEB/8QAHQABAAICAwEBAAAAAAAAAAAAAAMHBggEBQkCAf/EAGEQAAEDAgMEBQQJDA8EBwkAAAABAgMEBQYRMQcSIUEIEzJRYSJxgcEUI0KRobGy0dIXNTY3UlZic3SUlbQVFiYzNENFRlRydYKSk7MkU4XwJVV2g4Sj0wlEY6LDxOHi8f/EABoBAQACAwEAAAAAAAAAAAAAAAABAgMEBQb/xAA1EQEAAgIBAwIDBQYGAwAAAAAAAQIDEQQSITEFYRNBcRRCUYGhIjKRwdHwFSMzNFKiseHx/9oADAMBAAIRAxEAPwDcsAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAApbbjtmZhmWXD2GHRT3hvk1FSqI6OlX7lE0c/wXgnPNeCBaWJsT4ew1TJUX68UdvY5FViTSIj35a7re070Ipgs23zZrHKrG3Wrkai9ttFJl8KIvwGot3uNfda+SvudZPWVUq5ySzPVznelTgP8AWRtOm+WDtoODsXyLDYL7TVVQiKq07kdHLkmqox6IqoneiKhlB5zUNbV26uirqCplpaqB6PilicrXMcnNFQ272DbYqTG0DLJe3RUmIYmcOKNZWoicXMTk/vZ6U4Zo1smFvAAlAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACtukLjG54RwWjrTBOlVXvWBtW1vkUyZcVz5PVODfSvLJdO5XOe9XvcrnOXNVVc1VT0Du1uobtbZ7bcqWKqpKhm5LFImbXJ/zz5Go22/ZbWYHuC3C3tkqbBO/KKVeLqdy/wAW/wBTufnCYVa7kRP9ZK7kRP8AWQlC/mfMM01NUMqKeWSGaJyPjkjcrXMcnFFRU4oqLzPp/MhfzIS216P+2yHFLYsN4qnhp741EbT1C5MZW8su5JPBODuXcXieaqOcyRr2OVrmrmiouSoptP0d9uLbysGE8Z1TI7lkjKKvkXJKnuZIq6SdzvdadrtTtWYbBgAlAAAAAAAAAAABo7jDEeIYsV3eOO+3RjG106Na2rkRERJHZIiZm8RoTjX7L7z+Xz/6jgIX4mxJ98N20/pknzkL8T4l++G76f02T5zr3+ogf6iEuwfijE3H90V3/PZPnIn4pxPx/dHePz2T5zrX8yF/MhMOyfirFHH90l4/PpPpEL8V4p4/ulvP59L9I6x/MhfzC0OzfizFOa/ulvP59L9Iifi3FX3zXrX+nS/SOrfqpC/1iUu0fi7Ff3z3rX+ny/SIX4vxZ98971/p8v0jq3+sgf6yJTDtZMYYt++i+a/0+X6RA/GOLvvpvmv/AFhL9I6qT1kD/WVTDt3Yxxf99V9/SEv0iF2MsX/fXff0hL9I6h3IhdyCVydHjE+JazEWI2VmIbvUNjwzXyMbLWyPRr0YmTkzXgqd5y34mxJl9kF2/PJPnMe6Nn2S4m/7K3H5DTlv0Lx4Ysnl2EmJ8ScP3Q3b89k+cififEvD90N3/PZPnOul5EMnIlV2MmKMTb32RXfT+myfORPxTifP7I7x+eyfOdbKnHMifqEw7F+KsUby/ukvH59J9IiXFeKM1/dJefz6T6R1j+2pCuqkSlt10M7lcbng6+S3Kvq617LijWuqJnSK1OrbwRXKuSF7mvvQh+wq/wD9pp/ptNgiYVnyGoXTUvt7tm1G3QW28XGihdZInrHT1L42q7r50zyaqcckTj4G3ppj05/tsWz+wYv9eoEpr5Uy7F+LN1f3UXv8/l+kRPxhi3dX91N8/P5fpHUv7KkL+wpSWR2z8Y4uy+ym+fpCX6RG7GOL91f3VX39IS/SOnfoRv7KkLO2djLF/wB9d9/SEv0i+egtiC/XXa3dKe6Xu5V0LbDM9sdTVPkajvZFOmaI5VTPJV4+KmtLuRsH0Avtx3b/ALPzfrFOTHlFvEt4wAZGAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADg3q509rpVmmXeevCONF4uX5vEpkyVx1m1p1EJrWbTqH5fLpT2qkWaZd568I40Xi9fm8SurheblV1KzSVcrFzza1j1a1vmRD5uldUXCpfU1L95y6ImjU7k8DhO1PI871G/Jtqvasf33dnBxq4o795ZdhPFDusShuku8irlHO5dPBy+v3zLbnQ0dzt89vuFNHU0tQxWSxSJm17V5KU+7mZdgzE/VbltuUnkdmGZy9n8F3h3Ly+Lf8ATPU/GLNP0n+rX5XF+/Rrvty2TVmB6lbrbOsq7BM/JJFTN9K5V4Mf3pyR3PRclyzqd/rPRSupKavopqKtgjqKadixyxSNRzXtVMlRUXVDUPbvsjrMF1cl5s0ctTh2V+e92n0blXgx/NW58Ed6F45K70LnKhfzIX8yZ/MhfzIWRO1IXciZ2pC7kJGzfR0269d7Ewfjaq9t4RUNzlf2+TY5VXnyR/Pgi8eK7LHmQ7Q2P6N+3VaNYcI45r1Wm7FDc53/AL13RyuX3Pc5dNF4ZKiJRMNqAEVFRFRc0UFlQAAAAAAAA0Jxr9l95/L5/wDUcb7GhONfsvvP5fP/AKjgOjf6iB/qJ3+oyjY9h79s+0qy2p7N+BZ0mqEyzTqo/Lci+dE3fOqEJZ3QdGvFNZb6eqkvVrpnzRMkdDI2TejVURVauSapoSO6MGJ1/nHZ/wDDJ8xtWCdG2pdf0ZMTU1FPVOxFZ3JFG6RURsma5Jn9z4FAP5npNf8A6xXD8lk+Sp5sv5kLVlC/VSF/rJn6qQv9ZWVndbPsLVeNcZ2/DFDUw01RXOejJZs9xu6xz1zy46NUuR3RMxYv857J/gl+iYP0Xft9YZ/GT/q8pvyTEbVmZhp07okYtX+dFj/wS/RMN2udH3E2z3B8mJqy62640sU7I5m0zXo6NHrkj13kRMt7dT+8hvudBtFw5Di7At6w3PuZXCjkhY5yZoyRUzY/+65Gu9A6YIvLy+dyIXcjlVkE1LUyU1RG6KaJ7mSMcmStci5Ki+KKcV3Ioywu/oPwQ1O2qSnqIY5oZbRUskjkajmvaqsRUVF4Ki9xtviLY3s9vTXq6xsoJXJkktC9Yd3zNTyPfaamdBf7eP8Awqo+NhvmXr4Ysnlqxtf2FPwxYJ7/AIeuNRcKSl8qop54062OPm9HNyRyJz4Jkma8ijZOR6LzRRzQvhmjbJHI1WvY5M0ci8FRU5oaN7aMGSYHxxVWtjXewJfb6F68d6Fy8Ez5q1c2r5s+ZKsMIevHIgfqSydpPMRPTmEwktSW915pG3Z07beszEqnQZdY2PNN5W5oqZomeXA3CwxsB2VxUUFW2jqr1HK1s0U9TWuVHtXJWrlHuNVFTwNMn9tTazogY+/ZSxzYIuU+dZbWrLQq5eMlOq8Wedjl/wALkROyCV14cw9YsOUbqSw2iitsD3bz2U0KM33ZZZuy1XLmp2gBKoY9i3BGEMWKj8R4dt1ylSPqmzTQp1rWZqu6j08pEzVVyReamQgDWLb7sS2U4RwLccTwOulplhZuUlNDV77J53cGMylRzteK5Lwajl5Go7+wpfHTE2hftpxx+1m3T71qsTnRu3V4S1S8JHeKNy3E7lR/JSh39hTHLLVC/Q7vAGEbzjnFdFhqxQpJV1T+L3cGQsTi6R68mtTj3rwRM1VEOkfobu9DHZv+1fBS4vucG7dr7G10KORM4aTVied/B6+G5oqKIja1p1Dj4O6JmAra2KXEdxud/nRPLYj/AGNA5fBrPLT/ABlyYOwLg7BzFbhjDdttb3R9W+WCBEle3XJ0i+W5M0z4qpkQL6YZmZAASgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQXCpZR0UtU9M2xtzy7+5CtLnW1FfVuqKh+85dE5NTuTwLKudK2toJqVy5JI3JF7l1RffKzr6Wejqn09QxWPb8Pingec9enLusfd/m6fp/T3/FxHaEbtSR2hG7U89DoyidzIXaEzuZC7QtCGZYKxT1Sstlyk9r7MMzl7P4LvDuXl8Wb1dPBV0stLVQxzwTMVkkcjUc17VTJUVF4KioUk7UzTBGKuqVlsucvtfZhmcvZ7muXu7l5fF6H031LWsWWfpP8pc7lcX79GvnSA2Q1GDKmS+2KOSfD0z/ACm8XOonKvBrl1VirwR3oXjkrqafzPSKpghqaeSnqIo5oZWqySORqOa9q8FRUXgqGovSF2OzYRnlxHh2F81gldnLEmbnUTl5L3xqui8tF5KvoHPiVJO1IXciZ2pC7kRKULtCF/MmdoQv5kLNh+jbt1Sxtp8H41q3La0yjoLhIua0vJI5F/3fc73Oi+T2dtWOa9iPY5HNcmaKi5oqHl6/mbBdGjbo7Dj6fB+MqpXWVcmUVdIua0S8mPX/AHXcvuP6vZmJVmvzbgg+YZI5oWTQyMkjkajmPYubXIvFFRU1Q+iygAAAAAGhONfsvvP5fP8A6jjfY1hxFsCxrcL9cK6CssiRVNTLKxHVEiKiOcqpn7XrxAot/qNhuhxh7N96xVMzTdoKd3vPk/8Ap/CY27o6Y7X/AN9sP5zJ/wCmbD7KcLLg3AdtsMjo31ELFfUvjXNrpXKrnZLkmaJnknDREISykAEocK//AFiuH5LJ8lTzZfzPSa//AFiuH5LJ8lTzZfzIleqF+qkL/WTP1Uhf6ysrLG6Lv2+sM/jJ/wBXlN+TQbou/b6wz+Mn/V5TfktClvIACVXnx0t8LphjbZdlhYjKa7I25Qp+Mz6z/wAxsnoVCoHcjfrpTbHrptQpLLU4fnoKe5298jHrVyOYx8L0RdWtcuaOamSZe6cUM7ok7TV/lLDH55N/6RSY7s1bRpwegv8Abx/4VUfGw3zNZOjVsExps42k/tjv1ZZJqP2FLBu0lRI9+85W5cHRtTLgvM2bLV8KXnchWnSJwP8AtywLJLRw9ZdrZvVFJknlPTL2yNP6yJmic3NaWWCVHnBL2vQROXjl4Fs9JjAv7UscOuNDAjLTd1dPCjU8mKXP2yPw4rvImmTsk0KmfqRK0IJO0p2WDcQ1+FMVUGIbY/KpopkkRFXJHt0cxfBzVVq+CnXScfQQLqpEpejuE77b8TYboL9a5OspK2FJY88s297Vy901UVFTvRTtDVPocY+9g3iowHcp19j1rnT25XLwZMiZvjTuRzUzTlm1ebjawmFZjQVv0itoDdn+zmqrKaRG3euzpbciLxbIqcZPMxM3d2e6nMshyo1FVVRETiqryNAekhtAXH+0OqqKSZX2e350tvRF8lzUXypf77uOeu7uougmU1jcqvlc5285yq5yrmqqvFVIH9hSZ/ZUjRj5XNjjY573qjWtamaqqrwREKSyrH6N2zp20TaNTUtVC51ltytqrk7Lg5iL5MWfe9Uy/qo5U0PQtjWsY1jGo1rUyRETJETuK46OmztmzrZzS2+oial4rcqq5PTJVSVU4R59zEyb3Z7ypqWQXiNMVp3IACVQAAAAAAAAAAAAAAAAAAAAAAAAAAACu9tm1G27PbP1cfV1d8qWKtJSKvBqadZJlxRiLy1cqZJzVA+9tG1C07PLMuax1d6nZ/sdFvceadY/LijEVF8VVMk5qmuWzHbpiSwYxqa/ElZUXa2XKbfrYlXNYl0R8SaNyTJN1MkVEROGSKlZYivFyv13qbvd6uSrral+/LK9eKr3JyRETgiJwREREOrfzI2tEPSOy3S33q1U11tVXFV0VSxJIZo1za5PUvJUXii8FIb9aYLrTbj8mTNT2uTLTwXwNLdhG1q4bPLv7ErFkq8PVUiLU0yLmsKrw62Pud3po5E78lTdu03Chu1tp7lbaqKqo6liSQzRuza9q80KZMdctZpeNxKYmaTuFYXClnoqh9PUMVkjV4p3+KeBxHaloYgs8F2pdx+TJmp7XJlp4L4FbXGknoqt9NUxqyRmqd/ingeP5/Atxbdu9Z8S7GDkRmj3cN3MhdoTO5kLtDRhsI3akUmikrtSKTRS8DNsC4s6rq7XdJfa+zBO5ez3Ncvd3Ly+LPqiGGpp5KeoiZNDK1WSRvajmvaqZKiovBUVORQruRnWA8XdUsdqusvtfBsE7l7Pc1y93cvL4u/6b6jrWLLP0lzuVxfv0UP0hti0mFXS4mwvDJNY3LnUU6ZudRqq696x+Oqc+8ol3I9L5Y45onxSsbJG9qtexyZo5F1RU5oakdIjYjNhx1TirCcCy2TPfqqRmavo89XNTnH8LfNxTvTDnxKgXaEL+ZM7QhfzIXQv5kL/AFEz+ZC/1BML76NG3OTCc0GEsW1L5LBI7dpap65uoVXkvfF8nVOBuZDLHNCyaGRkkcjUcx7Fza5F4oqKmqHlm7kX50ZNusmEpoMJYuqXyYfkdu0tU9c3UKryXvi+TqnAmJVtX5w3SB8wyxzQsmhkZJFI1HMexc2uReKKipqh9FmMAAAAAAAAAAHCv/1iuH5LJ8lTzZfzPSa//WK4fksnyVPNl/MiV6oX6qQv9ZM/VSF/rKyssbou/b6wz+Mn/V5Tfk0G6Lv2+sM/jJ/1eU35LQpbyAAlUAAAAAAABiW1zB0GOcD1tle1iVWXXUUjv4udqLurnyReLV8HKaG19NPR1s1JVRPhqIHujljemTmOaqoqKneioej5qr0usCfsZfYcaW6DdpLiqRVqNTgyoROD/DfanvtVV4uIlMNfn9pSJ3aJX9tSFy8fSQs/aGrqbfcYK+imfBU00rZoZWL5THtXNrk8UVEU9AtkONKbHuA6DEEO42oe3qqyJv8AFTty32+bRyeDkPPZ3aXzlr9GLaTDgTF8tFeKlYrDc27tQ5c1SCRqLuSZJ6WrlyVF9yIkmNwvXpb7Qf2qYEXD9un3Lve2OiRWr5UNPpI/wV2e4nncqdk0hd2V8xme13GdVj3HVxxDUb7YZHdXSROX95gbmjG+fmv4TlUwx3ZXzETK1Y1CF/ZUvvoabOkxJjB+MbnBv2yxyJ7HRyeTLV5Irf8AAio/zqzxKVw5ZbhiO/UNitUPXVtdO2GFnLNV1VeSJqq8kRVPR3ZzhS34JwXbcM21M4qOLdfJlkssi8XvXxVyqvhpyER3RadMhABdjAAAAAAAAAAAAAAAAAAAAAAAAAAAAK3227UaDANq9jU3V1d+qWKtNTKuaRpp1kmWje5NXKndmqB9badqls2f232PEjKy+1EarTUufBiaJJJlxRvcmrlTJMuKpphiG73K/Xiou93q5KutqXq+WV68VXu7kRE4IicERERD7vdzr7zdai6XSqkq6yper5ppFzVy+pOSInBEREQ653IhaELtCF/MmdoQv5kJRP5lp7AtsFds9uaW25ulqsN1L85oU4upnLrJH628/OVY/mQv9QWemFpuFDdrbT3K21UVXR1MaSQzROza9q6KinGxDZqe70u4/Jk7E9rky08F8DTHo77YajZ9dUtV4fNUYaq3+2MTNzqR6/xrE5p901NdU4pku7VuraS40EFfQVMVTS1EaSQzRORzHtVM0VFTVCuTHXLWaXjcSrEzSdwqW40lRQ1UlNUxqyRmqd/ingcJ2hbWI7LT3ikVj8mTtT2qXLTwXwKtuVHUUFU+lqo1ZIxeKcl8U70PI87gW4tu3es+Jdjj8iMse7hu1IpNFJXakUmimlDYQu5Hyup9O5HyupaBnOAMX9V1dqu0vtfBsE7l7Pc1y93cvLzaWK9rXsVj2o5rkyVFTNFQ17doZ7s/xj1PV2m7S+19mCdy9nua5e7uXl5tO/6d6jrWLLP0lzuVxfv0Uj0kdiMlhfUYtwdSOfaHKslbQxNzWjXVXsRP4rvT3H9Xs68P5np6qIqZKmaKaS9KzZ3S4KxjDc7RCyC0XlHyxwsTJsEzVTrGInJq7zXIniqIiI1DuTDQrKln8yF/qJn8yF/qIXhE7kcd2hyHcjju0IlMNhui1t0/atPFg3GNY5bFK5G0NZI7NKFyr2XL/ul7/cr4KqpuexzXsa9jkc1yZoqLmip3nlI/1my/RQ26/sNJS4CxjVf9GOXq7bcJX/wVV0ikVf4tdGr7nRfJ7MxKlq/OG4oALsYAAAAAAADhX/6xXD8lk+Sp5sv5npNf/rFcPyWT5Knmy/mRK9UL9VIX+smfqpC/1lZWWN0Xft9YZ/GT/q8pvyaDdF37fWGfxk/6vKb8loUt5AASqAAAAAAAAHUYyw/QYqwvX4fuTc6eshWNXIiKrHatemfNqoip4oduAPObFdkr8OYkr7Hc4+rq6KZYpEyXJctHJnq1UyVF5oqHTu1U2o6YWAvZdvgx3bYM56VG09xRqdqJV8iRfFqruqvcreTTVddVKyvCF2pCTO7S+chIWRO7K+Ygd2V8xO7sr5jJtkmCqvH+O6DDlNvshld1lXM3+Jgb236Lxy4Jnw3lanMJX90Jtnfsejn2iXSBOsnR1NakcnFrEXKSVPOqbieCP5KbPHGtNBR2q10tst8DaejpIWQQRN0YxqIjUTzIiHJLwxTOwABAAAAAAAAAAAAAAAAAAAAAAAAAAVptu2o0mBrd7BoOrqL9UMzhiXi2Bq/xj/UnPzAfm23arRYEofYNCkVZfp2ZxQOXNsLV93Jl8DdV8xp/e7lX3i6VNzudVJVVlS9XyyyLmrl9SckROCIiIh93aurLnX1FwuFTJU1VQ9XyyyLm5zl5qcF/MiUwhdqQu5EztSF3IhMIXaEL+ZM7QhfzCUT+ZC/1Ez+ZC/1CVkLuRb3R220Vez+vZZb2+WpwxUSeU1M3PonKvGRic283NTzpxzR1Qu5ELtCE63D0/t1bSXGggr6CpiqqWojSSGaJyOZI1UzRUVNUOBiWyU95pNx+TJ2J7VLlovcvehUnRCw7juw4MldiaoWCz1OUltt07FWaHPi5+efkNdnnuKi8ePk5rvXiMmOuWs1vG4lji00tuJUrc6OooKx9LVRqyRmqcl8U70OFJopcOJrHTXqk3H5R1DE9qly4p4L3oVNdaKpt9XJS1UaxysXinJU7070PJc7g24tvxrPiXZ4/IjLHu4LuR8rqfTuR8rqaUNhC7QjdyJHaGTYGwnLe521dWjo7fG7iuiyqnuU8O9fQnhnw4r5rRSkd1Ml4pHVZYWAJ6ipwhQS1TnOk3XN3naq1HKjfgRCounA2Jdl1qeqN61L1GjV57qwzZ+jg34C+IY44YWQxMayNjUa1rUyRETRENOemLjmHEGMafDFumSSjsm8k7mr5L6l2W8njuIiN8FV6HssdZpjiszvUOFM9VpmFCP5kL/UTP5kL/USvCJ3I47tDkO5HHdoRKYQv9ZG/RSR/rI36KQlt10Qtt3s1lLs6xbV/7UxqR2esld++tRMkp3L90iJ5Krr2dUTe2mPJxkssEzJ4JHxSxqj2PY5Uc1ycUVFTRUPQbosbUX7ScA7tzkR1/tCsp7guWXXIqL1c2X4SNXP8JruCJkXrLHevzW6ACzGAAAAAOFf/AKxXD8lk+Sp5sv5npNf/AKxXD8lk+Sp5sv5kSvVC/VSF/rJn6qQv9ZWVljdF37fWGfxk/wCrym/JoN0Xft9YZ/GT/q8pvyWhS3kABKoAAAAAAAAAAOPcqKluVvqLfXQMnpamJ0U0T0zR7HJkqL50U8/dqeEanBGObjh6dXPjhk36aV2ssLuLHefLguXNFQ9CjTPpifbf/wCGwfG8iUwpR3aXzkS6kru0vnIl1KsiF3ZXzG6/RFwFDhfZ7HiKpYx1zvzG1CvTisdPrGxF8UXeXzoi9k0od2V8x6KbFvtQ4Q/sWl/0mkwrbwy4AFlAAAAAAAAAAAAAAAAAAAAAAAAAAADRva3X1Ny2k4iqaqRXvS4zRNz5MY5WMT0Naieg3kNLdvtnfZtql7iWNWx1U/suJV0ckqbyqn95XJ6AK/fzIn8yV/MifzIlMIXakLuRM7UhdyITCF2hC/mTO0IX8wlE/mQv9RM/mKWlqa6shoqOCSoqZ3pHFFG1XOe5VyRqImqqolZx4YZqmoip6eKSaaV6MjjjarnPcq5IiInFVVeRtl0ddgUVj9jYrxvTMmuqZSUdvfk5lKuqPfydJ3Jo3xXs5P0e9jFDgOhjvV8ihq8TTNzV3BzaJFTiyNfuslyc5PMnDNVuNVREzVckJiFJt+AqoiZquSGP1WLrTBULCnXzIi5K+NqK30ZqmZ0+LsQrVb1DQvyg0kkT+M8E8Pj82uKO1OBzfWJrfowfL5/0b+DhRMbyLdoaunraZtRSytkidoqf88DrsUWGmvlF1b8o6hie1S5cWr3L3oYFh69VFnqd9mb4HL7bFnr4p4lm2+sp6+kZVUsiPjenBeaeC9ym9xeXj52OaXjv84/p/fZgzYbce3VXwpK60NTbqx9JVxrHKxeKclTvTvQ4i6l04lsFFfaZI6hFjmZ+9zNTym+HingYxQbO2sq0fXV6SwNXPcjYrVencq58Dk5vSc1b6x94/vy3cfNpNd27S6DA+FJL3MlXVo6O3sdxXRZVT3KeHevo81swRRQQshhjbHGxqNa1qZIiJyQQxxU8DYomNiijbk1rUyRqIa0dIHbyq+yML4ErMkXOOrusTte9sKp7yv8A8Ped3icSnGpqPPzlz82a2a258Mh6R22yHDNPPhbClUya+SIsdTVRuzbRJoqIv+9+Tz4mnkznPe573K5zlzVVXNVXvJXqq8VXNVXipC/mbEypEaQv5kL/AFEz+ZC/1BaETuRx3aHIdyOO7QiUwhf6yN+ikj/WRv0UhKGTn5i8Og/iCW07bYbVvO6i9UU1M9ufDeY3rWu8/tbk/vKUfJz8xbHQ+ttbcdv9gkpGPWOibPU1L2/xcaQvbmvgrnsb/eJjyT4ehwAMjXAAAAAHCv8A9Yrh+SyfJU82X8z0mv8A9Yrh+SyfJU82X8yJXqhfqpC/1kz9VIX+srKyxui79vrDP4yf9XlN+TQbou/b6wz+Mn/V5TfktClvIACVQAAAAAAAAAADTPpifbf/AOGwfG83MNM+mJ9t/wD4bB8byJTClHdpfORLqSu7S+ciXUqyIXdlfMeimxb7UOEP7Fpf9Jp51u7K+Y9FNi32ocIf2LS/6TSYVt4ZcACygAAAAAAAAAAAAAAAAAAAAAAAAAABTvSZwBNiWxR4itUKyXO2RqksbU8qan4qqJ3q1c1ROaK7Vci4gB52v5kT+ZfPSP2VLaJ5sX4cp/8Ao2V29XU0bf4O9dZGonuF5p7lfBeFDP5kSmELtSF3ImdqQu5EJhC7QhfzJnaEL+YSifzNjuhfgaCrq67HdwhR/sSRaS3o5OzIrc5JPOjXNai/hO7jXF/M3e6JronbDrQkatVzZqlJMtUd1z14+OSp6MiS3hayqiJmvBDCcWYgWp36GhflAnCSRPd+CeHx+bXv8YPljsMyxKqZq1r1T7lV4/MV27mef9Z5t6T8Cnbcd/6N7hYK2/blG7QjdqSO0I3annIdOUL+yp2OHb5UWarR7M3wPX22LPgvincp1z+ypE7kZsWS2O0WrOphS1YtGpXNb6ynr6RlVSyI+N6cF5p4L3KcgqPDd8qbLW77M5Kd6p1sWfBfFO5S1LdW01wo2VdJIkkT04LzRe5e5T1/B51eVX8LR5hxuRx5xT7KB6X+JcaWyiprRQQPo8O1zN2ethcqumfxzhcvuEyTPL3SZ8ckchqo7kekd/tFtv1nqbRd6SOroalm5LFInBU7+9FReKKnFFRFQ0g257L7js6v2TOuqrHVOX2FWOT09W/Lgj09G8iZpzRN2WGJVq7QhfzJnaEL+ZC6F/Mhf6iZ/Mhf6gmETuRx3aHIdyOO7QiUwhf6yN+ikj/Wdrg3DF6xhiOlw/h+ifV11S7JrU4NY3m9y+5amqqpCXGwrh28YrxBS2Gw0MlbcKt25FEz4XKuiNROKqvBEPQPYBshs+yzDqxsWOtv1WxP2Qr93Xn1cefFI0X0uXivJE/dgeyKz7LsP7rerrb9VsT2fX7uvPq48+KRovpcvFeSJZheI0xWtvsAAsoAAAAdNesWYXsqq274itVA7LPcnq2McvmRVzUDmX/6xXD8lk+Sp5sv5m7OL9u2zWmtddSwXuWvqHQvjRlLSyORVVuSZOciNXXvNJn8yJXqhfqpC/1kz9VIX+srKyxui79vrDP4yf8AV5Tfk87dimJbbg/alZcR3frvYNHJIsvUs3nojonsTJM0z4uQ3Kse3PZVd3IyHF9JTPy4trI30+X96RqN95S0KWhZAOBZr3Zr1D19nu9vuMX3dJUslb77VU55KoAAAAAAAAAca43G322BZ7jXUtHEmr55WxtT0qqAck0z6Yn23/8AhsHxvNkL3ti2aWjhU4toJnZ5IlJvVPwxo5DU7pDYws+N9oj7zY1ndRtpY4EdNHuK5zVdmqJnpx55ESmFau7S+ciXUld2l85EupVkQu7K+Y9FNi32ocIf2LS/6TTzrd2V8xuNsh277NaDANgsV1vM9uraGggpZUnpJFYr2MRqqjmI5Ms01XImFbR2X4DHbHjvBV8cjLRiuy1si5e1xVsav/w55p7xkScUzQsoAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPieKKeF8E0bJYpGqx7Htza5qpkqKi6opqT0gNk02EKyS/2KB8mH53eW1M3LRvX3LvwFXR3oXjkq7ckNdS01dRzUVZBHUU08axyxSN3mvaqZKipzRUA86HakLuRbu3rZLVYIrXXi0MkqMOzvya7i51I5V4MevNv3LvQvHJXVE7kQtCF2hC/mTO0IX8yEon8y/uh/tFo7FcqnBd5qGU9LcpkmoZX8GtqFRGqxV/DRG5eLctXFAv5kL/UEzG3pxNGyaJ0UrUex6ZOauioV/iaxyWyVZokV9K9fJdzYvcvzlV9Gvbo2qZT4OxvXI2paiR2+5TvySVOUUrl913OXtaLxyV2yc0Uc0TopWI+N6ZOaqcFQ0+bwqcump7THiWTBntht7KjdoRu1O+xTYpLXL10KOfSPXyXc2L3L850LtTx2XDfDeaXjUw7NL1vHVVC/sqRO5Er+ypE7kVhZG7U7LDN+qbHV9YzOSneqdbFnwcnencp1rtSJ/ZUy4slsdotWdTCtqxaNSu+21tNcaOOrpJEkienBeaL3L3KcXE1jtWJLHU2W9UcdXQ1LN2SN/wACouqKi8UVOKKVbhi/1Nird9mclM/LrYs+Dk707lLbtldTXGijq6SVJInpwXmi9y9ynreDzq8mvftaPMONyOPOKfZoptv2W3XZxe913WVdkqXr7Crd3Xn1cmXBHonociZpzRK2fzPSrFNhtWJrDV2O9UrKqiqmKyRjk4p3OReTkXii8lQ0T21bMbxs3v609Qjqq1VDlWhrkbwkT7l3c9Oac9UN2YYonau38yF/qJn8yF/qIXhE7kcd2hyHcjtcD4TvmNcSU2H8P0i1NZOuaqvBkTE7Uj3e5amfFfMiZqqIsJcXB+GL1jDEdLYLBRPq66pfk1qcGsbze5fctTVVN/8AYZspsuy/DnsalRlXd6lqLX16tydI77hv3LE5Jz1XifexPZVYtmFhWlof9rulSiLXXB7MnyqnuWp7liLo30rmpYJaI0x2tsABZQAAAAAUl0xp54NnFubDNJGkt0ayRGOVN9vVS8Fy1TwNRn8zbXpm/a5tX9rM/wBGU1KfzIlaED+ZC/mTP5kL+ZC0IX6qQv8AWTP1Uhf6yJShf6yB/rJ3+sgf6yJTCLffFI2SN7mPY5HNc1clRU0VFPSrZHLLPspwhPPK+WWSx0T3ve5XOc5YGKqqq6qq8zzTk9Z6UbHPtRYN/sCh/V2E1Vv4ZWAC7GAAAAAI6pVSmlVFyVGL8R5xVk81TUOnqJpJpX8XPkcrnL51U9Hav+Cy/wBR3xHm+/UiUwgf21IV1Umf21IV1UiVkLu0vnIV1JndpfOQkLIndlfMQO7K+Ynd2V8xA7sr5gmEL+yptX0A6mokp8Y00k8r4IVolijc9VaxXeyN7JNEzyTPLXI1Uf2VNpv/AGf/APPb/wAB/wDcCPKLeG1IALsQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAhr6Slr6KairaeOopp2LHLFI1HNe1UyVFRdUNNNvOyeswHcVuVubJU4eqH5RS6up3LpG/wBTufnN0TjXSgorpbp7dcaaKqpKhixyxSNza9q8lCYeb7tCF/Mt/b7sgq8B1S3e0JNV4dmeiI93lPpHLox6/cqujvQvHJXVA/mQtCJ/Mhf6iZ/Mhf6iJWQu5Gz3Rl26fwXBWNazuit1xld6EilVfea5fMvJTWF3IhdoQa3D1DnijnhdDMxr43pk5qpwVCusVWGS1zddCjn0j18l3Ni9y/OUz0Z9vLqaSnwZjmuV1O5Ujt9zmdxiXRIpXL7nucumi8OKbVzxRTwuhmY18b0yc1U4Khqc3hU5dNT2mPEr4c1sNvZTL+ypE7kZFi3D8tpkWaFHPo3r5LubF7l+cx13I8hlw3w3ml41MOzS9b16qo3akT+ypK7Uif2VKwshdqdthbEFTYq3rGZyUz19uhz4OTvTuU6l2p8O0Uy48lsdotWdTCtqxaNSve2V9LcqKOso5UkienBeaL3L3KcLF+HLPivD9TY75SMqqKobk5q8FavJzV5OTVFKqwriGqsFb1keclM9U66HPg5O9O5S4rXX0tyoo6yjlSSKROC80XuXuU9ZwubXk1/C0eYcbkcecU+zQPbbsxu+zbEPsWpV1Va6lVdQ1qNySRqatd3PTNM09KFdv9R6XY3wtZcZYcqbDfqRKiknTgqcHxPTR7F9y5OS+hc0VUNK8UbBMbW7aTBhG30jq6CsVX0lx3VbB1KKm8+ReO4rc03k1zVMs8257kwx1srzBGFL3jTElNh/D9ItTWTrmqrwZExO1I93uWpnxXzImaqiLvvsX2Y2PZnhpLfb0SpuE6I6vr3NyfUPTkn3LEzXJvLXiqqqy7Idmtg2bYebb7VGk1bK1q11e9uUlS9PktTNcmpp4rmq5sTEK2tsABKoAAAAAAADr77ZLPfaVlLerXR3KBj+sZHVQtka12SpmiOTXJV4+J031OMAfeVh/wDR8XzGUgDFV2b7Pl/mTh79HxfRPz6muz37yMO/o6L6JlYAxP6mmzz7x8O/o6L6J+fUy2dfeNhz9HRfRMtAGI/Ux2c/eLhv9GxfRH1L9nH3iYb/AEbF9Ey4A2w9dl2zZdcB4a/RsX0TKqGlpqGigoqKnip6WnjbFDDE1GsjY1MmtaicERERERCYAAAAAAAAAHIjkVFRFReCopi31OcAfeXh/wDR8XzGUgDFV2b7P1/mVh79HxfMfn1Ntnv3k4e/R0X0TKwBif1NNnn3kYd/R0X0R9TPZ394+HP0dF9EywAYl9TLZ194uHP0bF9E/PqYbOPvEw3+jYvomXADEPqX7N/vDw1+jYvonb4bwthvDXsj9r1htlp9k7vX+w6ZkXWbue7vbqJnlvOy86ncAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAILhR0twoZ6Gup4qmlnYscsUjUc17VTJUVF1Q006QOx6rwLWvvNmjlqsNzv8AJdxc+jcvuHrzb9y70LxyV26RFWU1PWUktJVwRz08zFjlikajmvaqZK1UXgqKnIJidPNB/Mhf6i7OkNsYqsFVct/w9DNU4bmdm5qZufQuX3Ll1Vnc5fMvHJXUm/1FZXQu5ELtCZ3IhdoQtCF/rNlejTt9W2LSYLxxVZ0HCKgucruNPojY5VX3HJHe54Ivk8W61P8AWRP5gmNvU2eKKpgdFKxskUjcnNXiioVpi7D0tom66FHSUb18l3Ni9y/OUb0Y9vqWJtPgzHFYq2pMo7fcZFzWl5JHIv8Au+53uNF8ns7eSxwVdMscjWTQyt4pqjkU1eZw6cqmp8x4lbDmtht7KSdqRP7KmS4vw5LaJ1nhR0lE9fJdzYv3K/OY0/sqeRy4b4bzS8d3Zpet69VULtT4dop9u1Ph2ilYWQu0O3wniKqw/X9ZHnJTSKnXQ58HJ3p3KdQ7Q+6KkqK+tipKSJ0s0q5Nan/OniZsV70vE08qXrW0TFvC/qSoiqqWKpgdvRSsR7HZaoqZoSnEstH+x9opKFX76wQtYrk5qicVOWe1rMzWN+XAnW+wACyAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAR1MMNTTyU9RFHNDKxWSRyNRzXtVMlRUXgqKnI086RWxGowlJUYnwvC6fD7l3pqdM3PolX44+5dU0XvXcc+Zo45onwzRtkje1WvY5M0ci8FRU5oRMJidPMF3IhdobE9JDYW/DyzYrwZSSS2fi+soY0VzqPmr2c1i709zr2ezrs7QqyRO0L/AFkT+ZK/1kT+YShf6jYvovbenYbfTYLxpVq6yOVI6GvkdmtEvJj1/wB13L7j+r2ddH+ohdyI3pOtw9WJY4KumWORrJoJW8U1a5FKuxnhuWzSrPAjpKJ6+S7mxfuV+c166L23qTCE1PhDF9S+TD0jt2lqnqquoHKui98XydU4Zm6rm0tfRZL1VTTTsRUVFRzXtVOCovNF7zW5fEpyqanz8pThzWwW9lEO1Ph2imTY0wzNZp1qKdHSUL18l2qxr9y71KdBRUlTXVTaWkhdLM9cmtan/OSeJ5TJgvjv8O0d3YpkrevVE9kNHSVFdVR0lJE6WaR2TWt5/wDPeW/g3DNNYKTeduy1sie2y5afgt8Pj95E/cHYap7DSbzt2WtkT22XLT8Fvh8fvInfno/T/T4wx13/AHv/AA5fJ5PxJ6a+AAHVaYAAAAAHxPNFBGsk0rImJq57kRPfU+yCupIK2nWnqWb8blRVTNU+Irfq6Z6fKY1vuiW62xNbjR/57fnH7LWr/rOi/wA9vzmO4ow/bqO0T1lO2Rkke7km/mi5uROfnMIkOJyvU8/Gv0XrG/PmW9i4uPLG6zK11vFoT+VaH84Z85y4ZYp4mywyMljdxa9jkVF8yoYza8IWZ9DBLOyaZ8kbXOzkVEzVM+WRkdFSwUVLHS0zNyKNMmtzVcvSp1OPfPbvkiIj2lq5Ixx2rMpgAbLEAGKX/FiQSOp7ajJHJwdK7i1P6qc/Pp5zX5HKx8evVklkx4rZZ1WGVOVGornKiImqqcGa82mFytkuVKipqnWoqp7xg9DQ3fEk6vlqHrC1fKkkXyWr3Inf5jIqTBdpjanshZql2XHefup6ET5zTxczkcj9rFj1H4zLPbDjx9r27+znLiewp/KUXvL8xLFfrLIuTbrR5r91KjfjOIuEcPKn1v8A/Ok+kddcsBWudrlpJ56V/JM99qehePwmWbc2sb6az9Jn+asRgn5zDLI3skYj43te1dFauaKfRTl2t17wtWI5s0kTXL7XNC9Ua/wX5lMgwztAd1rKW+NburwSpY3LJfwkT4094x4vU6zfoyx0ytfiT09VJ3CwwfjXNc1HNVHNVM0VF4KhDXVdPRUzqipkRkbffVe5DpWtFY6pns1YiZnUJz4mlihZvzSMjb3uciIYLd8WV1S50dEnsaJeCLlm9fTy9Hvn3bsKV9x3aq6VT4t7juuXekVPHPT4Tl/4n8W3Rx6Tb38Q2vsvRG8k6ZRLfrNH2rnSr/VkR3xESYmsKqifslDx70X5jhxYLsrGZPbUSr906TJfgyPyXBNie3JI52L3tlX1mXq53/Gv8ZV1x/xl2LMQWNy5JdqNP60qJ8ZzKato6n+D1dPN+LkR3xGI1ezyhe1fYtwqYncusa16fBkYViqwVdhqWRzubLFImccrUyRctU8F+cw5eZysEdWTHGvaWSmDDknVbd11EVTU01MiLU1EUKO4Isj0bn750mzpznYNoXOcrl9s4qv/AMRx2tzttHco2MrIlkRi5t8pUy9434yXvii9I7zET3a01it5rbxD5W72pNbnRJ/37fnPlb1Zk1u1An/iWfOYfjmxUFsoIqmkSVrnS7itV2aZbqr6jCoIkqK+GByqiSSNYqpyzVEOVm9Tz4cnw7Ujf1bmPiY8leqJlca32xprebd+dM+c7FOKZoYvBgLDjGoklPNOqaq+ZyKv+HIyZzmRRq5yo1jEzVV5Ih1sU5ZiZyxEfRp3in3Nvo4tVcaCldu1NbTwu+5fIiL7xgl9xHW10rmQSPp6bPyWsXJXJ4r6tDtLVgqF0DJbjPJ1jkzWONURG+CrzNCvqN895pxqb1857QzzxoxxvLOndPxLYmOyW5Q5+Ga+o+48RWJ6ZpdaVP60iN+M4LsF2FyZdRMnj1qnW12z23yMX2HW1EL+XWIj2+pfhMk350d+ms/nKIrx5+csspq6iqv4NWU8/wCLlR3xKcgoa/W2ptNxloapuT2aOTRyclTwLzof4FB+Lb8Rbh8y3Im1bV1MIz4IxxExO9pgAb7WAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFRFRUVM0U1O6Smwb9jmVWMcD0jnUmay19sib+8JzkiRPcaqrfc6p5PBu2IGkxOnlm/1kT+Ztl0ltgXslKrGWBKNVqFVZa61Qs/fPupIUT3XNWc+Kpx4LqbKioqoqKipqilGWJ2hf6iF3Imf6iF3IiUwhdyL/AOjLt+qsFVNPhTF1RJUYZe7cgnciufb1VffdF3t1TVO5aAdyP2kpKqvrYaKip5amqqJEjhhiYrnyPcuSNaicVVV5CE6ifL1dglpq6jjngkiqaaeNHsexyPZIxyZoqKnBUVFPmloqOk3vYtJBBvdrq40bn58jA+jvgm64A2W26wXq4zVddm6eWN8m8ylV+S9TH+C3zqiuVypwUsMvqJ7zDBvXYABKAAAAAAAAAAAdPjT7Gav+58tpWUhZuNPsZq/7ny2lZSHlfXP9xH0/nLrcD/Tn6rctX1rpPxLPkock41q+tdJ+JZ8lDknp8f7sOVbzIAFVETNVRE8S6HQ42rn0trSGJytfULu5ovFGpr6k9JgG6r5GsambnKiIZrtAiV9JS1DeLWOc1VTxy+Ywtr1inZK3Vio5PQp4/wBXta3LmLeI1p2eFERh3HlatBSxUVHFSwpkyNuSePepOfEErJoWTRu3mPajmr3op9nrqxEViK+HHmZme4ACyHDvVvgutsnoZ0RWyNyRebXclTzKUNUxvilfFI3dex265O5U1NhXKjWq5VRERM1VSgLvMyouVVURpkyWZz28MuCqqnC9ZrXdLfPu6PAmf2o+Sy9kt0krLLNQzOVzqNyIxy/cOzyT0Ki+jI4eMa59XdnxI5eqgVY2tz5pqvv/ABH7sapZGUdwrHIqMleyNnDXdRVX5SHDxFC6C9VTHoqZyuenmVc0+M1ebkyfYMcT85/+MmGtftFnKwNRMqrx10rUc2nbvoi/dZ8PWvoLBME2fVDYrrLA5URZo/J8VRc8vez94zs6PosVjjbjzudtbnTPxe4ARVvsj2JN7E3PZG4vVb/Z3suGfhmdWZ1G2nHdKcO72yiutL7Gr4EmiRyORM1RUVOaKnExeoq8fRIqpQUsiJxzZur8G9mY67H1+T+i/wCV/wDk52X1DDWOnJWe/wCMNqnFvPesx/FZtsoaa20MdFRxrHBHnutVyrlmqqvFfFVOSV1a79jq60/sigoaaWHeVu+rWtRVTzuTMzLDa3l1A5b42BtT1i7qQ6bmSZZ+OeZscfk0yailZiPp2Y8uKad7TG3T7TvrJB+UJ8lxXVu+vFJ+UM+UhYu076yQflCfJcV1bvrxSflDPlIcD1P/AHn8HR4n+j/FeBBcadaqgqKZHbqyxuYi9yqhOfL3sZ23tbn3rkentETWYnw5MTMTuFTVtPNSVDqeojVkjFyVFT/ngd9ZsZTUkLKeugWdjUySRi5PyTvRdfgM1raKkrokZVQRzN5ZpxTzLqh0NXgu1SqqxSVEHciORU+FM/hOBX03k8W8249u39/k6M8rFljWSH3BjWxSIiySzQL3PiVfk5nY0t+stTl1NzpVV2jXSI1V9C5KYrU7P3rmsFzavcj4cvhRfUdBdMHX2ijV6U7apiJmq07t5U9C5L8Bmnl8/FH+Zj39P/UypGHj3/dssm8WO03lYn3ClbULGioxyPc3gvi1UzOxY1rGNY1MmtTJE8CibdeLpapN6hrJYO9qLm1fO1eCls4IxAmILU6eSNsVRC7cla1eCrlmjk7kX1KbPD52LPeYiurT+v5sefj3xx53DvgAdJqAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGuXSW2BMxF7KxfgmmRl5cvWVtvZk1lX3vZ3Sc1TR2va7WxoCYnTyrqYpIJnwzRvjljVWvY9uTmqnBUVF0U4zuR6Q442PbOcZ1r6++Yap310nF9VTvdBI9e9ysVN9fF2ZjEHRo2RRyte+xVkzU1Y+4TIi/4XIvwlOlki8NE8M4eveKLxDZ8P2uquVdL2YYGbyomaJvOXRrUzTNy5InNTd3o27BKHZyxMQ4gWC4YokaqMczjFRMVOLY89Xror+7gmSZq628LYXw7hahWiw5ZaC1wLlvNpoUYr1TRXKnFy+K5qduTFdK2vsABZQAAAAAAAAAAAA4t0pHVtG6nbO+BXKi77NUyK3mYrMxG5TERM93AxqqJhqqRVRFXcy8fLaVnIZ3Lg2OR28+4yuXvViL6yN2B4F/lCT/LT5zzvO4nK5eSL9Gu2vMOlgzYcNenq3+TJLU9n7GUvlt/eWc/wUORvs+7b75hzsA06/ylL/lp858O2e0zv5Tl/wApPnOpXNyoiI+F/wBoas0wzO+v9GadYz7tvvlc7Wr4j1jslNJm1uUlQrV1X3LfX7xzl2c0y/ypN/lJ85G7ZpSu/lWb/KT5zFyZ5ebHNIx637wvijBjt1Tbf5OzwBSw1mAaamnbnG/rEXw9sdxMdv1nq7ZMvWsV8K9mVqcF8/cvgZ5hy1sstnhtzJnTNiVyo9UyVc3KunpOwe1r2q17Uc1UyVFTNFGb02vIw0i3a0REbKcqceS0x3iZYJhbEraCNKKuRy06diREzVngqc0M1o6ylrI+spaiOZvex2eXn7jqq/CtoqlVzYn07l5xOyT3lzQ6ibAqdYj6e6Pjy0zizVPSioUwV53GjomsXiPfU/qnJODLPVvUszI6meCmiWWomjhjTVz3I1E9KmJLhO78WftkqEjy/D+LeIVwB10vWVt5mmXLLhHx99VU2p5HJnxi/WGKMeL53/SXDxxjKGelkttper2yIrZp8skVvNrfPzX/APpi2HMNXC+1CJCxYqZF8uoenkp5u9fBPTkWVbsF2GjVHOpnVT091O7eT3kyRfeMhYxrGIxjUa1qZIiJkiGr/h+XkZPicifyhm+00x16cUfm49roae22+GhpWbsMLd1veveq+KrxOuxPY23SNJoVRlUxMmqujk7lO6B0svHx5cfw7R2alMlqW6onuqeWOrttaiPa+CoidmmacUVOfiZpZMWUVVG2Kuc2ln0VV7DvHPl6TvqykpqyLqqqCOZnc5M8vN3HQVmDLXMqugknp15Ijt5qe/x+E5GPg8nh2mcExMT8pbluRizRrJGpZHFJHKxHxSNe1dFauaH0YQuBJWOV0N23V5e0qnxOPhcC1kjVbLefJ7urVUX/AOY3Y5PK+eH/ALQw/Cxf8/0ZhV3K30if7VXU0Pg+VEX3iiHljxbOoc857rI5O5kKN+NVObTYAscbs5X1c/4L5ERPgRF+E0uVx+VzJr1ViuvdsYcuHBvU737JtmH2Iw/jZPlKZOca2UFJbaRtJRQpFC1VVG5qvFdeK8SC9211yijjbVy0yNVVXc91n3nVpW+HBFYjcxENO01vkmd6iXR7TlT9hqduaZrPnlz7LiuqBzW3alc5Ua1J2KqqvBE3kLClwNDIqq64yqq6r1afOcd+zymcuf7Jyp/3SfOcTk8TlZ83xejX5w38ObDjp09X6M3RUVEVFRUXRTHMfx71rhkyz3ZsvQqL8yHUw7PUgXOnvtXCuefkN3fiUzGuoqetpfY1U1ZI1y5qi5pzOplpl5OC+O1emZ99tOs0xZItE7YnhbEkVLTpRXF7kYz96kyVck+5XLj5jKqW40FV/B6yCRdcmyJn72p0s+Dra9VWOapj8N5FT4jr5cCIq+13PJO50P8A+xqYPt/HrFJpFoj3Zsn2fJPVE6ZocW43Kgt0SyVtXFAiJmiOdxXzJqvoMPZgSrb5LbzuN5bsa/FvH1TbO6VF/wBquU0id0caM+Fcza+0cu0dsWvrMMXw8Meb/owLEtXFc77VVdNErI5pM2NROK8tO9dfSWVszsk9pssktXF1dTVPR6tXVrETyUXx4qvpO0s+G7NanJJS0bOuTSWTynp5lXT0ZHbmPh8CcWScuSd2/qvn5MXr0VjsAA6jTAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAH/2Q==" class="logo-img">'

st.markdown("""
<style>
:root {
    --bg: #f5f0eb; --panel: #ffffff; --text: #1c1008; --muted: #7a6a5a;
    --line: #ddd0bc; --gold: #b8913a; --gold-light: #f3dfad; --gold-pale: #faf5ee;
    --brown: #2c1a0e; --brown-mid: #4a2e1a; --brown-hover: #1e1008;
}
[data-testid="stHeader"], [data-testid="stToolbar"], [data-testid="stDecoration"], [data-testid="stStatusWidget"] {display: none;}
#MainMenu {visibility: hidden;} footer {visibility: hidden;}
.stApp { background: var(--bg); }
.block-container { max-width: 1200px; padding-top: 2.5rem; padding-bottom: 2rem; }
.badge { display:inline-block; background:var(--gold-pale); color:var(--brown); padding:8px 18px; border-radius:999px; font-size:13px; font-weight:800; margin-bottom:20px; border:1px solid var(--gold-light); }
.title { font-size:42px; line-height:1.15; font-weight:900; color:var(--brown); margin:0 0 18px 0; letter-spacing:-1.2px; }
.desc { color:var(--muted); font-size:16px; line-height:1.75; margin-bottom:30px; }
.logo-box { background:#fff; border-radius:14px; padding:18px; width:420px; max-width:100%; box-shadow:0 4px 24px rgba(44,26,14,0.08); border:1px solid var(--line); text-align:center; box-sizing:border-box; }
.logo-img { width:320px; max-width:100%; display:block; margin:0 auto; }
.version { margin-top:18px; background:var(--gold-pale); color:var(--muted); text-align:center; font-size:12px; padding:10px; width:420px; max-width:100%; border-radius:10px; border:1px solid var(--gold-light); box-sizing:border-box; letter-spacing:1px; }
.login-card { background:#fff; border-radius:18px; padding:40px; box-shadow:0 4px 32px rgba(44,26,14,0.10); border:1px solid var(--line); }
.login-gold-bar { height:3px; background:linear-gradient(90deg,var(--brown),var(--gold),var(--gold-light),var(--gold),var(--brown)); border-radius:2px; margin-bottom:28px; }
.login-title { font-size:30px; font-weight:900; color:var(--brown); margin-bottom:6px; }
.login-sub { color:var(--muted); font-size:15px; margin-bottom:28px; }
[data-testid="stTextInput"] label, [data-testid="stFileUploader"] label, [data-testid="stRadio"] label, [data-testid="stSlider"] label { color:var(--brown); font-weight:700; }
[data-testid="stTextInput"] input { background:var(--gold-pale); border:1px solid var(--line); color:var(--brown); border-radius:10px; }
[data-testid="stTextInput"] input:focus { border-color:var(--gold); box-shadow:0 0 0 2px rgba(184,145,58,0.15); }
div.stButton > button, div.stDownloadButton > button { background:var(--brown); color:#f3dfad; border:none; border-radius:10px; height:48px; font-weight:800; letter-spacing:0.5px; }
div.stButton > button:hover, div.stDownloadButton > button:hover { background:var(--brown-hover); color:#f3dfad; border:none; }
.small-text { color:#b0a090; font-size:12px; margin-top:16px; }
.topbar { background:#fff; border:1px solid var(--line); border-radius:18px; padding:22px 28px; display:flex; justify-content:space-between; align-items:center; box-shadow:0 4px 24px rgba(44,26,14,0.07); margin-bottom:26px; border-top:3px solid var(--gold); }
.topbar-title { font-size:22px; font-weight:900; color:var(--brown); }
.topbar-sub { font-size:14px; color:var(--muted); margin-top:5px; }
.dashboard-card { background:#fff; border:1px solid var(--line); border-radius:18px; padding:26px; box-shadow:0 4px 24px rgba(44,26,14,0.06); min-height:150px; border-top:3px solid var(--gold); }
.card-icon { font-size:28px; margin-bottom:12px; }
.card-title { font-size:20px; font-weight:900; color:var(--brown); margin-bottom:8px; }
.card-desc { font-size:14px; color:var(--muted); line-height:1.6; }
.section-title { font-size:20px; font-weight:900; color:var(--brown); margin:18px 0 14px 0; }
.content-card { background:#fff; border:1px solid var(--line); border-radius:18px; padding:28px; box-shadow:0 4px 24px rgba(44,26,14,0.06); margin-bottom:18px; border-top:3px solid var(--gold); }
.page-title { font-size:30px; font-weight:900; color:var(--brown); margin-bottom:8px; }
.page-sub { color:var(--muted); font-size:15px; }

/* ── 사이드바 ── */
[data-testid="stSidebar"] { background: #1e0e06 !important; border-right: 1px solid #3a2010; }
[data-testid="stSidebar"] .stButton > button { background: transparent; color: #c8a878; border: 1px solid #3a2010; border-radius: 8px; font-weight: 600; text-align: left; justify-content: flex-start; }
[data-testid="stSidebar"] .stButton > button:hover { background: #2c1a0e; color: #f3dfad; border-color: #b8913a; }
[data-testid="stSidebar"] hr { border-color: #3a2010; }

/* ── 탭 스타일 ── */
[data-testid="stTabs"] [role="tablist"] { gap:4px; border-bottom:2px solid var(--line); }
[data-testid="stTabs"] [role="tab"] { color:var(--muted); font-weight:700; font-size:14px; padding:10px 18px; border-radius:8px 8px 0 0; border:none; background:transparent; }
[data-testid="stTabs"] [role="tab"][aria-selected="true"] { color:var(--brown); border-bottom:3px solid var(--gold); background:var(--gold-pale); }
[data-testid="stTabs"] [role="tab"]:hover { background:var(--gold-pale); color:var(--brown); }
[data-testid="stTabs"] [data-testid="stTabsContent"] { padding-top:1.5rem; }

/* ── 업로더 ── */
[data-testid="stFileUploader"] { background:var(--gold-pale); border:1.5px dashed var(--gold); border-radius:12px; padding:8px; }
[data-testid="stFileUploader"] button { background:var(--brown) !important; color:#f3dfad !important; border-radius:8px !important; }

/* ── 데이터프레임 ── */
[data-testid="stDataFrame"] { border:1px solid var(--line); border-radius:12px; overflow:hidden; }

/* ── 슬라이더 ── */
[data-testid="stSlider"] [role="slider"] { background:var(--gold) !important; }
[data-testid="stSlider"] [data-testid="stSliderTrack"] { background:var(--gold-light) !important; }

/* ── 셀렉트박스/라디오 ── */
[data-testid="stSelectbox"] select, [data-testid="stMultiSelect"] { border-color:var(--line) !important; border-radius:10px !important; }
[data-testid="stRadio"] [role="radio"][aria-checked="true"] + div { color:var(--brown); font-weight:700; }

/* ── 섹션 구분선 ── */
hr { border-color:var(--line); }

/* ── success/info/error 메시지 ── */
[data-testid="stAlert"][data-baseweb="notification"] { border-radius:12px; border-left:4px solid var(--gold); }

/* ── 캡션 ── */
[data-testid="stCaptionContainer"] { color:var(--muted); }

/* ── topbar 골드 상단 라인 ── */
.topbar { border-top:3px solid var(--gold) !important; }
</style>
""", unsafe_allow_html=True)

def login_page():
    st.markdown("""
    <style>
    [data-testid="stAppViewContainer"] > .main { padding: 0 !important; }
    .block-container { padding: 0 !important; max-width: 100% !important; }
    [data-testid="stSidebar"] { display: none !important; }
    [data-testid="collapsedControl"] { display: none !important; }
    </style>
    """, unsafe_allow_html=True)

    left_col, right_col = st.columns([1, 1])

    with left_col:
        st.markdown("""
        <div style="background:#2c1a0e; min-height:100vh; padding:80px 56px; display:flex; flex-direction:column; justify-content:center; box-sizing:border-box;">
          <div style="font-size:36px; font-weight:900; color:#f3dfad; letter-spacing:1px; margin-bottom:6px;">TY LOGIS</div>
          <div style="font-size:12px; color:#b8913a; letter-spacing:5px; margin-bottom:28px;">INTERNAL SYSTEM</div>
          <div style="height:1px; background:#4a2e1a; margin-bottom:32px;"></div>
          <div style="display:flex; align-items:center; gap:12px; margin-bottom:16px;">
            <div style="width:7px; height:7px; background:#b8913a; border-radius:50%; flex-shrink:0;"></div>
            <div style="font-size:16px; color:#c8a878;">전자상거래 통관 · 택배 업무</div>
          </div>
          <div style="display:flex; align-items:center; gap:12px; margin-bottom:16px;">
            <div style="width:7px; height:7px; background:#b8913a; border-radius:50%; flex-shrink:0;"></div>
            <div style="font-size:16px; color:#c8a878;">3PL BL 변환 · 현장 운영</div>
          </div>
          <div style="display:flex; align-items:center; gap:12px; margin-bottom:16px;">
            <div style="width:7px; height:7px; background:#b8913a; border-radius:50%; flex-shrink:0;"></div>
            <div style="font-size:16px; color:#c8a878;">씨앤에어 자동화 신고</div>
          </div>
          <div style="display:flex; align-items:center; gap:12px; margin-bottom:40px;">
            <div style="width:7px; height:7px; background:#b8913a; border-radius:50%; flex-shrink:0;"></div>
            <div style="font-size:16px; color:#c8a878;">알리 HT 변환 · 주소 검증</div>
          </div>
          <div style="font-size:12px; color:#4a2e1a; letter-spacing:2px;">TY · KY · YST 통합 업무 포털 · v23.0</div>
        </div>
        """, unsafe_allow_html=True)

    with right_col:
        # 상단 여백으로 세로 중앙 맞추기
        st.markdown("<div style='height: 28vh;'></div>", unsafe_allow_html=True)
        st.markdown("""
        <div style="padding: 0 40px;">
        <div style="font-size:36px; font-weight:900; color:#2c1a0e; margin-bottom:8px;">로그인</div>
        <div style="font-size:15px; color:#9a7a60; margin-bottom:20px;">계정과 비밀번호를 입력하세요.</div>
        <div style="width:48px; height:3px; background:#b8913a; border-radius:2px; margin-bottom:32px;"></div>
        </div>
        """, unsafe_allow_html=True)

        _, fc, _ = st.columns([0.08, 0.84, 0.08])
        with fc:
            with st.form("login_form", clear_on_submit=False):
                user = st.text_input("사용자 계정", placeholder="예: admin")
                pw   = st.text_input("비밀번호", type="password", placeholder="비밀번호 입력")
                submitted = st.form_submit_button("로그인", use_container_width=True)
            if submitted:
                ok, role = check_login(user.strip(), pw)
                if ok:
                    st.session_state.login = True
                    st.session_state.user  = user.strip()
                    st.session_state.role  = role
                    st.session_state.page  = "main"
                    st.rerun()
                else:
                    st.error("아이디 또는 비밀번호가 틀렸습니다.")

def topbar():
    if "role" not in st.session_state:
        st.session_state.role = "user"
    page_map = {
        "main": "메인 대시보드",
        "ecommerce": "전자상거래",
        "seaair": "SEA & AIR",
        "threepl": "3PL",
        "bl_convert": "3PL BL PDF 변환",
        "kyungdong": "전자상 경동리스트",
        "address_verify": "주소 / 통관 검증",
        "meni_convert": "메니변환",
    }
    page_name = page_map.get(st.session_state.page, "메인 대시보드")
    st.markdown(f"""
    <div class="topbar"><div><div class="topbar-title">TY LOGIS 업무 자동화 시스템</div><div class="topbar-sub">접속 계정: {st.session_state.user} · {page_name}</div></div><div class="badge">v23.0</div></div>
    """, unsafe_allow_html=True)


def main_page():
    today_str = datetime.now().strftime("%Y년 %m월 %d일")
    load_dashboard_from_github()

    # ── 사이드바 ──
    with st.sidebar:
        st.markdown(f"""
        <div style="background:#2c1a0e; margin:-1rem -1rem 0; padding:28px 20px 20px; text-align:center;">
          <div style="font-size:22px; font-weight:900; color:#f3dfad; letter-spacing:1px;">TY LOGIS</div>
          <div style="font-size:9px; color:#b8913a; letter-spacing:4px; margin-top:3px;">INTERNAL SYSTEM</div>
          <div style="height:1px; background:#4a2e1a; margin:14px 0 10px;"></div>
          <div style="font-size:11px; color:#7a5a30;">{st.session_state.user} 님</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='margin-top:16px;'></div>", unsafe_allow_html=True)
        st.markdown('<div style="font-size:11px;color:#b8913a;letter-spacing:2px;padding:0 4px;margin-bottom:6px;">NAVIGATION</div>', unsafe_allow_html=True)

        if st.button("🏠  메인 대시보드", use_container_width=True, key="nav_main"):
            st.session_state.page = "main"; st.rerun()
        if st.button("🛒  전자상거래", use_container_width=True, key="nav_ecom"):
            st.session_state.page = "ecommerce"; st.rerun()
        if st.button("🚢  SEA & AIR", use_container_width=True, key="nav_sea"):
            st.session_state.page = "seaair"; st.rerun()
        if st.button("🏭  3PL", use_container_width=True, key="nav_3pl"):
            st.session_state.page = "threepl"; st.rerun()

        st.markdown("<div style='margin-top:16px;'></div>", unsafe_allow_html=True)
        st.divider()

        if is_admin():
            if st.button("👤  계정 관리", use_container_width=True, key="nav_admin"):
                st.session_state.page = "admin"; st.rerun()
        if st.button("🚪  로그아웃", use_container_width=True, key="nav_logout"):
            st.session_state.login = False
            st.session_state.user  = ""
            st.session_state.role  = "user"
            st.session_state.page  = "main"
            st.rerun()

        st.markdown(f'<div style="font-size:10px;color:#b8913a;text-align:center;margin-top:20px;">{today_str}</div>', unsafe_allow_html=True)

    # ── 메인 콘텐츠: 현장 현황 대시보드 ──
    st.markdown(f'<div class="topbar"><div><div class="topbar-title">오늘의 현장 현황</div><div class="topbar-sub">{today_str} · 접속: {st.session_state.user}</div></div><div class="badge">v23.0</div></div>', unsafe_allow_html=True)

    if is_admin():
        with st.expander("🔧 관리자 — 이미지 업로드", expanded=False):
            ua, ub, uc = st.columns(3)
            with ua:
                up_sched = st.file_uploader("📅 작업일정", type=["png","jpg","jpeg","webp"], key="up_sched")
                if up_sched:
                    img_bytes = up_sched.read()
                    st.session_state.dash_schedule = img_bytes
                    with st.spinner("저장 중..."):
                        ok = github_upload_image(img_bytes, "schedule.png")
                    st.success("작업일정 등록 완료!" + (" (영구저장)" if ok else " (임시저장)"))
                if st.session_state.dash_schedule:
                    if st.button("🗑 작업일정 삭제", key="del_sched"):
                        st.session_state.dash_schedule = None
                        github_delete_image("schedule.png")
                        st.rerun()
            with ub:
                up_ty = st.file_uploader("📊 TY 현황표", type=["png","jpg","jpeg","webp"], key="up_ty")
                if up_ty:
                    img_bytes = up_ty.read()
                    st.session_state.dash_ty = img_bytes
                    with st.spinner("저장 중..."):
                        ok = github_upload_image(img_bytes, "ty.png")
                    st.success("TY 현황표 등록 완료!" + (" (영구저장)" if ok else " (임시저장)"))
                if st.session_state.dash_ty:
                    if st.button("🗑 TY 현황표 삭제", key="del_ty"):
                        st.session_state.dash_ty = None
                        github_delete_image("ty.png")
                        st.rerun()
            with uc:
                up_ky = st.file_uploader("📊 KY 현황표", type=["png","jpg","jpeg","webp"], key="up_ky")
                if up_ky:
                    img_bytes = up_ky.read()
                    st.session_state.dash_ky = img_bytes
                    with st.spinner("저장 중..."):
                        ok = github_upload_image(img_bytes, "ky.png")
                    st.success("KY 현황표 등록 완료!" + (" (영구저장)" if ok else " (임시저장)"))
                if st.session_state.dash_ky:
                    if st.button("🗑 KY 현황표 삭제", key="del_ky"):
                        st.session_state.dash_ky = None
                        github_delete_image("ky.png")
                        st.rerun()

    tab_sched, tab_ty, tab_ky = st.tabs(["📅 작업일정", "📊 TY 현황표", "📊 KY 현황표"])
    with tab_sched:
        if st.session_state.dash_schedule:
            st.markdown(f'<div style="font-size:13px;color:#9a7a60;margin-bottom:8px;">{today_str} 작업일정</div>', unsafe_allow_html=True)
            st.image(st.session_state.dash_schedule, use_container_width=True)
        else:
            st.markdown('<div style="text-align:center;padding:60px 0;color:#b8913a;font-size:15px;">등록된 작업일정이 없습니다.</div>', unsafe_allow_html=True)
    with tab_ty:
        if st.session_state.dash_ty:
            st.markdown(f'<div style="font-size:13px;color:#9a7a60;margin-bottom:8px;">{today_str} TY 작업현황표</div>', unsafe_allow_html=True)
            st.image(st.session_state.dash_ty, use_container_width=True)
        else:
            st.markdown('<div style="text-align:center;padding:60px 0;color:#b8913a;font-size:15px;">등록된 TY 현황표가 없습니다.</div>', unsafe_allow_html=True)
    with tab_ky:
        if st.session_state.dash_ky:
            st.markdown(f'<div style="font-size:13px;color:#9a7a60;margin-bottom:8px;">{today_str} KY 작업현황표</div>', unsafe_allow_html=True)
            st.image(st.session_state.dash_ky, use_container_width=True)
        else:
            st.markdown('<div style="text-align:center;padding:60px 0;color:#b8913a;font-size:15px;">등록된 KY 현황표가 없습니다.</div>', unsafe_allow_html=True)


def admin_page():
    """관리자 전용: 계정 생성/삭제/비밀번호 변경."""
    if not is_admin():
        st.error("관리자 계정으로만 접근할 수 있습니다.")
        return
    topbar()
    if st.button("← 메인으로 돌아가기", key="admin_back"):
        st.session_state.page = "main"
        st.rerun()

    st.markdown(
        '<div class="content-card"><div class="page-title">👤 계정 관리</div>'
        '<div class="page-sub">사용자 계정을 생성·수정·삭제합니다. 관리자 전용 페이지입니다.</div></div>',
        unsafe_allow_html=True,
    )

    users = load_users()

    # ── 현재 계정 목록 ──
    st.markdown('<div class="section-title">현재 등록 계정</div>', unsafe_allow_html=True)
    for uname, uinfo in list(users.items()):
        c1, c2, c3 = st.columns([2, 2, 1])
        c1.write(f"**{uname}**")
        c2.write(f"권한: {'🔑 관리자' if uinfo.get('role') == 'admin' else '👤 일반'}")
        if uname != st.session_state.user:   # 자기 자신은 삭제 불가
            if c3.button("삭제", key=f"del_{uname}"):
                del users[uname]
                save_users(users)
                st.success(f"{uname} 계정이 삭제되었습니다.")
                st.rerun()
        else:
            c3.write("(본인)")

    st.divider()

    # ── 새 계정 만들기 ──
    st.markdown('<div class="section-title">새 계정 만들기</div>', unsafe_allow_html=True)
    with st.form("create_user_form"):
        new_id   = st.text_input("아이디")
        new_pw   = st.text_input("비밀번호", type="password")
        new_pw2  = st.text_input("비밀번호 확인", type="password")
        new_role = st.selectbox("권한", ["user", "admin"], format_func=lambda x: "관리자" if x == "admin" else "일반 사용자")
        ok_btn   = st.form_submit_button("계정 생성", use_container_width=True)
    if ok_btn:
        if not new_id or not new_pw:
            st.error("아이디와 비밀번호를 입력하세요.")
        elif new_pw != new_pw2:
            st.error("비밀번호가 일치하지 않습니다.")
        elif new_id in users:
            st.error("이미 존재하는 아이디입니다.")
        else:
            users[new_id] = {"password": new_pw, "role": new_role}
            save_users(users)
            st.success(f"'{new_id}' 계정이 생성되었습니다.")
            st.rerun()

    st.divider()

    # ── 비밀번호 변경 ──
    st.markdown('<div class="section-title">비밀번호 변경</div>', unsafe_allow_html=True)
    with st.form("change_pw_form"):
        target  = st.selectbox("계정 선택", list(users.keys()))
        chg_pw  = st.text_input("새 비밀번호", type="password")
        chg_pw2 = st.text_input("새 비밀번호 확인", type="password")
        chg_btn = st.form_submit_button("비밀번호 변경", use_container_width=True)
    if chg_btn:
        if not chg_pw:
            st.error("비밀번호를 입력하세요.")
        elif chg_pw != chg_pw2:
            st.error("비밀번호가 일치하지 않습니다.")
        else:
            users[target]["password"] = chg_pw
            save_users(users)
            st.success(f"'{target}' 비밀번호가 변경되었습니다.")
            st.rerun()


def ecommerce_page():
    topbar()
    if st.button("← 메인으로 돌아가기", key="ecom_back"):
        st.session_state.page = "main"
        st.rerun()

    st.markdown(
        '<div class="content-card"><div class="page-title">🛒 전자상거래</div>'
        '<div class="page-sub">전자상 통관·택배·경동리스트 업무 메뉴입니다.</div></div>',
        unsafe_allow_html=True,
    )

    c1, c2, c3, c4 = st.columns(4, gap="large")
    with c1:
        st.markdown(
            '<div class="dashboard-card"><div class="card-icon">📦</div>'
            '<div class="card-title">전자상 경동리스트</div>'
            '<div class="card-desc">멀티건 송장 매칭, 동춘경동, 학익경동 자동변환을 처리합니다.</div></div>',
            unsafe_allow_html=True,
        )
        if st.button("전자상 경동리스트 열기", use_container_width=True, key="open_kd_from_ecom"):
            st.session_state.page = "kyungdong"
            st.rerun()

    with c2:
        st.markdown(
            '<div class="dashboard-card"><div class="card-icon">📍</div>'
            '<div class="card-title">주소 / 통관 검증</div>'
            '<div class="card-desc">주소 정리, 우편번호 확인, 개인통관부호 검증 메뉴를 추가할 수 있습니다.</div></div>',
            unsafe_allow_html=True,
        )
        if st.button("주소 / 통관 검증 열기", use_container_width=True, key="open_addr_from_ecom"):
            st.session_state.page = "address_verify"
            st.rerun()

    with c3:
        st.markdown(
            '<div class="dashboard-card"><div class="card-icon">🧾</div>'
            '<div class="card-title">메니변환</div>'
            '<div class="card-desc">HDFC 메니 파일의 HS코드, 용도구분, 중량, FTA 관련 처리를 자동 변환합니다.</div></div>',
            unsafe_allow_html=True,
        )
        if st.button("메니변환 열기", use_container_width=True, key="open_meni_from_ecom"):
            st.session_state.page = "meni_convert"
            st.rerun()

    with c4:
        st.markdown(
            '<div class="dashboard-card"><div class="card-icon">📊</div>'
            '<div class="card-title">알리 HT변환</div>'
            '<div class="card-desc">원본 기준 금액·V·HS CODE를 최종 수정하고 메모 시트를 생성합니다.</div></div>',
            unsafe_allow_html=True,
        )
        if st.button("알리 HT변환 열기", use_container_width=True, key="open_ali_ht_from_ecom"):
            st.session_state.page = "ali_ht_convert"
            st.rerun()


def address_verify_page():
    topbar()
    if st.button("← 전자상거래로 돌아가기", key="addr_back"):
        st.session_state.page = "ecommerce"
        st.rerun()

    st.markdown(
        '<div class="content-card"><div class="page-title">📍 주소 / 통관 검증</div>'
        '<div class="page-sub">카카오 주소 API로 엑셀 주소를 검증하고 정상/오류 결과를 다운로드합니다.</div></div>',
        unsafe_allow_html=True,
    )

    with st.expander("사용 안내", expanded=True):
        st.write("- 엑셀은 최소 2개 컬럼이 필요합니다. 예: 송장 / 주소")
        st.write("- 카카오 REST API 키를 입력한 뒤 엑셀을 업로드하세요.")
        st.write("- 결과는 전체결과, 정상, 오류 시트로 나누어 다운로드됩니다.")
        st.write("- 빠른 검증 모드는 중복 주소를 한 번만 조회하고, 여러 건을 동시에 처리합니다.")

    kakao_api_key = st.text_input(
        "카카오 REST API 키",
        type="password",
        help="Kakao Developers > 내 애플리케이션 > 앱 키 > REST API 키",
        key="addr_kakao_api_key",
    )

    uploaded = st.file_uploader("주소 검증용 엑셀 업로드", type=["xlsx"], key="addr_excel_file")

    if uploaded is None:
        st.info("엑셀 파일을 업로드하면 주소 검증을 진행할 수 있습니다.")
        return

    if not kakao_api_key:
        st.warning("카카오 REST API 키를 먼저 입력해주세요.")
        return

    try:
        df = pd.read_excel(uploaded)
    except Exception as e:
        st.error(f"엑셀 파일을 읽는 중 오류가 발생했습니다: {e}")
        return

    if df.empty:
        st.error("업로드한 엑셀에 데이터가 없습니다.")
        return

    try:
        default_invoice_col, default_address_col = detect_columns(df)
    except Exception as e:
        st.error(str(e))
        return

    st.markdown('<div class="section-title">컬럼 확인</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        invoice_col = st.selectbox("송장 컬럼", options=list(df.columns), index=list(df.columns).index(default_invoice_col), key="addr_invoice_col")
    with c2:
        address_col = st.selectbox("주소 컬럼", options=list(df.columns), index=list(df.columns).index(default_address_col), key="addr_address_col")

    st.markdown('<div class="section-title">미리보기</div>', unsafe_allow_html=True)
    st.dataframe(df[[invoice_col, address_col]].head(20), use_container_width=True)

    unique_count = df[address_col].astype(str).nunique(dropna=False)
    total_count = len(df)
    c_speed1, c_speed2 = st.columns(2)
    with c_speed1:
        max_workers = st.slider("동시 처리 개수", min_value=1, max_value=10, value=6, step=1, help="값이 높을수록 빠르지만, 너무 높으면 카카오 API 호출 제한이 걸릴 수 있습니다.")
    with c_speed2:
        st.metric("실제 조회할 고유 주소 수", f"{unique_count:,} / {total_count:,}")

    if st.button("✅ 주소 검증 실행", type="primary", use_container_width=True, key="addr_run"):
        progress = st.progress(0)
        status = st.empty()
        results_by_addr = {}
        results = []
        total_unique = unique_count

        thread_local = threading.local()

        def get_session():
            if not hasattr(thread_local, "session"):
                thread_local.session = requests.Session()
            return thread_local.session

        def worker(raw_addr):
            return raw_addr, classify_kakao_result(raw_addr, kakao_api_key, get_session())

        unique_addresses = list(dict.fromkeys(df[address_col].astype(str).tolist()))
        done = 0

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_map = {executor.submit(worker, addr): addr for addr in unique_addresses}
            for future in as_completed(future_map):
                addr = future_map[future]
                try:
                    raw_addr, r = future.result()
                except Exception as e:
                    raw_addr = addr
                    r = {
                        "원본주소": addr, "조회주소": "", "판정": "오류",
                        "도로명주소": "", "지번주소": "", "위도": "", "경도": "",
                        "오류사유": f"처리 오류: {e}"
                    }
                results_by_addr[raw_addr] = r
                done += 1
                if done == total_unique or done % 20 == 0:
                    status.text(f"처리 중... 고유주소 {done:,} / {total_unique:,}건")
                    progress.progress(done / total_unique)

        status.text("결과 정리 중...")
        for _, row in df.iterrows():
            addr_key = str(row[address_col])
            r = dict(results_by_addr.get(addr_key, {
                "원본주소": row[address_col], "조회주소": "", "판정": "오류",
                "도로명주소": "", "지번주소": "", "위도": "", "경도": "",
                "오류사유": "결과 매칭 실패"
            }))
            r["송장"] = row[invoice_col]
            results.append(r)

        progress.progress(1.0)
        status.text("완료")

        result_df = pd.DataFrame(results)[["송장", "원본주소", "조회주소", "판정", "도로명주소", "지번주소", "위도", "경도", "오류사유"]]
        ok_df = result_df[result_df["판정"] == "정상"].copy()
        err_df = result_df[result_df["판정"] == "오류"].copy()

        st.success("주소 검증이 완료되었습니다.")
        m1, m2, m3 = st.columns(3)
        m1.metric("총 건수", len(result_df))
        m2.metric("정상", len(ok_df))
        m3.metric("오류", len(err_df))

        st.markdown('<div class="section-title">결과 미리보기</div>', unsafe_allow_html=True)
        st.dataframe(result_df.head(50), use_container_width=True)

        detail_df = build_detail_address_df(ok_df)
        excel_bytes = to_excel_bytes(result_df, ok_df, err_df, detail_df)
        st.download_button(
            "⬇️ 주소검증 결과 다운로드",
            excel_bytes,
            file_name="카카오_주소검증_결과_v8.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="addr_download",
        )


def seaair_page():
    topbar()
    if st.button("← 메인으로 돌아가기", key="seaair_back"):
        st.session_state.page = "main"
        st.rerun()

    st.markdown(
        '<div class="content-card"><div class="page-title">🚢 SEA & AIR</div>'
        '<div class="page-sub">해상·항공 포워딩 관련 업무 메뉴입니다.</div></div>',
        unsafe_allow_html=True,
    )

    c1, c2, c3, c4 = st.columns(4, gap="large")
    with c1:
        st.markdown(
            '<div class="dashboard-card"><div class="card-icon">📄</div>'
            '<div class="card-title">서류 변환</div>'
            '<div class="card-desc">해상·항공 서류 자동화 메뉴를 추가할 수 있습니다.</div></div>',
            unsafe_allow_html=True,
        )
        st.button("준비중", use_container_width=True, disabled=True, key="seaair_doc_ready")

    with c2:
        st.markdown(
            '<div class="dashboard-card"><div class="card-icon">✈️</div>'
            '<div class="card-title">항공 업무</div>'
            '<div class="card-desc">항공 관련 업무 메뉴를 추가할 수 있습니다.</div></div>',
            unsafe_allow_html=True,
        )
        st.button("준비중", use_container_width=True, disabled=True, key="seaair_air_ready")

    with c3:
        st.markdown(
            '<div class="dashboard-card"><div class="card-icon">🚢</div>'
            '<div class="card-title">해상 업무</div>'
            '<div class="card-desc">해상 관련 업무 메뉴를 추가할 수 있습니다.</div></div>',
            unsafe_allow_html=True,
        )
        st.button("준비중", use_container_width=True, disabled=True, key="seaair_sea_ready")


def threepl_page():
    topbar()
    if st.button("← 메인으로 돌아가기", key="threepl_back"):
        st.session_state.page = "main"
        st.rerun()

    st.markdown(
        '<div class="content-card"><div class="page-title">🏭 3PL</div>'
        '<div class="page-sub">3PL BL/PDF 변환 및 현장 운영 업무 메뉴입니다.</div></div>',
        unsafe_allow_html=True,
    )

    c1, c2, c3, c4 = st.columns(4, gap="large")
    with c1:
        st.markdown(
            '<div class="dashboard-card"><div class="card-icon">📄</div>'
            '<div class="card-title">3PL BL / PDF 변환</div>'
            '<div class="card-desc">PDF BL 파일을 업로드하면 엑셀 변환 결과를 생성합니다.</div></div>',
            unsafe_allow_html=True,
        )
        if st.button("BL 변환 열기", use_container_width=True, key="open_bl_from_3pl"):
            st.session_state.page = "bl_convert"
            st.rerun()

    with c2:
        st.markdown(
            '<div class="dashboard-card"><div class="card-icon">📦</div>'
            '<div class="card-title">팔레트 적재</div>'
            '<div class="card-desc">현장 적재 및 팔레트 계산 메뉴를 추가할 수 있습니다.</div></div>',
            unsafe_allow_html=True,
        )
        st.button("준비중", use_container_width=True, disabled=True, key="threepl_pallet_ready")

    with c3:
        st.markdown(
            '<div class="dashboard-card"><div class="card-icon">🏷️</div>'
            '<div class="card-title">현장 운영</div>'
            '<div class="card-desc">입출고, 현장 작업자료 메뉴를 추가할 수 있습니다.</div></div>',
            unsafe_allow_html=True,
        )
        st.button("준비중", use_container_width=True, disabled=True, key="threepl_ops_ready")


def bl_convert_page():
    topbar()
    if st.button("← 메인으로 돌아가기"):
        st.session_state.page = "main"; st.rerun()
    st.markdown('<div class="content-card"><div class="page-title">📄 3PL BL PDF → Excel 변환</div><div class="page-sub">여러 개의 BL PDF 파일을 업로드하면 엑셀 파일로 변환합니다.</div></div>', unsafe_allow_html=True)
    with st.expander("추출 설정", expanded=False):
        desc_mode = st.radio("품명 오른쪽 범위", ["안전모드", "넓게 인식"], help="안전모드는 중량칸 혼입을 더 강하게 막고, 넓게 인식은 길게 넘어간 품명을 더 많이 잡습니다.")
        desc_right_ratio = 0.80 if desc_mode == "안전모드" else 0.89
        table_bottom_ratio = st.slider("MARK/품명 하단 범위", min_value=0.55, max_value=0.78, value=0.69, step=0.01, help="마크가 아래로 길게 이어지는 특수 PDF면 조금 올려주세요. 너무 올리면 하단 문구가 포함될 수 있습니다.")
    uploaded = st.file_uploader("BL PDF 파일 업로드", type=["pdf"], accept_multiple_files=True)
    if uploaded:
        rows = []
        progress = st.progress(0)
        for idx, file in enumerate(uploaded, start=1):
            try:
                rows.append(extract_one_pdf(file.getvalue(), file.name, desc_right_ratio, table_bottom_ratio))
            except Exception as e:
                rows.append({**{c: "" for c in COLUMNS}, "원본파일명": file.name, "확인필요": f"오류: {e}"})
            progress.progress(idx / len(uploaded))
        df = pd.DataFrame(rows, columns=COLUMNS)
        st.subheader("미리보기")
        st.dataframe(df, use_container_width=True, hide_index=True)
        excel_bytes = make_excel(df)
        file_name = f"BL_PDF_변환결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button("엑셀 다운로드", data=excel_bytes, file_name=file_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with st.expander("반영된 변환 기능"):
            st.write("- SREJY / SILKROAD 특수 BL 라벨 기준 인식 보정")
            st.write("- HANSUNG INCHEON 선명/항차 분리 보정")
            st.write("- WOOYOUNG / INUF 화주 전용 BL 양식 라벨 기준 인식 보정")
            st.write("- 항차 패턴 보정: 숫자 4자리, 숫자4자리+E, 숫자3자리+E")
            st.write("- 컨사이니 주소, 선명/항차, 마크, 수량, 중량, CBM 인식 범위 보정")
            st.write("- 품명은 SAID TO CONTAIN 아래 내용 기준으로 추출")
    else:
        st.info("PDF를 업로드하면 자동으로 엑셀 변환 결과가 생성됩니다.")


# ==============================
# 전자상 경동리스트
# ==============================
def kd_parse_list(text: str):
    if not text or not text.strip():
        return []
    items = re.split(r"[,\s;]+", text.strip())
    out, seen = [], set()
    for x in items:
        x = x.strip()
        if x and x not in seen:
            out.append(x)
            seen.add(x)
    return out

def kd_norm_str(x):
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if s.lower() == "nan":
        return ""
    return s

def kd_to_float(x):
    try:
        if pd.isna(x):
            return None
        s = str(x).replace(",", "").strip()
        if s == "" or s.lower() == "nan":
            return None
        return float(s)
    except:
        return None

def kd_set_number_like_original(value):
    s = kd_norm_str(value)
    if s == "":
        return ""
    try:
        return int(s)
    except:
        return s

def kd_replace_delivery_terms(df):
    replace_map = {
        "경착": "착택",
        "착불": "착택",
        "신용": "현택",
        "대납": "현택",
        "선불": "현택",
        "경선": "현택",
    }

    def convert_cell(x):
        if pd.isna(x):
            return x

        s = str(x).strip()
        s_key = s.replace(" ", "").replace("\n", "").replace("\t", "")

        return replace_map.get(s_key, x)

    return df.map(convert_cell)

def kyungdong_page():
    topbar()
    if st.button("← 메인으로 돌아가기", key="kd_back"):
        st.session_state.page = "main"
        st.rerun()

    st.markdown(
        '<div class="content-card"><div class="page-title">📦 전자상 경동리스트</div>'
        '<div class="page-sub">멀티건 송장 매칭 · 동춘경동 자동변환 · 학익경동 자동변환</div></div>',
        unsafe_allow_html=True,
    )

    tab1, tab2, tab3 = st.tabs(["① 멀티건 송장 매칭", "② 동춘경동 자동변환", "③ 학익경동 자동변환"])

    # ---------------- TAB 1 ----------------
    with tab1:
        st.subheader("멀티건 송장 매칭 (조회불가 포함)")
        uploaded = st.file_uploader("멀티건모음 엑셀 업로드", type=["xlsx", "xls"], key="kd_multi")

        e_text = st.text_area(
            "조회할 택배송장(E) 목록 (선택)",
            height=120,
            placeholder="예)\n507144526830\n507144526841\n...",
            key="kd_e_text",
        )

        if uploaded:
            xls = pd.ExcelFile(uploaded)
            sheet = st.selectbox("시트 선택", xls.sheet_names, key="kd_sheet")
            header_mode = st.radio("첫 줄이 컬럼명(헤더)인가요?", ["네", "아니요"], horizontal=True, key="kd_header")

            if header_mode == "네":
                df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
                st.caption(f"행 {len(df):,} / 컬럼 {len(df.columns)}")

                e_col = st.selectbox(
                    "E열(택배송장) 컬럼",
                    df.columns,
                    index=df.columns.get_loc("CJ单号") if "CJ单号" in df.columns else 0,
                    key="kd_e_col",
                )

                f_col = st.selectbox(
                    "F열(세관신고송장) 컬럼",
                    df.columns,
                    index=df.columns.get_loc("主单号") if "主单号" in df.columns else 0,
                    key="kd_f_col",
                )

            else:
                df = pd.read_excel(xls, sheet_name=sheet, dtype=str, header=None)
                df.columns = [
                    string.ascii_uppercase[i] if i < 26 else f"COL{i+1}"
                    for i in range(len(df.columns))
                ]

                st.caption(f"행 {len(df):,} / 컬럼 {len(df.columns)} (헤더 없음)")

                e_col = st.selectbox(
                    "E열 문자",
                    df.columns,
                    index=df.columns.get_loc("E") if "E" in df.columns else 0,
                    key="kd_e_col_nohead",
                )

                f_col = st.selectbox(
                    "F열 문자",
                    df.columns,
                    index=df.columns.get_loc("F") if "F" in df.columns else 0,
                    key="kd_f_col_nohead",
                )

            with st.expander("데이터 미리보기(상위 30행)"):
                st.dataframe(df.head(30), use_container_width=True)

            if st.button("✅ 매칭 생성", type="primary", use_container_width=True, key="kd_make_match"):
                query = kd_parse_list(e_text)

                base = df.copy()
                base[e_col] = base[e_col].apply(kd_norm_str)
                base[f_col] = base[f_col].apply(kd_norm_str)
                base = base[base[e_col] != ""]

                not_in_file = []

                if query:
                    all_e = set(base[e_col].tolist())
                    not_in_file = [e for e in query if e not in all_e]
                    base = base[base[e_col].isin(query)].copy()

                valid = base[base[f_col] != ""][[e_col, f_col]].copy()
                invalid = base[base[f_col] == ""][[e_col]].copy()

                out1 = valid.rename(columns={e_col: "택배송장(E)", f_col: "세관신고송장(F)"})
                out1["비고"] = "정상 매칭"

                out2 = invalid.rename(columns={e_col: "택배송장(E)"})
                out2["비고"] = "조회안되는송장(F 없음)"

                if not_in_file:
                    add = pd.DataFrame({
                        "택배송장(E)": not_in_file,
                        "비고": ["조회안되는송장(파일에 없음)"] * len(not_in_file),
                    })
                    out2 = pd.concat([out2, add], ignore_index=True)

                out1 = out1.drop_duplicates().reset_index(drop=True)
                out2 = out2.drop_duplicates().reset_index(drop=True)

                st.success(f"정상매칭 {len(out1):,}건 / 조회안되는송장 {len(out2):,}건")

                c1, c2 = st.columns(2)
                with c1:
                    st.markdown("### 정상매칭")
                    st.dataframe(out1.head(200), use_container_width=True)
                with c2:
                    st.markdown("### 조회안되는송장")
                    st.dataframe(out2.head(200), use_container_width=True)

                bio = io.BytesIO()
                with pd.ExcelWriter(bio, engine="openpyxl") as w:
                    out1.to_excel(w, index=False, sheet_name="정상매칭")
                    out2.to_excel(w, index=False, sheet_name="조회안되는송장")
                bio.seek(0)

                st.download_button(
                    "⬇️ 엑셀 다운로드 (정상매칭/조회안되는송장)",
                    bio.getvalue(),
                    file_name="멀티건_E-F_매칭_조회불가포함.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="kd_match_download",
                )

    # ---------------- TAB 2 ----------------
    with tab2:
        st.subheader("동춘경동 자동변환 (멀티건 분할 + 중량/BOX 처리 + A열 숫자)")
        st.write("**멀티건 매칭파일(정상매칭 시트)** + **동춘경동 신규파일**을 업로드하세요.")

        match_file = st.file_uploader(
            "멀티건 매칭파일(.xlsx) — 정상매칭 시트 포함",
            type=["xlsx"],
            key="kd_dc_match",
        )

        hak_file = st.file_uploader(
            "동춘경동 신규파일(.xlsx)",
            type=["xlsx"],
            key="kd_dc_file",
        )

        if match_file and hak_file:
            match_df = pd.read_excel(match_file, sheet_name="정상매칭", dtype=str)
            hak_df = pd.read_excel(hak_file, dtype=str)

            if "세관신고송장(F)" not in match_df.columns or "택배송장(E)" not in match_df.columns:
                st.error("매칭파일의 '정상매칭' 시트 컬럼명이 예상과 달라요. (택배송장(E), 세관신고송장(F))")
                st.stop()

            required_dc_cols = ["HBL NO", "BOX 수량", "중량"]
            missing_dc_cols = [c for c in required_dc_cols if c not in hak_df.columns]
            if missing_dc_cols:
                st.error(
                    "동춘경동 신규파일 양식이 아니에요. "
                    f"필요 컬럼: {', '.join(required_dc_cols)} / 없는 컬럼: {', '.join(missing_dc_cols)}\n\n"
                    "업로드한 파일이 학익경동 양식이면 위의 '③ 학익경동 자동변환' 탭에서 처리해주세요."
                )
                st.stop()

            map_group = match_df.groupby("세관신고송장(F)")["택배송장(E)"].apply(list).to_dict()
            rows = []

            for _, r in hak_df.iterrows():
                hbl = kd_norm_str(r.get("HBL NO"))
                e_list = map_group.get(hbl)

                box = kd_to_float(r.get("BOX 수량"))
                wt = kd_to_float(r.get("중량"))

                denom = box if (box and box > 0) else (len(e_list) if e_list else None)
                per = (wt / denom) if (wt is not None and denom) else wt

                def base_row(rr):
                    rr = rr.copy()
                    if "BOX 수량" in rr.index:
                        rr.loc["BOX 수량"] = "1"
                    if per is not None and "중량" in rr.index:
                        rr.loc["중량"] = str(round(per, 3))
                    return rr

                if e_list:
                    for e in e_list:
                        rr = base_row(r)
                        rr["HBL NO"] = kd_norm_str(e)
                        rows.append(rr)
                else:
                    rr = base_row(r)
                    rows.append(rr)

            out = pd.DataFrame(rows)
            if set(hak_df.columns).issubset(set(out.columns)):
                out = out[hak_df.columns]

            out = kd_replace_delivery_terms(out)

            bio = io.BytesIO()
            with pd.ExcelWriter(bio, engine="openpyxl") as w:
                out.to_excel(w, index=False, sheet_name="변환결과")
            bio.seek(0)

            st.success(f"변환 완료: {len(out):,}행")
            st.dataframe(out.head(50), use_container_width=True)

            st.download_button(
                "⬇️ 동춘경동 변환파일 다운로드",
                bio.getvalue(),
                file_name="동춘경동_택배송장변환.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="kd_dc_download",
            )

    # ---------------- TAB 3 ----------------
    with tab3:
        st.subheader("학익경동 자동변환 (멀티건 분할 + C/T·W/T 처리)")
        st.write("**멀티건 매칭파일(정상매칭 시트)** + **학익경동 신규파일**을 업로드하세요.")
        st.caption("학익경동 양식 기준: HBL NO(또는 송장번호) / C/T / W/T 컬럼을 사용합니다.")

        hakik_match_file = st.file_uploader(
            "멀티건 매칭파일(.xlsx) — 정상매칭 시트 포함",
            type=["xlsx"],
            key="kd_hakik_match",
        )

        hakik_file = st.file_uploader(
            "학익경동 신규파일(.xlsx)",
            type=["xlsx"],
            key="kd_hakik_file",
        )

        if hakik_match_file and hakik_file:
            match_df = pd.read_excel(hakik_match_file, sheet_name="정상매칭", dtype=str)
            hakik_df = pd.read_excel(hakik_file, dtype=str)

            required_match_cols = ["세관신고송장(F)", "택배송장(E)"]
            missing_match_cols = [c for c in required_match_cols if c not in match_df.columns]

            if missing_match_cols:
                st.error("매칭파일의 '정상매칭' 시트 컬럼명이 예상과 달라요. (택배송장(E), 세관신고송장(F))")
                st.stop()

            if "HBL NO" in hakik_df.columns:
                waybill_col = "HBL NO"
            elif "송장번호" in hakik_df.columns:
                waybill_col = "송장번호"
            else:
                st.error("학익경동 신규파일에 필요한 컬럼이 없어요: HBL NO 또는 송장번호")
                st.stop()

            required_hakik_cols = [waybill_col, "C/T", "W/T"]
            missing_hakik_cols = [c for c in required_hakik_cols if c not in hakik_df.columns]

            if missing_hakik_cols:
                st.error(f"학익경동 신규파일에 필요한 컬럼이 없어요: {', '.join(missing_hakik_cols)}")
                st.stop()

            match_df["세관신고송장(F)"] = match_df["세관신고송장(F)"].apply(kd_norm_str)
            match_df["택배송장(E)"] = match_df["택배송장(E)"].apply(kd_norm_str)

            map_group = (
                match_df[match_df["세관신고송장(F)"] != ""]
                .groupby("세관신고송장(F)")["택배송장(E)"]
                .apply(list)
                .to_dict()
            )

            rows = []
            unmatched_rows = []

            for _, r in hakik_df.iterrows():
                waybill = kd_norm_str(r.get(waybill_col))
                e_list = map_group.get(waybill)

                ct = kd_to_float(r.get("C/T"))
                wt = kd_to_float(r.get("W/T"))

                if e_list:
                    repeat_count = len(e_list)
                elif ct and ct > 0:
                    repeat_count = int(ct)
                else:
                    repeat_count = 1

                weight_divisor = ct if (ct and ct > 0) else repeat_count
                per_wt = (wt / weight_divisor) if (wt is not None and weight_divisor) else wt

                def make_row(rr):
                    rr = rr.copy()
                    if "C/T" in rr.index:
                        rr.loc["C/T"] = "1"
                    if per_wt is not None and "W/T" in rr.index:
                        rr.loc["W/T"] = str(round(per_wt, 3))
                    return rr

                if e_list:
                    for e in e_list:
                        rr = make_row(r)
                        rr[waybill_col] = kd_norm_str(e)
                        rows.append(rr)
                else:
                    for _ in range(repeat_count):
                        rr = make_row(r)
                        rows.append(rr)

                    if ct and ct > 1:
                        unmatched_rows.append({
                            waybill_col: waybill,
                            "C/T": ct,
                            "W/T": wt,
                            "비고": "매칭 송장 없이 C/T 기준으로 자동 분할됨",
                        })

            out = pd.DataFrame(rows)
            if set(hakik_df.columns).issubset(set(out.columns)):
                out = out[hakik_df.columns]

            out = kd_replace_delivery_terms(out)

            bio = io.BytesIO()
            with pd.ExcelWriter(bio, engine="openpyxl") as w:
                out.to_excel(w, index=False, sheet_name="변환결과")
                if unmatched_rows:
                    pd.DataFrame(unmatched_rows).to_excel(w, index=False, sheet_name="확인필요")
            bio.seek(0)

            st.success(f"변환 완료: {len(out):,}행")

            if unmatched_rows:
                st.warning(f"확인필요 {len(unmatched_rows):,}건이 있습니다. 다운로드 파일의 '확인필요' 시트를 확인해주세요.")

            st.dataframe(out.head(50), use_container_width=True)

            st.download_button(
                "⬇️ 학익경동 변환파일 다운로드",
                bio.getvalue(),
                file_name="학익경동_택배송장변환.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="kd_hakik_download",
            )



# ==============================
# 메니변환
# ==============================
MENI_KEYWORDS_V = [
    "mini PC", "Vibrator", "smart watch", "Earphones",
    "tablet android", "tablet pc", "Speakers", "smartphone",
    "lenovo xiaoxin pad", "IPAD TABLET", "TABLET", "DRAWING BOARD",
    "BLUETOOTH SPEAKER", "BLUETOOTH EARPHONE CLIP",
    "BLUETOOTH EARBUDS", "BLUETOOTH",
]

MENI_WIRELESS_TERMS = [
    "Wireless Earphones",
    "Wireless Headphones",
    "Wireless Earbuds",
    "Wireless Bluetooth Earphones",
    "Wireless Bluetooth Headphones",
]

MENI_WEIGHT_CYCLE = [1.5, 1.6, 1.7, 1.8, 1.9]
MENI_STEP_WEIGHT = 0.1
MENI_HS_EXCEPT_FOR_V3 = {"900290", "900410", "902920"}

MENI_FTA_HS_CODES = {
    "330430", "420292", "630622", "630629", "640219", "640299",
    "640419", "732112", "846729", "847130", "847141", "847150",
    "847160", "847180", "850440", "850760", "850811", "851511",
    "851671", "851762", "851810", "851830", "851840", "852351",
    "852499", "852589", "852691", "852852", "852862", "852869",
    "854231", "854370", "870829", "870894", "870899", "871200",
    "871496", "871499", "900290", "900410", "902920", "910212",
    "910591", "940169", "950450", "950490", "950691",
}

def meni_find_column_name(columns, keyword, startswith=False):
    for c in columns:
        s = str(c)
        if startswith:
            if s.startswith(keyword):
                return c
        else:
            if keyword in s:
                return c
    raise ValueError(f"'{keyword}' 컬럼명을 찾을 수 없습니다.")

def meni_distribute_to_target(df, col_af, target_total):
    w_series = pd.to_numeric(df[col_af], errors="coerce")
    current_total = float(w_series.sum())
    remaining = float(target_total) - current_total

    if remaining <= 0:
        return df, current_total, 0.0

    candidates = df.index[pd.to_numeric(df[col_af], errors="coerce") >= 2].tolist()
    if not candidates:
        return df, current_total, 0.0

    distributed = 0.0
    i = 0
    n = len(candidates)
    max_loops = 2_000_000

    while remaining >= MENI_STEP_WEIGHT - 1e-12 and max_loops > 0:
        idx = candidates[i % n]
        cur = float(w_series.loc[idx])
        new_val = cur + MENI_STEP_WEIGHT

        if abs(new_val - 30.0) < 1e-12:
            i += 1
            max_loops -= 1
            continue

        w_series.loc[idx] = new_val
        df.at[idx, col_af] = new_val

        distributed += MENI_STEP_WEIGHT
        remaining -= MENI_STEP_WEIGHT
        i += 1
        max_loops -= 1

    if remaining > 0:
        w_series = pd.to_numeric(df[col_af], errors="coerce")
        max_idx = w_series.idxmax()
        cur = float(w_series.loc[max_idx])
        new_val = cur + remaining

        if abs(new_val - 30.0) >= 1e-12:
            df.at[max_idx, col_af] = new_val
            distributed += remaining
            w_series.loc[max_idx] = new_val
            remaining = 0.0

    new_total = float(w_series.sum())
    return df, new_total, distributed



def normalize_meni_special_chars(value):
    """품명 컬럼에서만 é/É 문자를 e/E로 변경합니다."""
    if pd.isna(value):
        return value
    if not isinstance(value, str):
        return value
    return value.replace("é", "e").replace("É", "E")

def meni_process_excel_to_bytes(uploaded_file, target_total=None):
    df = pd.read_excel(uploaded_file).astype("object")

    col_hs    = meni_find_column_name(df.columns, "허용품목코드")
    col_zip   = meni_find_column_name(df.columns, "ZIP CODE")
    col_v     = meni_find_column_name(df.columns, "용도구분")
    col_desc1 = meni_find_column_name(df.columns, "1.DESCRIPTION", startswith=True)
    col_desc2 = meni_find_column_name(df.columns, "2.DESCRIPTION", startswith=True)
    col_af    = meni_find_column_name(df.columns, "Total W/T")
    col_hawb  = meni_find_column_name(df.columns, "HAWB NO")
    col_tel   = meni_find_column_name(df.columns, "C/TEL")
    col_total = meni_find_column_name(df.columns, "总金额")

    w_orig = pd.to_numeric(df[col_af], errors="coerce")
    count_le2_orig = int(((w_orig <= 2) & w_orig.notna()).sum())

    def convert_hs_row(row):
        val_hs = row[col_hs]
        v_val = row[col_v]
        if pd.isna(val_hs):
            return val_hs
        s = str(val_hs).strip()

        if str(v_val).strip() == "3" and s in MENI_HS_EXCEPT_FOR_V3:
            return s

        if s.startswith(("1", "2", "30", "90")):
            return "960719"
        return s

    df[col_hs] = df.apply(convert_hs_row, axis=1)

    def fix_zip(z):
        if pd.isna(z):
            return z
        s = str(z).strip()
        if len(s) == 4:
            return "0" + s
        return s

    df[col_zip] = df[col_zip].apply(fix_zip)

    v_before_str = df[col_v].astype(str).str.strip()
    wireless_mask = (v_before_str == "1") & df[col_desc1].astype(str).apply(
        lambda x: any(term.upper() in str(x).upper() for term in MENI_WIRELESS_TERMS)
    )
    df.loc[wireless_mask, col_v] = "3"
    rows_v_blue = df.index[wireless_mask].tolist()
    wireless_changed_cnt = int(wireless_mask.sum())
    wireless_ratio_all = (wireless_changed_cnt / len(df) * 100) if len(df) else 0.0

    def match_v(desc, v):
        if pd.isna(desc) or pd.isna(v):
            return False
        if str(v).strip() != "1":
            return False
        t = str(desc).upper()
        return any(kw.upper() in t for kw in MENI_KEYWORDS_V)

    mask_v = df.apply(lambda r: match_v(r[col_desc1], r[col_v]), axis=1)
    df.loc[mask_v, col_v] = "3"
    rows_v_red = df.index[mask_v].tolist()

    w = pd.to_numeric(df[col_af], errors="coerce")
    mask_range = (w >= 2) & (w <= 5)
    bp_empty = df[col_desc2].isna() | (df[col_desc2].astype(str).str.strip() == "")
    bh_no_elec = ~df[col_desc1].astype(str).str.upper().str.contains("ELECTRIC", na=False)
    mask_target = mask_range & bp_empty & bh_no_elec
    t_idx = df.index[mask_target].tolist()

    for i, idx in enumerate(t_idx):
        df.at[idx, col_af] = MENI_WEIGHT_CYCLE[i % len(MENI_WEIGHT_CYCLE)]

    distributed_total = None
    new_total = None
    if target_total is not None:
        df, new_total, distributed_total = meni_distribute_to_target(df, col_af, target_total)

    w_after = pd.to_numeric(df[col_af], errors="coerce")
    count_le2_after = int(((w_after <= 2) & w_after.notna()).sum())

    tel = df[col_tel].astype(str).str.strip()
    v_str = df[col_v].astype(str).str.strip()
    amt = pd.to_numeric(df[col_total], errors="coerce")

    mask_v1 = (v_str == "1") & tel.notna() & (tel != "")
    df_v1 = pd.DataFrame({"TEL": tel.where(mask_v1), "AMT": amt.where(mask_v1)})
    tel_sum = df_v1.groupby("TEL", dropna=True)["AMT"].sum()
    bad_tels = set(tel_sum[tel_sum >= 150].index.tolist())

    mask_phone_rule = mask_v1 & tel.isin(bad_tels)
    rows_v_orange = df.index[mask_phone_rule].tolist()
    df.loc[mask_phone_rule, col_v] = "3"
    hawb_list = df.loc[mask_phone_rule, col_hawb].astype(str).tolist()

    v_after_str = df[col_v].astype(str).str.strip()
    hs_str = df[col_hs].astype(str).str.strip()
    amt_after = pd.to_numeric(df[col_total], errors="coerce")

    mask_fta = (v_after_str == "3") & (amt_after >= 150) & hs_str.isin(MENI_FTA_HS_CODES)
    fta_hawb_list = df.loc[mask_fta, col_hawb].astype(str).tolist()

    # 메니변환본 특수문자 정리 (é -> e 등)
    # 품명 DESCRIPTION 컬럼에서만 é/É 문자를 e/E로 변경
    for _col in [c for c in df.columns if "DESCRIPTION" in str(c).upper()]:
        df[_col] = df[_col].map(normalize_meni_special_chars)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="처리결과")
        wb = writer.book
        ws = writer.sheets["처리결과"]

        header = next(ws.iter_rows(min_row=1, max_row=1))
        col_idx = {cell.value: cell.column for cell in header}
        v_idx = col_idx[col_v]
        af_idx = col_idx[col_af]

        blue   = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")
        red    = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        green  = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
        orange = PatternFill(start_color="FFFF9900", end_color="FFFF9900", fill_type="solid")

        rows_v_blue_excel   = {i + 2 for i in rows_v_blue}
        rows_v_red_excel    = {i + 2 for i in rows_v_red}
        rows_af_excel       = {i + 2 for i in t_idx}
        rows_v_orange_excel = {i + 2 for i in rows_v_orange}

        for r in range(2, ws.max_row + 1):
            if r in rows_v_blue_excel:
                ws.cell(row=r, column=v_idx).fill = blue
            if r in rows_v_red_excel:
                ws.cell(row=r, column=v_idx).fill = red
            if r in rows_af_excel:
                ws.cell(row=r, column=af_idx).fill = green
            if r in rows_v_orange_excel:
                ws.cell(row=r, column=v_idx).fill = orange

        memo = wb.create_sheet("메모")
        memo["A1"] = "AF ≤ 2 (원본)"
        memo["B1"] = count_le2_orig

        memo["A2"] = "목표 총중량(kg)"
        memo["B2"] = "" if target_total is None else float(target_total)

        memo["A3"] = "분배 후 총중량(kg)"
        memo["B3"] = "" if new_total is None else float(new_total)

        memo["A4"] = "2차 분배량 합계(kg)"
        memo["B4"] = "" if distributed_total is None else float(distributed_total)

        memo["A5"] = "AF ≤ 2 (분배 후)"
        memo["B5"] = count_le2_after

        memo["A6"] = "WIRELESS 변경 건수(V=1→3)"
        memo["B6"] = wireless_changed_cnt

        memo["A7"] = "전화번호 중복 + 총금액합계≥150 행 수"
        memo["B7"] = len(rows_v_orange)

        memo["A8"] = "해당 전화번호 수"
        memo["B8"] = len(bad_tels)

        memo["A9"] = "WIRELESS 전체 대비 변경 비율(%)"
        memo["B9"] = round(wireless_ratio_all, 2)

        memo["A10"] = "해당 HAWB NO 리스트"
        row = 11
        for h in hawb_list:
            memo[f"A{row}"] = h
            row += 1

        row += 1
        memo[f"A{row}"] = "FTA적용건 HAWB 리스트"
        memo[f"B{row}"] = len(fta_hawb_list)
        row += 1
        for h in fta_hawb_list:
            memo[f"A{row}"] = h
            row += 1

    output.seek(0)

    summary = {
        "원본 AF≤2": count_le2_orig,
        "분배 후 AF≤2": count_le2_after,
        "WIRELESS 변경": wireless_changed_cnt,
        "키워드 V변경": len(rows_v_red),
        "중량 1차 재분배": len(t_idx),
        "전화번호/총금액 변경": len(rows_v_orange),
        "FTA 적용건": len(fta_hawb_list),
    }

    return output.getvalue(), summary

def meni_convert_page():
    topbar()
    if st.button("← 전자상거래로 돌아가기", key="meni_back"):
        st.session_state.page = "ecommerce"
        st.rerun()

    st.markdown(
        '<div class="content-card"><div class="page-title">🧾 메니변환</div>'
        '<div class="page-sub">HDFC 메니변환</div></div>',
        unsafe_allow_html=True,
    )

    uploaded = st.file_uploader("메니 엑셀 파일 업로드", type=["xlsx", "xls"], key="meni_file")
    target_text = st.text_input("목표 총중량(kg) (선택)", placeholder="없으면 비워두세요", key="meni_target")

    target_total = None
    if target_text.strip():
        try:
            target_total = float(target_text.replace(",", "").strip())
        except Exception:
            st.warning("목표 총중량은 숫자로 입력해주세요. 숫자가 아니면 목표 중량 없이 처리됩니다.")
            target_total = None

    if uploaded:
        if st.button("✅ 메니변환 실행", type="primary", use_container_width=True, key="meni_run"):
            try:
                with st.spinner("메니변환 처리 중입니다. 파일 용량에 따라 10초~1분 정도 걸릴 수 있어요..."):
                    result_bytes, summary = meni_process_excel_to_bytes(uploaded, target_total=target_total)

                st.success("메니변환 완료")

                st.write("처리 요약")
                st.dataframe(pd.DataFrame([summary]), use_container_width=True, hide_index=True)

                st.download_button(
                    "⬇️ 메니변환 결과 다운로드",
                    result_bytes,
                    file_name="메니변환_중량조정_최종.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="meni_download",
                )
            except Exception as e:
                st.error(f"처리 중 오류가 발생했습니다: {e}")
                st.exception(e)
    else:
        st.info("엑셀 파일을 업로드하면 메니변환을 실행할 수 있습니다.")


# ==============================
# 알리HT변환
# ==============================
ALI_HT_NAME_MAP = {
    "Dried Bamboo Shoots": "Seasoned Dried Bamboo Shoots",
    "Peanuts": "Processed peanuts",
    "Seed": "Seasoned Seed",
    "Walnut Kernels": "Processed Walnut Kernels",
    "Rice": "rice cake",
}

def ali_ht_norm_col(c):
    return re.sub(r"\s+", "", str(c)).upper()

def ali_ht_find_column(columns, keywords, required=True):
    for col in columns:
        n = ali_ht_norm_col(col)
        if all(str(k).upper() in n for k in keywords):
            return col
    if required:
        raise ValueError(f"필수 컬럼을 찾지 못했습니다: {keywords}")
    return None

def ali_ht_clean_text(v):
    if pd.isna(v):
        return ""
    return str(v).strip()

def ali_ht_to_number(v):
    if pd.isna(v):
        return 0.0
    s = str(v).replace(",", "").strip()
    try:
        return float(s)
    except Exception:
        return 0.0


def ali_ht_money(v):
    """금액/단가 계산용: 소수점 2자리 반올림."""
    return round(float(v or 0), 2)


def ali_ht_set_text(value, width=None):
    """엑셀 숫자형으로 읽힌 코드를 문자형으로 정리."""
    if pd.isna(value):
        return ""
    s = str(value).strip()
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    s = re.sub(r"\s+", "", s)
    if width and s:
        s = s.zfill(width)
    return s


def ali_ht_pick_column(columns, exact_letter_index=None, keywords=None, required=True):
    """헤더명이 흔들릴 때를 대비해 키워드 우선, 실패 시 엑셀 열 위치로 보정."""
    keywords = keywords or []
    try:
        if keywords:
            return ali_ht_find_column(columns, keywords, required=False)
    except Exception:
        pass
    cols = list(columns)
    if exact_letter_index is not None and len(cols) > exact_letter_index:
        return cols[exact_letter_index]
    if required:
        raise ValueError(f"필수 컬럼을 찾지 못했습니다: {keywords}")
    return None


def ali_ht_add_log(logs, row_no, hawb, name, tel, field, before, after, reason):
    logs.append({
        "원본행": row_no,
        "HAWB NO": hawb,
        "수취인": name,
        "전화번호": tel,
        "변경항목": field,
        "변경전": before,
        "변경후": after,
        "사유": reason,
    })


def ali_ht_process_excel_to_bytes(uploaded_file):
    """알리HT 원본 엑셀을 원본 서식 유지 상태로 직접 수정하고 메모 시트를 추가합니다."""
    from openpyxl.utils import get_column_letter

    # 업로드 파일은 pandas/openpyxl에서 각각 읽어야 하므로 bytes로 고정
    file_bytes = uploaded_file.getvalue() if hasattr(uploaded_file, "getvalue") else uploaded_file.read()
    df = pd.read_excel(io.BytesIO(file_bytes), dtype=object).astype("object")
    original_columns = list(df.columns)
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    # 주요 컬럼: 헤더명 우선, 실패 시 사용자가 지정한 열 위치 기준
    col_hawb = ali_ht_pick_column(df.columns, 4, ["HAWB", "NO"])          # E
    col_v = ali_ht_pick_column(df.columns, 21, ["용도구분"])              # V
    col_hs = ali_ht_pick_column(df.columns, 29, ["허용품목코드"], False)  # AD
    col_name = ali_ht_pick_column(df.columns, 10, ["C/NAME", "KOR"])     # K
    col_tel = ali_ht_pick_column(df.columns, 14, ["C/TEL"])              # O
    col_total = ali_ht_pick_column(df.columns, 52, ["总金额"])            # BA

    col_pos = {col: idx + 1 for idx, col in enumerate(original_columns)}

    # 상세 품명/수량/단가 반복 구조 찾기: DESCRIPTION 기준 +2=QTY, +6=INVOICE VALUE
    detail_groups = []
    cols = list(df.columns)
    for idx, col in enumerate(cols):
        norm = ali_ht_norm_col(col)
        if "DESCRIPTION" in norm and "ITEMREMARK" not in norm:
            qty_idx = idx + 2
            unit_idx = idx + 6
            if qty_idx < len(cols) and unit_idx < len(cols):
                qty_col = cols[qty_idx]
                unit_col = cols[unit_idx]
                if "QTY" in ali_ht_norm_col(qty_col) and ("INVOICE" in ali_ht_norm_col(unit_col) or "VALUE" in ali_ht_norm_col(unit_col)):
                    detail_groups.append((col, qty_col, unit_col))
    if not detail_groups:
        # 안전장치: 기존 위치 기준
        desc_col = ali_ht_pick_column(df.columns, 60, ["1.DESCRIPTION"], False)
        qty_col = ali_ht_pick_column(df.columns, 62, ["QTY"], False)
        unit_col = ali_ht_pick_column(df.columns, 66, ["INVOICE", "VALUE"], False)
        if not (desc_col and qty_col and unit_col):
            raise ValueError("상세 품명/QTY/INVOICE VALUE 반복 컬럼을 찾지 못했습니다.")
        detail_groups = [(desc_col, qty_col, unit_col)]

    logs = []
    money_changed_cells = set()
    v_changed_cells = set()
    hs_changed_cells = set()
    text_changed_cells = set()
    desc_changed_cells = set()
    excluded_from_list_hawbs = []

    def excel_set(i, col, value):
        df.at[i, col] = value
        ws.cell(row=i + 2, column=col_pos[col]).value = value

    def row_hawb(i): return ali_ht_clean_text(df.at[i, col_hawb])
    def row_name(i): return ali_ht_clean_text(df.at[i, col_name])
    def row_tel(i): return ali_ht_clean_text(df.at[i, col_tel])
    def group_key(i): return (row_name(i), row_tel(i))
    def total_val(i): return ali_ht_money(ali_ht_to_number(df.at[i, col_total]))
    def qty_val(i, qty_col):
        q = ali_ht_to_number(df.at[i, qty_col])
        return q if q else 1.0
    def unit_val(i, unit_col): return ali_ht_money(ali_ht_to_number(df.at[i, unit_col]))
    def detail_signature(i):
        names = []
        for desc_col, qty_col, unit_col in detail_groups:
            d = ali_ht_clean_text(df.at[i, desc_col])
            if d:
                names.append(d.upper())
        return tuple(names)
    def first_desc(i):
        sig = detail_signature(i)
        return sig[0] if sig else ""
    def recalc_row_total(i):
        total = 0.0
        for desc_col, qty_col, unit_col in detail_groups:
            desc = ali_ht_clean_text(df.at[i, desc_col])
            unit = unit_val(i, unit_col)
            qty = qty_val(i, qty_col)
            if desc or unit:
                total += unit * qty
        return ali_ht_money(total)
    log_group_no = 0

    def _pick_hawb_line(text, hawb):
        """그룹 변경 전/후 문자열에서 해당 HAWB 줄만 뽑아 메모를 세로형으로 보기 좋게 표시."""
        text = ali_ht_clean_text(text)
        hawb = ali_ht_clean_text(hawb)
        if not text or not hawb:
            return text
        parts = [p.strip() for p in text.split(" / ") if p.strip()]
        for p in parts:
            if hawb in p:
                return p
        return text

    def add_log(kind, idxs, modified_idxs, field, before, after, reason, row_i=None):
        """메모 시트는 옆으로 송장 나열하지 않고, 관련 송장을 한 줄씩 세로로 기록.
        수정제외 로그는 남기지 않으며, 실제 수정이 발생한 그룹만 기록한다.
        """
        nonlocal log_group_no
        if kind == "수정제외":
            return
        idxs = list(dict.fromkeys(idxs or []))
        modified_idxs = list(dict.fromkeys(modified_idxs or []))
        if not modified_idxs:
            return
        rows_to_write = idxs if len(idxs) > 1 else modified_idxs
        log_group_no += 1
        modified_set = set(modified_idxs)
        for ii in rows_to_write:
            hawb = row_hawb(ii)
            is_modified = ii in modified_set
            logs.append({
                "구분": kind,
                "그룹번호": log_group_no,
                "수취인": row_name(ii),
                "전화번호": row_tel(ii),
                "HAWB NO": hawb,
                "원본행": ii + 2,
                "처리상태": "수정" if is_modified else "유지",
                "변경항목": field if is_modified else "그룹 내 유지",
                "변경전": _pick_hawb_line(before, hawb) if is_modified else "",
                "변경후": _pick_hawb_line(after, hawb) if is_modified else "",
                "사유": reason,
            })

    # E열 HAWB 문자 처리
    for i in df.index:
        before = df.at[i, col_hawb]
        after = ali_ht_set_text(before)
        if str(before) != str(after):
            excel_set(i, col_hawb, after)
            text_changed_cells.add((i, col_hawb))
            add_log("개별수정", [i], [i], "HAWB 문자형", before, after, "운송장번호 문자형 정리", i)

    # 품명 변경: 왼쪽 원본 품명 → 오른쪽 변경 품명 단방향 처리
    # Dried는 단순 전체 치환이 아니라, 품명 안에 Dried가 포함되고 Snack이 없을 때만
    # 첫 Dried 앞에 Seasoned를 붙인다. 예: Dried Fruit Mix → Seasoned Dried Fruit Mix
    name_map_norm = {k.strip().lower(): v for k, v in ALI_HT_NAME_MAP.items()}
    name_change_count = 0
    for i in df.index:
        for desc_col, _qty_col, _unit_col in detail_groups:
            before = ali_ht_clean_text(df.at[i, desc_col])
            if not before:
                continue

            before_norm = before.strip().lower()
            after = name_map_norm.get(before_norm)

            if not after:
                has_dried = re.search(r"\bdried\b", before, flags=re.I) is not None
                has_snack = re.search(r"\bsnack\b", before, flags=re.I) is not None
                already_seasoned_dried = re.search(r"\bseasoned\s+dried\b", before, flags=re.I) is not None
                if has_dried and not has_snack and not already_seasoned_dried:
                    after = re.sub(r"\bDried\b", "Seasoned Dried", before, count=1, flags=re.I)

            if after and before != after:
                excel_set(i, desc_col, after)
                desc_changed_cells.add((i, desc_col))
                name_change_count += 1
                add_log("품명변경", [i], [i], desc_col, before, after, "요청 품명 자동 변경", i)

    # AD 허용품목코드 6자리 문자형 + 30 시작코드 960719 변경
    if col_hs is not None:
        for i in df.index:
            before = df.at[i, col_hs]
            code = ali_ht_set_text(before, 6)
            after = "960719" if code.startswith("30") else code
            if str(before) != str(after):
                excel_set(i, col_hs, after)
                hs_changed_cells.add((i, col_hs))
                reason = "30 시작 HS CODE → 960719 변경" if code.startswith("30") else "HS CODE 6자리 문자형 정리"
                add_log("HS수정", [i], [i], "AD 허용품목코드", before, after, reason, i)

    # V=1 목록건 처리
    v1_indices = [i for i in df.index if ali_ht_clean_text(df.at[i, col_v]) == "1"]
    groups = {}
    for i in v1_indices:
        groups.setdefault(group_key(i), []).append(i)

    split_groups_count = 0
    adjusted_150_count = 0
    moved_to_v3_count = 0
    skipped_under_150_count = 0

    for key, idxs in groups.items():
        original_group_total = ali_ht_money(sum(total_val(i) for i in idxs))

        # 수취인 총합이 150불 미만이면 같은 품명이어도 수정 제외
        if original_group_total < 150:
            # 단순 정상건은 로그 과다 방지를 위해 중복 후보가 있을 때만 제외 기록
            tmp = {}
            for i in idxs:
                tmp.setdefault((detail_signature(i), total_val(i)), []).append(i)
            if any(len(v) >= 2 for v in tmp.values()):
                skipped_under_150_count += 1
            continue

        # 분할배송 중복금액: 같은 수취인+전화+품명구성+총금액이 2건 이상이면 각 단가를 건수로 나눔
        dup_map = {}
        for i in idxs:
            sig = detail_signature(i)
            amt = total_val(i)
            if amt > 0 and sig:
                dup_map.setdefault((sig, amt), []).append(i)

        for (_sig, amt), dup_idxs in dup_map.items():
            if len(dup_idxs) < 2:
                continue
            split_groups_count += 1
            n = len(dup_idxs)
            modified = []
            before_lines = []
            after_lines = []
            for i in dup_idxs:
                row_before_total = total_val(i)
                before_units = []
                after_units = []
                for desc_col, qty_col, unit_col in detail_groups:
                    desc = ali_ht_clean_text(df.at[i, desc_col])
                    if not desc:
                        continue
                    old_unit = unit_val(i, unit_col)
                    if old_unit <= 0:
                        continue
                    new_unit = ali_ht_money(old_unit / n)
                    if new_unit != old_unit:
                        excel_set(i, unit_col, new_unit)
                        money_changed_cells.add((i, unit_col))
                        before_units.append(f"{desc}: {old_unit}")
                        after_units.append(f"{desc}: {new_unit}")
                new_total = recalc_row_total(i)
                if new_total != row_before_total:
                    excel_set(i, col_total, new_total)
                    money_changed_cells.add((i, col_total))
                if before_units or new_total != row_before_total:
                    modified.append(i)
                    before_lines.append(f"{row_hawb(i)} 단가[{'; '.join(before_units)}], 총금액 {row_before_total}")
                    after_lines.append(f"{row_hawb(i)} 단가[{'; '.join(after_units)}], 총금액 {new_total}")
            if modified:
                add_log("목록분할", idxs, modified, "상세단가/BA 총금액", " / ".join(before_lines), " / ".join(after_lines), f"동일 수취인·동일 품명·동일 금액 {n}건 분할배송 처리", dup_idxs[0])

        group_total = ali_ht_money(sum(total_val(i) for i in idxs))

        # 합계 150~160불은 149.00 이하로 금액 보정. 단가 큰 상세품목부터, 한 품명 단가 최대 5불 차감.
        if 150 <= group_total <= 160:
            need = ali_ht_money(group_total - 149.00)
            if need > 0:
                candidates = []
                for i in idxs:
                    for desc_col, qty_col, unit_col in detail_groups:
                        desc = ali_ht_clean_text(df.at[i, desc_col])
                        unit = unit_val(i, unit_col)
                        qty = qty_val(i, qty_col)
                        if unit > 0 and qty > 0 and (desc or unit):
                            candidates.append((unit, unit * qty, i, desc_col, qty_col, unit_col, desc))
                candidates.sort(reverse=True, key=lambda x: (x[0], x[1]))
                remaining = need
                modified = []
                before_lines = []
                after_lines = []
                for _unit_sort, _total_sort, i, desc_col, qty_col, unit_col, desc in candidates:
                    if remaining <= 0:
                        break
                    q = qty_val(i, qty_col)
                    old_unit = unit_val(i, unit_col)
                    old_row_total = total_val(i)
                    # 한 품명 단가에서 최대 5불까지만 차감
                    reduce_unit = min(5.00, ali_ht_money(remaining / q), old_unit - 0.01)
                    if reduce_unit <= 0:
                        continue
                    new_unit = ali_ht_money(old_unit - reduce_unit)
                    excel_set(i, unit_col, new_unit)
                    money_changed_cells.add((i, unit_col))
                    new_row_total = recalc_row_total(i)
                    excel_set(i, col_total, new_row_total)
                    money_changed_cells.add((i, col_total))
                    actual_reduce = ali_ht_money(old_row_total - new_row_total)
                    if actual_reduce <= 0:
                        continue
                    modified.append(i)
                    before_lines.append(f"{row_hawb(i)} {desc} 단가 {old_unit}, 총금액 {old_row_total}")
                    after_lines.append(f"{row_hawb(i)} {desc} 단가 {new_unit}, 총금액 {new_row_total}")
                    remaining = ali_ht_money(remaining - actual_reduce)
                if modified:
                    adjusted_150_count += 1
                    add_log("150~160보정", idxs, modified, "상세단가/BA 총금액", " / ".join(before_lines), " / ".join(after_lines), f"수취인별 합계 {group_total}불 → 149불대 보정", modified[0])

        # 보정 후에도 160불 초과면 V=3 배제 처리
        final_total = ali_ht_money(sum(total_val(i) for i in idxs))
        if final_total > 160:
            moved_to_v3_count += 1
            modified = []
            for i in idxs:
                before = df.at[i, col_v]
                if ali_ht_clean_text(before) != "3":
                    excel_set(i, col_v, "3")
                    v_changed_cells.add((i, col_v))
                    hawb = row_hawb(i)
                    excluded_from_list_hawbs.append(hawb)
                    modified.append(i)
            if modified:
                add_log("목록→배제", idxs, modified, "V 용도구분", "1", "3", f"수취인별 합계 {final_total}불로 160불 초과, 배제 처리", modified[0])

    # V=3 배제건 처리: 같은 수취인+전화+품명구성+동일 총금액 분할배송만 1건 유지, 나머지 1~3불 표시용으로 조정
    v3_indices = [i for i in df.index if ali_ht_clean_text(df.at[i, col_v]) == "3"]
    v3_dup_map = {}
    for i in v3_indices:
        amt = total_val(i)
        sig = detail_signature(i)
        if amt > 0 and sig:
            v3_dup_map.setdefault((group_key(i), sig, amt), []).append(i)

    v3_split_adjust_count = 0
    for (_gkey, _sig, _amt), dup_idxs in v3_dup_map.items():
        if len(dup_idxs) < 2:
            continue
        v3_split_adjust_count += 1
        keep = dup_idxs[0]
        modified = []
        before_lines = []
        after_lines = []
        for i in dup_idxs[1:]:
            old_total = total_val(i)
            old_units = []
            new_units = []
            # 첫 번째 상세 단가를 조정해서 총금액 2.00 내외로 표시
            target_total = 2.00
            first_detail = None
            for desc_col, qty_col, unit_col in detail_groups:
                if ali_ht_clean_text(df.at[i, desc_col]) or unit_val(i, unit_col) > 0:
                    first_detail = (desc_col, qty_col, unit_col)
                    break
            if not first_detail:
                continue
            desc_col, qty_col, unit_col = first_detail
            q = qty_val(i, qty_col)
            old_unit = unit_val(i, unit_col)
            new_unit = ali_ht_money(target_total / q) if q else 2.00
            excel_set(i, unit_col, new_unit)
            money_changed_cells.add((i, unit_col))
            old_units.append(f"{ali_ht_clean_text(df.at[i, desc_col])}: {old_unit}")
            new_units.append(f"{ali_ht_clean_text(df.at[i, desc_col])}: {new_unit}")

            # 나머지 상세 단가가 있으면 0으로 조정해서 BA가 2.00에 맞도록 함
            for dcol2, qcol2, ucol2 in detail_groups:
                if ucol2 == unit_col:
                    continue
                if unit_val(i, ucol2) > 0:
                    ou = unit_val(i, ucol2)
                    excel_set(i, ucol2, 0.00)
                    money_changed_cells.add((i, ucol2))
                    old_units.append(f"{ali_ht_clean_text(df.at[i, dcol2])}: {ou}")
                    new_units.append(f"{ali_ht_clean_text(df.at[i, dcol2])}: 0.00")
            new_total = recalc_row_total(i)
            excel_set(i, col_total, new_total)
            money_changed_cells.add((i, col_total))
            modified.append(i)
            before_lines.append(f"{row_hawb(i)} 단가[{'; '.join(old_units)}], 총금액 {old_total}")
            after_lines.append(f"{row_hawb(i)} 단가[{'; '.join(new_units)}], 총금액 {new_total}")
        if modified:
            add_log("배제분할", dup_idxs, modified, "상세단가/BA 총금액", " / ".join(before_lines), " / ".join(after_lines), "V=3 동일 수취인·동일 품명·동일 금액 분할배송: 1건 원금액 유지, 나머지 1~3불 표시", modified[0])

    memo_columns = ["구분", "그룹번호", "수취인", "전화번호", "HAWB NO", "원본행", "처리상태", "변경항목", "변경전", "변경후", "사유"]
    memo_df = pd.DataFrame(logs, columns=memo_columns)
    excluded_df = pd.DataFrame({"목록건에서 배제로 변경된 HAWB NO": excluded_from_list_hawbs})

    summary = {
        "전체 행": len(df),
        "V=1 대상 행": len(v1_indices),
        "150불 미만 수정제외 묶음": skipped_under_150_count,
        "분할배송 금액분할 묶음": split_groups_count,
        "150~160불 금액보정 묶음": adjusted_150_count,
        "160불 초과 V=3 변경 묶음": moved_to_v3_count,
        "배제건 분할표시 묶음": v3_split_adjust_count,
        "품명 변경 셀 수": name_change_count,
        "목록→배제 변경 HAWB 수": len(excluded_from_list_hawbs),
        "전체 변경/확인 로그 수": len(memo_df),
    }

    # 색상 표시: 요청대로 금액 수정 셀 중심 + V/HS/품명/문자형 변경 확인용 표시
    money_fill = PatternFill("solid", fgColor="FFF2CC")   # 금액 수정: 연노랑
    v_fill = PatternFill("solid", fgColor="F4CCCC")       # V 변경: 연빨강
    hs_fill = PatternFill("solid", fgColor="D9EAD3")      # HS 변경: 연초록
    text_fill = PatternFill("solid", fgColor="D9EAF7")    # 문자형 정리: 연파랑
    desc_fill = PatternFill("solid", fgColor="EADCF8")    # 품명 변경: 연보라

    for i, col in money_changed_cells:
        ws.cell(row=i + 2, column=col_pos[col]).fill = money_fill
    for i, col in v_changed_cells:
        ws.cell(row=i + 2, column=col_pos[col]).fill = v_fill
    for i, col in hs_changed_cells:
        ws.cell(row=i + 2, column=col_pos[col]).fill = hs_fill
    for i, col in text_changed_cells:
        ws.cell(row=i + 2, column=col_pos[col]).fill = text_fill
    for i, col in desc_changed_cells:
        ws.cell(row=i + 2, column=col_pos[col]).fill = desc_fill

    # 메모/목록배제송장 시트만 새로 생성. 원본 시트 서식은 건드리지 않음.
    for sheet_name in ["메모", "목록배제송장"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    memo_ws = wb.create_sheet("메모")
    memo_ws.append(memo_columns)
    for row in memo_df.itertuples(index=False):
        memo_ws.append(list(row))
    memo_ws.freeze_panes = "A2"
    for cell in memo_ws[1]:
        cell.font = cell.font.copy(bold=True)
        cell.alignment = cell.alignment.copy(horizontal="center", vertical="center", wrap_text=True)
    for row in memo_ws.iter_rows():
        for cell in row:
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
    widths = {
        "A": 14, "B": 10, "C": 18, "D": 18, "E": 22, "F": 10,
        "G": 12, "H": 22, "I": 55, "J": 55, "K": 55,
    }
    for col_letter, width in widths.items():
        memo_ws.column_dimensions[col_letter].width = width

    if not excluded_df.empty:
        ex_ws = wb.create_sheet("목록배제송장")
        ex_ws.append(["목록건에서 배제로 변경된 HAWB NO"])
        for hawb in excluded_from_list_hawbs:
            ex_ws.append([hawb])
        ex_ws.freeze_panes = "A2"
        ex_ws.column_dimensions["A"].width = 28
        ex_ws["A1"].font = ex_ws["A1"].font.copy(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), summary

def ali_ht_convert_page():
    topbar()
    if st.button("← 전자상거래로 돌아가기", key="ali_ht_back"):
        st.session_state.page = "ecommerce"
        st.rerun()

    st.markdown(
        '<div class="content-card"><div class="page-title">📊 알리 HT변환</div>'
        '<div class="page-sub">원본 기준 금액/V/HS CODE 최종 수정 · 메모 시트 생성</div></div>',
        unsafe_allow_html=True,
    )

    uploaded = st.file_uploader("알리 HT 엑셀 파일 업로드", type=["xlsx", "xls"], key="ali_ht_file")
    st.caption("기준: 원본 전체를 기준으로 분할배송 금액 조정, 150~160불 보정, 160불 초과 V=3 변경, HS CODE 정리 후 메모 시트를 생성합니다.")

    if uploaded:
        if st.button("✅ 알리 HT변환 실행", type="primary", use_container_width=True, key="ali_ht_run"):
            try:
                with st.spinner("알리 HT변환 처리 중입니다..."):
                    result_bytes, summary = ali_ht_process_excel_to_bytes(uploaded)
                st.success("알리 HT변환 완료")
                st.write("처리 요약")
                st.dataframe(pd.DataFrame([summary]), use_container_width=True, hide_index=True)
                st.download_button(
                    "⬇️ 알리HT 결과 다운로드",
                    result_bytes,
                    file_name="알리HT_최종수정본.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="ali_ht_download",
                )
            except Exception as e:
                st.error(f"처리 중 오류가 발생했습니다: {e}")
                st.exception(e)
    else:
        st.info("엑셀 파일을 업로드하면 알리 HT변환을 실행할 수 있습니다.")


if not st.session_state.login:
    login_page()
else:
    if "role" not in st.session_state:
        st.session_state.role = "user"
    if st.session_state.page == "admin":
        admin_page()
    elif st.session_state.page == "bl_convert":
        bl_convert_page()
    elif st.session_state.page == "kyungdong":
        kyungdong_page()
    elif st.session_state.page == "meni_convert":
        meni_convert_page()
    elif st.session_state.page == "ali_ht_convert":
        ali_ht_convert_page()
    elif st.session_state.page == "address_verify":
        address_verify_page()
    elif st.session_state.page == "ecommerce":
        ecommerce_page()
    elif st.session_state.page == "seaair":
        seaair_page()
    elif st.session_state.page == "threepl":
        threepl_page()
    else:
        main_page()
